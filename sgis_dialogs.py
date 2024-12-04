from __future__ import absolute_import
from __future__ import print_function

import base64
import datetime
import distutils
import os.path
import shutil
import sys
import threading
from builtins import range
from functools import partial

import ezdxf
import openpyxl
import processing
import psycopg2
import pyperclip

import os
import pypdf
from pypdf import PdfWriter

from win32com import client

from PyQt5.QtCore import QVariant, Qt, QRectF, QSettings
from PyQt5.QtGui import QIcon, QGuiApplication, QColor
from PyQt5.QtSql import QSqlDatabase
from PyQt5.QtWidgets import *
from ezdxf.addons import Importer
from openpyxl.styles import Alignment
from processing.core.Processing import Processing
from qgis._gui import QgsFileWidget
from qgis.core import *
from qgis.core import QgsProject, QgsMessageLog, QgsDataSourceUri, Qgis, QgsPrintLayout, QgsUnitTypes, QgsLayoutSize, \
    QgsLayoutItemMap, QgsLayoutItemLabel, QgsLayoutItemScaleBar, QgsLayoutExporter

sys.path.append(os.path.dirname(os.path.abspath(__file__)) + "/forms")
sys.path.append(os.path.dirname(os.path.abspath(__file__)) + "/QML")
sys.path.append(os.path.dirname(os.path.abspath(__file__)) + "/360Tools")

from .forms.sgis_label_form import *
from .forms.sgis_supp_pre_form import *
from .forms.sgis_cadOutputs_form import *
from .forms.sgis_potree_form import *

# read config from .json 
import json
jsonpath = os.path.dirname(os.path.abspath(__file__)) + "\\config.json"

with open(jsonpath, "r") as f:
    config = json.load(f)
    
dbServer = config["dbServer"]
dbPort = config["dbPort"]
db = config["db"]
dbUser = config["dbUser"]
dbPass = config["dbPass"]
jobsPath = config["jobsPath"]
estimatesPath = config["estimatesPath"]
reportsPath = config["reportsPath"]
wwwURL = config["wwwURL"]
wwwRootFolder = config["wwwRootFolder"]

QgsMessageLog.logMessage('surveyorGIS (sGIS) plugin | v3.40.1 | STATUS: loaded successfully...', 'sGIS', level=Qgis.Info)
QgsMessageLog.logMessage(db + '@' + dbServer + ':' + dbPort,  'sGIS', level=Qgis.Info)
QgsMessageLog.logMessage('jobsPath | ' + jobsPath, 'sGIS', level=Qgis.Info)
QgsMessageLog.logMessage('welcome.', 'sGIS', level=Qgis.Info)

# common functions
def formatLL(l_l):
    lat_lon = l_l
    ll = len(lat_lon)

    if ll <= 30:
        pass
    else:
        dd = float(lat_lon.split(',')[0])
        dd2 = float(lat_lon.split(',')[1])

        d = int(float(dd))
        m = int(float((dd - d)) * 60)
        s = (dd - d - m / 60) * 3600.00
        z = round(s, 2)

        if d >= 0:
            lat = (str(abs(d)) + '°' + str(abs(m)) + '\'' + str(abs(z)) + '"' + 'N')
        else:
            lat = (str(abs(d)) + '°' + str(abs(m)) + '\'' + str(abs(z)) + '"' + 'S')

        d = int(float(dd2))
        m = int(float((dd2 - d)) * 60)
        s = (dd2 - d - m / 60) * 3600.00
        z = round(s, 2)

        if d >= 0:
            lon = (str(abs(d)) + '°' + str(abs(m)) + '\'' + str(abs(z)) + '"' + 'E')
        else:
            lon = (str(abs(d)) + '°' + str(abs(m)) + '\'' + str(abs(z)) + '"' + 'W')

        lat_lon = str(lat) + '    ' + str(lon)

        return lat_lon

def newParcel(self):
    self.vl = QgsProject.instance().mapLayersByName('Parcels')[0]
    cLayer = self.iface.mapCanvas().currentLayer()

    self.selectLastFeature()
    f = self.vl.selectedFeatures()[0]

    QgsMessageLog.logMessage('SELECTED | ' + str(f.id()), 'sGIS', level=Qgis.Info)

    layer = self.iface.activeLayer()

    feature_dict = {f.id(): f for f in layer.getFeatures()}

    index = QgsSpatialIndex()
    for f in feature_dict.values():
        index.insertFeature(f)

    f = self.vl.selectedFeatures()[0]
    geom = f.geometry()
    intersecting_ids = index.intersects(geom.boundingBox())

    QgsMessageLog.logMessage('intersectingIds | ' + str(intersecting_ids), 'sGIS',
                             level=Qgis.Info)
    neighbors = []
    for intersecting_id in intersecting_ids:

        if intersecting_id < 0:
            pass

        else:
            intersecting_f = feature_dict[intersecting_id]
            QgsMessageLog.logMessage('intersecting_f | ' + str(intersecting_f), 'sGIS',
                                     level=Qgis.Info)
            if len(neighbors) < 1:
                neighbors.append(intersecting_f['objectid'])
                neighbors.append(intersecting_f['town'])
                neighbors.append(intersecting_f['county'])
                neighbors.append(intersecting_f['zipcode'])

                QgsMessageLog.logMessage('neighbors | ' + str(neighbors), 'sGIS',
                                         level=Qgis.Info)
            else:
                pass

    town = neighbors[1]
    county = neighbors[2]
    zipcode = neighbors[3]

    f['town'] = town
    f['county'] = county
    f['zipcode'] = zipcode

    QgsMessageLog.logMessage('attributes : ' + str(town) + ' | ' + str(county) + ' | ' + str(zipcode), 'sGIS',
                             level=Qgis.Info)

    dataProvider = self.vl.dataProvider()

    self.iface.actionToggleEditing().trigger()
    idx = dataProvider.fieldNameIndex('town')
    self.vl.changeAttributeValue(f.id(), idx, town)
    self.vl.updateFields()
    self.iface.activeLayer().commitChanges()

    self.iface.actionToggleEditing().trigger()
    idx2 = dataProvider.fieldNameIndex('county')
    self.vl.changeAttributeValue(f.id(), idx2, county)
    self.vl.updateFields()
    self.iface.activeLayer().commitChanges()

    self.iface.actionToggleEditing().trigger()
    idx3 = dataProvider.fieldNameIndex('zipcode')
    self.vl.changeAttributeValue(f.id(), idx3, zipcode)
    self.vl.updateFields()
    self.iface.activeLayer().commitChanges()

    self.iface.actionToggleEditing().trigger()
    idx2 = dataProvider.fieldNameIndex('objectid')
    self.vl.changeAttributeValue(f.id(), idx2, str(f.id()))
    self.vl.updateFields()
    self.iface.activeLayer().commitChanges()

    self.selectLastFeature()
    self.iface.actionIdentify().trigger()
    QgsMessageLog.logMessage('PARCEL UPDATE | DONE.', 'sGIS', level=Qgis.Info)
    QMessageBox.critical(self.iface.mainWindow(), "DONE!",
                         "PARCEL has been created | " + str(f['map_bk_lot']))

def resetLegend(self):
    root = QgsProject.instance().layerTreeRoot()
    for child in root.children():
        child.setExpanded(False)
    aGroup = root.findGroup('Surveyor')
    aGroup.setExpanded(True)
    aGroup = root.findGroup('Misc')
    aGroup.setExpanded(False)
    aGroup = root.findGroup('FEMA')
    aGroup.setExpanded(False)
    aGroup = root.findGroup("2' SHP files")
    aGroup.setExpanded(False)
    aGroup = root.findGroup("Topo")
    aGroup.setExpanded(False)
    aGroup = root.findGroup("always on")
    aGroup.setExpanded(False)

def getSQL(self):
    base64_pw = dbPass
    base64_bytes = base64_pw.encode('ascii')
    pw_bytes = base64.b64decode(base64_bytes)
    pw = pw_bytes.decode('ascii')
    pw = str(pw)
    return pw
    # QgsMessageLog.logMessage('user: postgres | pass: ' + str(pw), 'sGIS', level=Qgis.Info)

def jobFolders(jobNo):
    import datetime
    year = datetime.datetime.today().strftime('%Y')
    jobYear = '20' + jobNo[:2]
    if jobYear == year:
        year = year
    else:
        year = jobYear
    path = os.path.join(jobsPath, jobYear, jobNo)
    jipath = os.path.join(path, "Job_Info")
    dwgpath = os.path.join(path, "dwg")
    frcpath = os.path.join(path, "From_Others")
    gispath = os.path.join(path, "GIS")
    ppath = os.path.join(path, "prints")
    supath = os.path.join(path, "survey")
    panopath = os.path.join(supath, "360Pano")
    rspath = os.path.join(path, "research")

    QgsMessageLog.logMessage('Checking output folder: ' + path + '...', 'sGIS', level=Qgis.Info)
    if not os.path.exists(path):
        os.makedirs(path)
        os.makedirs(jipath)
        os.makedirs(dwgpath)
        os.makedirs(gispath)
        os.makedirs(frcpath)
        os.makedirs(ppath)
        os.makedirs(supath)
        os.makedirs(panopath)
        os.makedirs(rspath)

    if not os.path.exists(jipath):
        os.makedirs(jipath)
    if not os.path.exists(dwgpath):
        os.makedirs(dwgpath)
    if not os.path.exists(gispath):
        os.makedirs(gispath)
    if not os.path.exists(frcpath):
        os.makedirs(frcpath)
    if not os.path.exists(ppath):
        os.makedirs(ppath)
    if not os.path.exists(supath):
        os.makedirs(supath)
    if not os.path.exists(rspath):
        os.makedirs(rspath)
    if not os.path.exists(panopath):
        os.makedirs(panopath)

    cmdpath = os.path.dirname(os.path.realpath(__file__)) + '\\360Tools\\coords.cmd'
    shutil.copy(cmdpath, panopath)

# classes
class sgis_newPlan(object):
    newJob = 0
    selComp = 0
    multiFeat = 0
    count = 0

    def __init__(self, iface):
        # save reference to the QGIS interface
        self.iface = iface

    def initGui(self):

        self.action = QAction("Create New Plan", self.iface.mainWindow())
        self.action.triggered.connect(self.run)
        self.action.trigger()
        self.iface.mapCanvas().selectionChanged.connect(self.select_changed)

    def run(self):
        self.vl = QgsProject.instance().mapLayersByName('Parcels')[0]
        self.iface.setActiveLayer(self.vl)

        reply = QMessageBox.question(self.iface.mainWindow(), 'New Plan',
                                     'Click OK and select the correct parcel(s) for the new plan.',
                                     QMessageBox.Ok, QMessageBox.Cancel)
        if reply == QMessageBox.Ok:
            if self.selComp == 1:
                return
            else:
                self.newJob = 1
                if self.count == 0:
                    for a in self.iface.mainWindow().children():
                        if a.objectName() == 'mActionDeselectAll':
                            a.trigger()
                            QgsMessageLog.logMessage('Job creation starting...', 'sGIS', level=Qgis.Info)
                            self.iface.actionSelectFreehand().trigger()
                            self.count = self.count + 1
                else:
                    self.iface.actionSelect().trigger()

        else:
            QgsMessageLog.logMessage('DEBUG: Plan creation cancelled.', 'sGIS', level=Qgis.Info)

    def select_changed(self):

        if self.newJob == 0:
            return
        else:
            vLayer = self.iface.activeLayer()
            feats_count = vLayer.selectedFeatureCount()

            msg = QMessageBox()
            msg.setWindowTitle('Selection')
            msg.setText(str(feats_count) + ' features have been selected. Continue?')
            create = msg.addButton('Continue', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(create)
            msg.exec_()
            msg.deleteLater()

            QGuiApplication.restoreOverrideCursor()

            if msg.clickedButton() is create:

                    QgsMessageLog.logMessage('Plan creation will begin...', 'sGIS', level=Qgis.Info)
                    self.iface.actionCopyFeatures().trigger()
                    self.newJob = 0
                    self.vl = QgsProject.instance().mapLayersByName('plans')[0]
                    self.iface.setActiveLayer(self.vl)

                    self.iface.actionIdentify().trigger()
                    self.iface.actionToggleEditing().trigger()
                    self.iface.actionPasteFeatures().trigger()

                    self.iface.mainWindow().findChild(QAction, 'mActionMergeFeatures').trigger()
                    # end merge features, launch form

                    QgsMessageLog.logMessage('Launching form for merged feature...', 'sGIS', level=Qgis.Info)
                    self.iface.activeLayer().commitChanges()

                    self.vl = QgsProject.instance().mapLayersByName('plans')[0]

                    feats_count = self.vl.selectedFeatureCount()

                    if feats_count == 0:
                        self.selectLastFeature()
                        jobNo = self.vl.selectedFeatures()[0]
                    else:
                        jobNo = self.vl.selectedFeatures()[0]

                    try:
                        self.iface.actionToggleEditing().trigger()
                        f = self.vl.selectedFeatures()[0]
                        QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
                        self.iface.openFeatureForm(self.iface.activeLayer(), f, False, True)
                        QGuiApplication.restoreOverrideCursor()
                    except Exception as e:
                        pass

                    QgsMessageLog.logMessage('Saving changes...', 'sGIS', level=Qgis.Info)

                    #result = self.active_edit()
                    #QgsMessageLog.logMessage('RESULT: ' + str(result), 'sGIS', level=Qgis.Info)

                    self.iface.activeLayer().commitChanges()

                    lat_lon = jobNo['lat_lon']
                    dd = float(lat_lon.split(',')[0])
                    dd2 = float(lat_lon.split(',')[1])

                    d = int(float(dd))
                    m = int(float((dd - d)) * 60)
                    s = (dd - d - m / 60) * 3600.00
                    z = round(s, 2)

                    if d >= 0:
                        lat = (str(abs(d)) + '°' + str(abs(m)) + "'" + str(abs(z)) + '"' + 'N')
                    else:
                        lat = (str(abs(d)) + '°' + str(abs(m)) + "'" + str(abs(z)) + '"' + 'S')

                    d = int(float(dd2))
                    m = int(float((dd2 - d)) * 60)
                    s = (dd2 - d - m / 60) * 3600.00
                    z = round(s, 2)

                    if d >= 0:
                        lon = (str(abs(d)) + '°' + str(abs(m)) + "'" + str(abs(z)) + '"' + 'E')
                    else:
                        lon = (str(abs(d)) + '°' + str(abs(m)) + "'" + str(abs(z)) + '"' + 'W')

                    lat_lon = str(lat) + '    ' + str(lon)
                    lyr = QgsProject.instance().mapLayersByName('plans')[0]
                    # start editing, change field value
                    self.iface.actionToggleEditing().trigger()
                    layerData = lyr.dataProvider()
                    idx3 = layerData.fieldNameIndex('lat_lon')
                    lyr.changeAttributeValue(jobNo.id(), idx3, lat_lon)
                    lyr.updateFields()
                    # end field update, save layer
                    self.iface.activeLayer().commitChanges()
                    jobNo = self.vl.selectedFeatures()[0]
                    QgsMessageLog.logMessage('Plan: ' + str(jobNo) + ' has been created and saved.',
                                             'sGIS', level=Qgis.Info)
                    lyr = QgsProject.instance().mapLayersByName('plans')[0]
                    self.iface.setActiveLayer(lyr)
                    self.iface.messageBar().clearWidgets()

            elif msg.clickedButton() is cancel:
                self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                # QGuiApplication.restoreOverrideCursor()
                return

    def active_edit(self):

        #try:

            self.vl = QgsProject.instance().mapLayersByName('plans')[0]
            self.iface.setActiveLayer(self.vl)
            self.iface.activeLayer().commitChanges()
            self.iface.actionToggleEditing().trigger()

            f = self.vl.selectedFeatures()[0]

            QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
            self.iface.openFeatureForm(self.iface.activeLayer(), f, False, True)
            QGuiApplication.restoreOverrideCursor()

            QgsMessageLog.logMessage('Saving changes...', 'sGIS', level=Qgis.Info)

        #except Exception as e:
        #    exc_type, exc_obj, exc_tb = sys.exc_info()
        #    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        #    QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
        #                         "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
        #                             exc_tb.tb_lineno) + ' ' + str(e))
        #    return

            self.vl = QgsProject.instance().mapLayersByName('plans')[0]
            self.iface.setActiveLayer(self.vl)
            self.iface.activeLayer().commitChanges()

    def selectLastFeature(self):

        f2 = self.vl.getFeatures()
        fCount = self.vl.featureCount()

        fIds = []
        fIds = [f["gid"] for f in f2]
        fIds.sort()

        fId = [fIds[-1]]
        self.vl.selectByIds(fId)


class sgis_newFreehandPlan(object):

    def __init__(self, iface):
        # save reference to the QGIS interface
        self.iface = iface

    def initGui(self):

        self.action = QAction("Create New Freehand Plan", self.iface.mainWindow())
        self.action.triggered.connect(self.run)
        self.action.trigger()
        # self.iface.mapCanvas().selectionChanged.connect(self.select_changed)

    def run(self):
        self.vl = QgsProject.instance().mapLayersByName('plans')[0]
        self.iface.setActiveLayer(self.vl)

        reply = QMessageBox.question(self.iface.mainWindow(), 'New Freehand Plan',
                                     'Click OK and draw the freehand polygon for the new plan.',
                                     QMessageBox.Ok, QMessageBox.Cancel)
        if reply == QMessageBox.Ok:
            for a in self.iface.mainWindow().children():
                if a.objectName() == 'mActionDeselectAll':
                    a.trigger()
                    QgsMessageLog.logMessage('FIRST RUN: Previously selected parcel(s) have been cleared.',
                                             'sGIS', level=Qgis.Info)
                    QgsMessageLog.logMessage('Plan creation starting...', 'sGIS', level=Qgis.Info)
                    for a in self.iface.mainWindow().children():
                        if a.objectName() == 'mActionAddFeature':
                            a.trigger()
                    self.iface.actionToggleEditing().trigger()
                    self.iface.activeLayer().featureAdded.connect(self.feature_added)

        else:
            QgsMessageLog.logMessage('DEBUG: Plan creation cancelled.', 'sGIS', level=Qgis.Info)

    def feature_added(self, featureAdded):
        layer = self.iface.activeLayer()
        layer.featureAdded.disconnect()
        layer.select(featureAdded)
        # layer.commitChanges()
        # self.selectLastFeature(layer)
        QgsMessageLog.logMessage('SELECTED: ' + str(featureAdded), 'sGIS', level=Qgis.Info)
        self.select_changed()

    def select_changed(self):

        self.vl = QgsProject.instance().mapLayersByName('plans')[0]
        self.iface.setActiveLayer(self.vl)
        self.iface.activeLayer().commitChanges()
        self.iface.actionToggleEditing().trigger()

        f = self.vl.selectedFeatures()[0]

        QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
        self.iface.openFeatureForm(self.iface.activeLayer(), f, False, True)
        QGuiApplication.restoreOverrideCursor()

        QgsMessageLog.logMessage('Saving changes...', 'sGIS', level=Qgis.Info)

        self.vl = QgsProject.instance().mapLayersByName('plans')[0]
        self.iface.setActiveLayer(self.vl)
        self.iface.activeLayer().commitChanges()

    def selectLastFeature(self):

        f2 = self.vl.getFeatures()
        fCount = self.vl.featureCount()

        fIds = []
        fIds = [f["gid"] for f in f2]
        fIds.sort()

        fId = [fIds[-1]]
        self.vl.selectByIds(fId)


class sgis_newJob_ORIGINAL(object):
    newJob = 0
    selComp = 0
    multiFeat = 0
    count = 0

    def __init__(self, iface):
        # save reference to the QGIS interface
        self.iface = iface

    def initGui(self):

        self.action = QAction("Create New Job", self.iface.mainWindow())
        self.action.triggered.connect(self.run)
        self.action.trigger()
        self.iface.mapCanvas().selectionChanged.connect(self.select_changed)

    def run(self):

        self.vl = QgsProject.instance().mapLayersByName('Parcels')[0]
        self.iface.setActiveLayer(self.vl)

        reply = QMessageBox.question(self.iface.mainWindow(), 'New Job',
                                     'Click OK and select the correct parcel for the new job.',
                                     QMessageBox.Ok, QMessageBox.Cancel)
        if reply == QMessageBox.Ok:
            if self.selComp == 1:
                return
            else:
                self.newJob = 1
                if self.count == 0:
                    # for a in self.iface.attributesToolBar().actions():
                    for a in self.iface.mainWindow().children():
                        if a.objectName() == 'mActionDeselectAll':
                            a.trigger()
                            QgsMessageLog.logMessage('FIRST RUN: Previously selected parcel(s) have been cleared.',
                                                     'sGIS', level=Qgis.Info)
                            QgsMessageLog.logMessage('Job creation starting...', 'sGIS', level=Qgis.Info)
                            self.iface.actionSelect().trigger()
                            self.count = self.count + 1
                else:
                    self.iface.actionSelect().trigger()

        else:
            QgsMessageLog.logMessage('DEBUG: Job creation cancelled.', 'sGIS', level=Qgis.Info)

    def select_changed(self):

        if self.newJob == 0:
            return
        else:
            try:
                parcel = self.iface.activeLayer().selectedFeatures()[0]
                map_bk_lot = parcel["map_bk_lot"]
                origMBL = parcel["map_bk_lot"]

                try:
                    proplocnum = str(int(parcel["proplocnum"]))
                    proploc = str(parcel["prop_loc"])

                    if proplocnum == '0':
                        proplocnum = ''

                    address = proplocnum + ' ' + proploc

                except Exception:
                    address = ''
                    pass

                mbl = map_bk_lot.split('-')
                mbLen = len(mbl)

                try:
                    if mbLen == 1:
                        map_bk_lot = map_bk_lot
                    elif mbLen == 2:
                        map_bk_lot = 'Map ' + mbl[0].lstrip('0') + ', Lot ' + mbl[1].lstrip('0')
                    elif mbLen == 3:
                        map_bk_lot = 'Map ' + mbl[0].lstrip('0') + ', Lot ' + mbl[1].lstrip('0') + '-' + mbl[2].lstrip(
                            '0')
                except Exception as e:
                    map_bk_lot = map_bk_lot

                msg = QMessageBox()
                msg.setWindowTitle('New Job')
                msg.setText(map_bk_lot + ' has been selected. Continue?')
                create = msg.addButton('Create Job', QMessageBox.AcceptRole)
                add = msg.addButton(' Select Additional Parcel(s) ', QMessageBox.AcceptRole)
                cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
                msg.setDefaultButton(create)
                QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
                msg.exec_()
                msg.deleteLater()
                QGuiApplication.restoreOverrideCursor()

                if msg.clickedButton() is create:
                    self.newJob = 0
                    if self.multiFeat == 0:
                        QgsMessageLog.logMessage('Job creation will begin for: ' + map_bk_lot,
                                                 'sGIS', level=Qgis.Info)

                        self.iface.actionCopyFeatures().trigger()
                        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
                        self.iface.setActiveLayer(self.vl)
                        self.iface.actionToggleEditing().trigger()
                        self.iface.actionPasteFeatures().trigger()
                        self.iface.activeLayer().commitChanges()

                    else:
                        QgsMessageLog.logMessage('Job creation will begin for multiple Parcels:',
                                                 'sGIS', level=Qgis.Info)

                        self.iface.actionCopyFeatures().trigger()
                        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
                        self.iface.setActiveLayer(self.vl)
                        self.iface.actionToggleEditing().trigger()
                        self.iface.actionPasteFeatures().trigger()
                        self.iface.mainWindow().findChild(QAction, 'mActionMergeFeatures').trigger()
                        self.iface.activeLayer().commitChanges()

                    self.iface.actionToggleEditing().trigger()

                    layerData = self.vl.dataProvider()
                    idx = layerData.fieldNameIndex('sid')
                    sid = self.vl.maximumValue(idx)
                    it = self.vl.getFeatures(QgsFeatureRequest().setFilterExpression(u'"sid" = {0}'.format(sid)))

                    for feature in it:
                        f = feature.id()
                        self.vl.select(f)

                    jobNo = self.vl.selectedFeatures()[0]

                    job_num = jobNo["job_no"]
                    lat_lon0 = jobNo['lat_lon']
                    lat_lon = formatLL(lat_lon0)

                    idx3 = layerData.fieldNameIndex('lat_lon')
                    self.vl.changeAttributeValue(jobNo.id(), idx3, lat_lon)
                    self.vl.updateFields()
                    self.iface.activeLayer().commitChanges()

                    self.iface.actionToggleEditing().trigger()
                    idx4 = layerData.fieldNameIndex('locus_addr')
                    self.vl.changeAttributeValue(jobNo.id(), idx4, address)
                    self.vl.updateFields()
                    self.iface.activeLayer().commitChanges()

                    QgsMessageLog.logMessage('Launching form for editing...', 'sGIS', level=Qgis.Info)

                    self.iface.actionIdentify().trigger()
                    self.iface.actionToggleEditing().trigger()
                    self.vl = QgsProject.instance().mapLayersByName('contacts')[0]
                    self.iface.setActiveLayer(self.vl)
                    self.iface.actionToggleEditing().trigger()

                    result = self.active_edit()

                    if result:
                        self.iface.activeLayer().commitChanges()
                    else:
                        self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                        self.iface.actionRollbackAllEdits().trigger()
                        self.vl = QgsProject.instance().mapLayersByName('contacts')[0]
                        self.iface.setActiveLayer(self.vl)
                        self.iface.actionToggleEditing().trigger()
                        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
                        self.iface.setActiveLayer(self.vl)

                        for f in self.vl.selectedFeatures():
                            self.vl.deleteFeature(f.id())

                        self.iface.activeLayer().commitChanges()
                        return

                    self.vl = QgsProject.instance().mapLayersByName('contacts')[0]
                    self.iface.setActiveLayer(self.vl)
                    self.iface.activeLayer().commitChanges()

                    self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
                    self.iface.setActiveLayer(self.vl)
                    self.iface.messageBar().clearWidgets()
                    resetLegend(self)

                    QgsMessageLog.logMessage('JobNo:' + str(job_num) + ' has been created and saved.',
                                             'sGIS', level=Qgis.Info)

                elif msg.clickedButton() is add:
                    self.multiFeat = 1
                elif msg.clickedButton() is cancel:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                    return

            except Exception as e:
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                     "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                         exc_tb.tb_lineno) + ' ' + str(e))
                return

    def active_edit(self):

        try:

            self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
            self.iface.setActiveLayer(self.vl)
            f = self.vl.selectedFeatures()[0]

            QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
            result = self.iface.openFeatureForm(self.iface.activeLayer(), f, False, True)
            if result:
                pass
            else:
                QGuiApplication.restoreOverrideCursor()
                QgsMessageLog.logMessage('DEBUG: active_edit cancelled.', 'sGIS', level=Qgis.Info)
                return 0

            self.abutters_dialog = sgis_abutters(self.iface)
            self.abutters_dialog.run()
            QGuiApplication.restoreOverrideCursor()
            QgsMessageLog.logMessage('Saving changes...', 'sGIS', level=Qgis.Info)
            resetLegend(self)

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))
            return

        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(self.vl)
        self.iface.activeLayer().commitChanges()
        resetLegend(self)

        return 1


class sgis_newLPJob(object):

    def __init__(self, iface, type, type2):
        # save reference to the QGIS interface
        self.iface = iface
        self.type = type
        self.type2 = type2

    def initGui(self, type, type2):

        self.run(type, type2)

    def run(self, sType, pType):

        if sType:
            if sType == 'X':
                self.supp_type = 'X'
                QgsMessageLog.logMessage('STANDARD JOB.', 'sGIS', level=Qgis.Info)

            else:
                QgsMessageLog.logMessage('SUPPLEMENTAL.', 'sGIS', level=Qgis.Info)
                self.supp_type = sType
        else:
            self.supp_type = 'X'

        if pType == 'P':
            QgsMessageLog.logMessage('PARCEL: ' + sType, 'sGIS', level=Qgis.Info)
            self.supp_type = sType
        elif pType == 'LP':
            QgsMessageLog.logMessage('LINE/FREEHAND: ' + sType, 'sGIS', level=Qgis.Info)
            self.supp_type = sType
        else:
            QgsMessageLog.logMessage('STANDARD JOB.', 'sGIS', level=Qgis.Info)
            self.supp_type = 'X'

        if pType == 'LP':

            self.objectType = ''
            msg = QMessageBox()
            msg.setWindowTitle('New Line/Free Polygonal Job')
            msg.setText('Which type of new job would you like to create?')
            line = msg.addButton('Line/Polyline', QMessageBox.AcceptRole)
            poly = msg.addButton('Freehand Polygon', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(line)
            QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
            msg.exec_()
            msg.deleteLater()
            QGuiApplication.restoreOverrideCursor()

            if msg.clickedButton() is line:
                self.line = 1
                self.poly = 0
                msg = QMessageBox()
                msg.setWindowTitle('New Line/Free Polygonal Job')
                msg.setText('Should we copy an existing feature or draw a new one?')
                existing = msg.addButton('Select Existing', QMessageBox.AcceptRole)
                new = msg.addButton('Draw New', QMessageBox.AcceptRole)
                cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
                msg.setDefaultButton(cancel)
                QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
                msg.exec_()
                msg.deleteLater()
                QGuiApplication.restoreOverrideCursor()

                if msg.clickedButton() is existing:
                    self.type = 'existing'
                    msg = QMessageBox()
                    msg.setWindowTitle('Select Existing Line/Polyline(s)')
                    result = self.setLineSourceLayers(msg, self.type)
                    # source layer has now been chosen
                    #QgsMessageLog.logMessage('RESULT: ' + str(result), 'sGIS', level=Qgis.Info)
                    if result:

                        reply = QMessageBox.question(self.iface.mainWindow(), 'Select Existing Line/Polyline',
                                                     'Click OK and select the correct line/polyline for the new job.',
                                                     QMessageBox.Ok, QMessageBox.Cancel)
                        if reply == QMessageBox.Ok:
                           for a in self.iface.mainWindow().children():
                                if a.objectName() == 'mActionDeselectAll':
                                    a.trigger()
                                    self.iface.actionSelectFreehand().trigger()
                                    self.iface.mapCanvas().selectionChanged.connect(self.feature_selected)
                        else:
                            QgsMessageLog.logMessage('CANCELLED: Job creation cancelled before source selection.',
                                                     'sGIS', level=Qgis.Info)
                            return
                    else:
                        QgsMessageLog.logMessage('NORESULT: Job creation cancelled before source selection.', 'sGIS',
                                             level=Qgis.Info)
                    return

                elif msg.clickedButton() is new:
                    self.type = 'new'

                    msg = QMessageBox()
                    msg.setWindowTitle('Draw New Line/Polyline')
                    result = self.setLineSourceLayers(msg, self.type)

                    if result:

                        reply = QMessageBox.question(self.iface.mainWindow(), 'Draw New Line/Polyline',
                                                     'Click OK and draw the new line/polyline.',
                                                     QMessageBox.Ok, QMessageBox.Cancel)

                        if reply == QMessageBox.Ok:
                            for a in self.iface.mainWindow().children():
                                if a.objectName() == 'mActionDeselectAll':
                                    a.trigger()
                                    for a in self.iface.mainWindow().children():
                                        if a.objectName() == 'mActionAddFeature':
                                            a.trigger()
                                    self.iface.actionToggleEditing().trigger()
                                    self.iface.activeLayer().featureAdded.connect(self.feature_added)

                        else:
                            QgsMessageLog.logMessage('DEBUG: Job creation cancelled before source selection.',
                                                 'sGIS', level=Qgis.Info)
                            return

                    else:
                        try:
                            self.iface.mapCanvas().selectionChanged.disconnect(self.feature_selected)
                            return
                        except Exception:
                            return

                elif msg.clickedButton() is cancel:
                    try:
                        self.iface.mapCanvas().selectionChanged.disconnect(self.feature_selected)
                        return
                    except Exception:
                        return

            elif msg.clickedButton() is poly:
                    self.poly = 1
                    self.type = 'new'

                    reply = QMessageBox.question(self.iface.mainWindow(), 'Draw New Polygon',
                                                 'Click OK and draw the new free polygonal feature.',
                                                 QMessageBox.Ok, QMessageBox.Cancel)

                    if self.supp_type == 'X':
                        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
                        self.iface.setActiveLayer(self.vl)
                    else:
                        self.vl = QgsProject.instance().mapLayersByName('supplementals')[0]
                        self.iface.setActiveLayer(self.vl)

                    if reply == QMessageBox.Ok:
                        for a in self.iface.mainWindow().children():
                            if a.objectName() == 'mActionDeselectAll':
                                a.trigger()
                                self.iface.mainWindow().findChild(QAction, 'mActionAddFeature').trigger()
                                QgsMessageLog.logMessage('READY TO DRAW.', 'sGIS',
                                                         level=Qgis.Info)
                                self.iface.actionToggleEditing().trigger()
                                self.iface.activeLayer().featureAdded.connect(self.feature_added)
                            else:
                                pass
                    else:
                        QgsMessageLog.logMessage('DEBUG: Job creation cancelled before source selection.', 'sGIS', level=Qgis.Info)

            elif msg.clickedButton() is cancel:
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.feature_selected)
                except Exception:
                    pass

        elif pType == 'P':
            self.poly = 2
            self.vl = QgsProject.instance().mapLayersByName('Parcels')[0]
            self.iface.setActiveLayer(self.vl)

            reply = QMessageBox.question(self.iface.mainWindow(), 'New Supplemental (Parcel-based)',
                                         'Click OK and select the correct parcel(s) for the new plan.',
                                         QMessageBox.Ok, QMessageBox.Cancel)
            if reply == QMessageBox.Ok:
                for a in self.iface.mainWindow().children():
                    if a.objectName() == 'mActionDeselectAll':
                        a.trigger()
                        QgsMessageLog.logMessage('Supplemental (Parcel-based) creation starting...', 'sGIS', level=Qgis.Info)
                        self.iface.actionSelectFreehand().trigger()
                        self.iface.mapCanvas().selectionChanged.connect(self.feature_selected)
            else:
                QgsMessageLog.logMessage('DEBUG: Supplemental creation cancelled.', 'sGIS', level=Qgis.Info)

        resetLegend(self)
        self.iface.messageBar().clearWidgets()

    def feature_added(self, featureAdded):
        layer = self.iface.activeLayer()
        layer.featureAdded.disconnect()
        layer.select(featureAdded)
        # layer.commitChanges()
        # self.selectLastFeature(layer)
        QgsMessageLog.logMessage('SELECTED: ' + str(featureAdded), 'sGIS', level=Qgis.Info)
        self.select_changed('new', self.poly, self.supp_type)

    def feature_selected(self):
        self.iface.mapCanvas().selectionChanged.disconnect()
        layer = self.iface.activeLayer()
        # QgsMessageLog.logMessage('feature_selected: ' + str(self.type) + ' | ' + str(self.objectType) + ' | ' + str(self.poly), 'sGIS', level=Qgis.Info)
        self.select_changed('existing', self.poly, self.supp_type)

    def select_changed(self, type, poly, supp):

        try:
            self.iface.mapCanvas().selectionChanged.disconnect()
        except Exception as e:
            # QgsMessageLog.logMessage('NOT connected.' + str(supp), 'sGIS', level=Qgis.Info)
            pass

        activeLayer = self.iface.activeLayer()
        aLayer = activeLayer.name()

        if supp == "Plan by others (K)":
            self.supp_type = 'K'
        elif supp == "Document (D)":
            self.supp_type = 'D'
        elif supp == "Raster (R)":
            self.supp_type = 'R'
        elif supp == "Proposed / Design (P)":
            self.supp_type = 'P'
        elif supp == "Historical Map / 1857 County (H)":
            self.supp_type = 'H'
        elif supp == "Tax Maps (T)":
            self.supp_type = 'T'
        elif supp == "Map (M)":
            self.supp_type = 'M'
        elif supp == "Other (O)":
            self.supp_type = 'O'
        elif supp == "USGS QUADS (Q)":
            self.supp_type = 'Q'
        elif supp == "Folder (F)":
            self.supp_type = 'F'
        elif supp == "Bathymetric (B)":
            self.supp_type = 'B'
        else:
            self.supp_type = 'X'

        # QgsMessageLog.logMessage('SUPP.: ' + str(supp), 'sGIS', level=Qgis.Info)
        # QgsMessageLog.logMessage('SUPP. TYPE: ' + str(type) + ' | ' + str(self.supp_type), 'sGIS', level=Qgis.Info)
        try:
            source = self.iface.activeLayer().selectedFeatures()[0]

            if aLayer == 'new_roads':
                streetname = 'New Road Feature'
                self.street = str(streetname)
            elif aLayer == 'ng911rdss':
                number = source["SOURCE"]
                streetname = source["RDNAME"]
                self.street = str(str(number) + ' ' + str(streetname))
                town = source["TOWN"]
                county = source["LCOUNTY"]
                map_bk_lot = source["MAP_LABEL"]
                zip = source["LZIPCODE"]
            elif aLayer == 'Parcels':
                self.street = 'New Polygon Feature'
                town = source["town"]
                county = source["county"]
                map_bk_lot = source["map_bk_lot"]
                zip = source["zipcode"]
            elif aLayer == 'jobs':
                self.street = source['job_no']
            elif aLayer == 'supplementals':
                self.street = 'New Polygon Feature'
            else:
                self.street = 'New Line Feature'

            if type == 'new':
                layerData = self.vl.dataProvider()
                idx = layerData.fieldNameIndex('supp_type')
                self.vl.changeAttributeValue(source.id(), idx, self.supp_type)
                self.vl.updateFields()
                self.iface.activeLayer().commitChanges()

                try:
                    address = self.street
                    if poly == 2:
                        self.objectType = 'polygon'
                    elif poly == 1:
                        self.objectType = 'poly'
                    else:
                        self.objectType = 'line'

                    self.genJobFromLine()
                    try:
                        self.updateJobGL()
                    except Exception as e:
                        pass

                    self.launchForm()

                    result = self.active_edit()

                    QgsMessageLog.logMessage('Finalizing...', 'sGIS', level=Qgis.Info)
                    new_job = self.vl.selectedFeatures()[0]
                    job_num = str(new_job["job_no"])

                    if result == 1:

                        if self.supp_type == 'X':
                            self.iface.activeLayer().commitChanges()
                            self.vl = QgsProject.instance().mapLayersByName('contacts')[0]
                            self.iface.setActiveLayer(self.vl)
                            self.iface.activeLayer().commitChanges()

                            self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
                            self.iface.setActiveLayer(self.vl)
                            self.iface.activeLayer().commitChanges()
                            resetLegend(self)

                            QgsMessageLog.logMessage('JobNo:' + str(job_num) + ' has been created and saved.',
                                                     'sGIS', level=Qgis.Info)
                        else:
                            QgsMessageLog.logMessage('SuppNo: ' + str(job_num) + ' has been created and saved.',
                                                     'sGIS', level=Qgis.Info)
                            resetLegend(self)

                    # ROLLBACK EVERYTHING
                    else:
                        QgsMessageLog.logMessage('OOPS...', 'sGIS', level=Qgis.Info)
                        self.iface.actionRollbackAllEdits().trigger()
                        self.vl = QgsProject.instance().mapLayersByName('contacts')[0]
                        self.iface.setActiveLayer(self.vl)
                        self.iface.actionToggleEditing().trigger()
                        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
                        self.iface.setActiveLayer(self.vl)

                        for f in self.vl.selectedFeatures():
                            self.vl.deleteFeature(f.id())

                        self.iface.activeLayer().commitChanges()

                        if self.objectType == 'road':
                            self.vl = QgsProject.instance().mapLayersByName('new_roads')[0]
                            self.iface.setActiveLayer(self.vl)

                        elif self.objectType == 'utility':
                            self.vl = QgsProject.instance().mapLayersByName('utilities')[0]
                            self.iface.setActiveLayer(self.vl)

                        elif self.objectType == 'easement':
                            self.vl = QgsProject.instance().mapLayersByName('easements')[0]
                            self.iface.setActiveLayer(self.vl)

                        try:
                            for f in self.vl.selectedFeatures():
                                self.iface.actionToggleEditing().trigger()
                                self.vl.deleteFeature(f.id())
                                self.iface.activeLayer().commitChanges()
                            self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
                            self.iface.setActiveLayer(self.vl)

                        except Exception as e:
                            exc_type, exc_obj, exc_tb = sys.exc_info()
                            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                            QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                                     exc_tb.tb_lineno) + ' ' + str(e))
                        pass

                    self.resetLegend()

                except Exception as e:
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                         "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                             exc_tb.tb_lineno) + ' ' + str(e))
                    return

            # type == 'existing':
            else:
                address = str(self.street)

                if self.poly == 2:
                    self.objectType = 'polygon'
                elif self.poly == 1:
                    self.objectType = 'poly'
                else:
                    self.objectType = 'line'

                QgsMessageLog.logMessage('Creation will begin for: ' + address, 'sGIS', level=Qgis.Info)
                try:
                    canvas = self.iface.mapCanvas()
                    canvas.zoomToSelected(activeLayer)

                    ex = canvas.extent()
                    ex.scale(2.0)
                    canvas.setExtent(ex)
                    canvas.refresh()

                    # <EXISTING.PARCEL.JOB>
                    if self.objectType == 'polygon':
                        try:
                            vLayer = self.iface.activeLayer()
                            feats_count = vLayer.selectedFeatureCount()
                            f = vLayer.selectedFeatures()[0]
                            msg = QMessageBox()
                            msg.setWindowTitle('Selection')
                            msg.setText(str(feats_count) + ' features have been selected. Continue?')
                            create = msg.addButton('Continue', QMessageBox.AcceptRole)
                            again = msg.addButton('Select Again', QMessageBox.AcceptRole)
                            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
                            msg.setDefaultButton(create)
                            msg.exec_()
                            msg.deleteLater()
                            QGuiApplication.restoreOverrideCursor()

                            if msg.clickedButton() is create:
                                self.genJobFromLine()
                                self.updateJobGL()
                                self.launchForm()
                                result = self.active_edit()

                                QgsMessageLog.logMessage('Finalizing...', 'sGIS', level=Qgis.Info)

                                self.selectLastFeature(self.iface.activeLayer())
                                new_job = self.vl.selectedFeatures()[0]
                                job_num = str(new_job["job_no"])

                                if result == 1:

                                    if self.supp_type == 'X':
                                        self.iface.activeLayer().commitChanges()
                                        self.vl = QgsProject.instance().mapLayersByName('contacts')[0]
                                        self.iface.setActiveLayer(self.vl)
                                        self.iface.activeLayer().commitChanges()

                                        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
                                        self.iface.setActiveLayer(self.vl)
                                        self.iface.activeLayer().commitChanges()
                                        QgsMessageLog.logMessage(
                                            'JobNo: ' + str(job_num) + ' has been created and saved.',
                                            'sGIS', level=Qgis.Info)
                                    else:
                                        self.vl = QgsProject.instance().mapLayersByName('supplementals')[0]
                                        self.iface.setActiveLayer(self.vl)
                                        self.iface.activeLayer().commitChanges()
                                        QgsMessageLog.logMessage(
                                            'SuppNo: ' + str(job_num) + ' has been created and saved.',
                                            'sGIS', level=Qgis.Info)
                                else:
                                    self.iface.actionRollbackAllEdits().trigger()

                                    if self.supp_type == 'X':
                                        self.vl = QgsProject.instance().mapLayersByName('contacts')[0]
                                        self.iface.setActiveLayer(self.vl)
                                        self.iface.actionToggleEditing().trigger()
                                        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
                                        self.iface.setActiveLayer(self.vl)

                                        for f in self.vl.selectedFeatures():
                                            self.vl.deleteFeature(f.id())

                                        self.iface.activeLayer().commitChanges()

                                        if self.objectType == 'road':
                                            try:
                                                self.vl = QgsProject.instance().mapLayersByName('ng911rdss')[0]
                                                self.iface.setActiveLayer(self.vl)
                                                for f in self.vl.selectedFeatures():
                                                    self.iface.actionToggleEditing().trigger()
                                                    self.vl.deleteFeature(f.id())
                                                    self.iface.activeLayer().commitChanges()
                                                self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
                                                self.iface.setActiveLayer(self.vl)

                                            except Exception as e:
                                                exc_type, exc_obj, exc_tb = sys.exc_info()
                                                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                                                QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                                                     "Details: " + str(exc_type) + ' ' + str(
                                                                         fname) + ' ' + str(
                                                                         exc_tb.tb_lineno) + ' ' + str(e))
                                            return

                                        elif self.objectType == 'utility':
                                            try:
                                                self.vl = QgsProject.instance().mapLayersByName('utilities')[0]
                                                self.iface.setActiveLayer(self.vl)
                                                for f in self.vl.selectedFeatures():
                                                    self.iface.actionToggleEditing().trigger()
                                                    self.vl.deleteFeature(f.id())
                                                    self.iface.activeLayer().commitChanges()
                                                self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
                                                self.iface.setActiveLayer(self.vl)

                                            except Exception as e:
                                                exc_type, exc_obj, exc_tb = sys.exc_info()
                                                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                                                QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                                                     "Details: " + str(exc_type) + ' ' + str(
                                                                         fname) + ' ' + str(
                                                                         exc_tb.tb_lineno) + ' ' + str(e))
                                            return

                                        elif self.objectType == 'easement':
                                            try:
                                                self.vl = QgsProject.instance().mapLayersByName('easements')[0]
                                                self.iface.setActiveLayer(self.vl)
                                                for f in self.vl.selectedFeatures():
                                                    self.iface.actionToggleEditing().trigger()
                                                    self.vl.deleteFeature(f.id())
                                                    self.iface.activeLayer().commitChanges()
                                                self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
                                                self.iface.setActiveLayer(self.vl)

                                            except Exception as e:
                                                exc_type, exc_obj, exc_tb = sys.exc_info()
                                                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                                                QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                                                     "Details: " + str(exc_type) + ' ' + str(
                                                                         fname) + ' ' + str(
                                                                         exc_tb.tb_lineno) + ' ' + str(e))

                                        pass

                                    else:

                                        self.vl = QgsProject.instance().mapLayersByName('supplementals')[0]
                                        self.iface.setActiveLayer(self.vl)

                                        for f in self.vl.selectedFeatures():
                                            self.vl.deleteFeature(f.id())

                                        self.iface.activeLayer().commitChanges()

                                self.resetLegend()

                            elif msg.clickedButton() is again:
                                try:
                                    for a in self.iface.mainWindow().children():
                                        if a.objectName() == 'mActionDeselectAll':
                                            a.trigger()
                                            self.iface.actionSelectFreehand().trigger()
                                            self.iface.mapCanvas().selectionChanged.connect(self.feature_selected)
                                except Exception as e:
                                    pass
                                return

                            elif msg.clickedButton() is cancel:
                                try:
                                    self.iface.mapCanvas().selectionChanged.disconnect()
                                except Exception as e:
                                    pass
                                return


                        except Exception as e:
                            try:
                                self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                            except Exception as e:
                                pass
                            return

                    # <EXISTING.POLY.JOB>
                    elif self.objectType == 'poly':
                        msg = QMessageBox()
                        msg.setWindowTitle('New Existing Polygonal Job')
                        msg.setText(address + ' has been selected. Continue?')
                        create = msg.addButton('Create Job', QMessageBox.AcceptRole)
                        more = msg.addButton('Refine Selection', QMessageBox.AcceptRole)
                        cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
                        msg.setDefaultButton(create)
                        QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
                        msg.exec_()
                        msg.deleteLater()
                        QGuiApplication.restoreOverrideCursor()

                        if msg.clickedButton() is create:
                            self.newJob = 0
                            QgsMessageLog.logMessage('Job creation will begin for: ' + str(address),
                                                     'sGIS', level=Qgis.Info)
                            self.genJobFromLine()
                            self.updateJobGL()
                            self.launchForm()
                            result = self.active_edit()

                            QgsMessageLog.logMessage('Finalizing...', 'sGIS', level=Qgis.Info)

                            self.selectLastFeature(self.iface.activeLayer())

                            new_job = self.vl.selectedFeatures()[0]
                            job_num = str(new_job["job_no"])

                    # <EXISTING.LINE.JOB>
                    else:
                        msg = QMessageBox()
                        msg.setWindowTitle('New Line/Polyline Job')
                        msg.setText(address + ' has been selected. Continue?')
                        create = msg.addButton('Create Job', QMessageBox.AcceptRole)
                        more = msg.addButton('Refine Selection', QMessageBox.AcceptRole)
                        cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
                        msg.setDefaultButton(create)
                        QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
                        msg.exec_()
                        msg.deleteLater()
                        QGuiApplication.restoreOverrideCursor()

                        if msg.clickedButton() is create:

                            self.genJobFromLine()
                            self.updateJobGL()
                            self.launchForm()

                            result = self.active_edit()
                            QgsMessageLog.logMessage('Finalizing...', 'sGIS', level=Qgis.Info)

                            self.selectLastFeature(self.iface.activeLayer())
                            new_job = self.vl.selectedFeatures()[0]
                            job_num = str(new_job["job_no"])

                            if result == 1:

                                if self.supp_type == 'X':
                                    self.iface.activeLayer().commitChanges()
                                    self.vl = QgsProject.instance().mapLayersByName('contacts')[0]
                                    self.iface.setActiveLayer(self.vl)
                                    self.iface.activeLayer().commitChanges()

                                    self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
                                    self.iface.setActiveLayer(self.vl)
                                    self.iface.activeLayer().commitChanges()
                                    QgsMessageLog.logMessage('JobNo: ' + str(job_num) + ' has been created and saved.',
                                                             'sGIS', level=Qgis.Info)
                                else:
                                    self.vl = QgsProject.instance().mapLayersByName('supplementals')[0]
                                    self.iface.setActiveLayer(self.vl)
                                    self.iface.activeLayer().commitChanges()
                                    QgsMessageLog.logMessage('SuppNo: ' + str(job_num) + ' has been created and saved.',
                                                             'sGIS', level=Qgis.Info)
                            else:
                                self.iface.actionRollbackAllEdits().trigger()

                                if self.supp_type == 'X':
                                    self.vl = QgsProject.instance().mapLayersByName('contacts')[0]
                                    self.iface.setActiveLayer(self.vl)
                                    self.iface.actionToggleEditing().trigger()
                                    self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
                                    self.iface.setActiveLayer(self.vl)


                                    for f in self.vl.selectedFeatures():
                                        self.vl.deleteFeature(f.id())

                                    self.iface.activeLayer().commitChanges()

                                    if self.objectType == 'road':
                                        try:
                                            self.vl = QgsProject.instance().mapLayersByName('ng911rdss')[0]
                                            self.iface.setActiveLayer(self.vl)
                                            for f in self.vl.selectedFeatures():
                                                self.iface.actionToggleEditing().trigger()
                                                self.vl.deleteFeature(f.id())
                                                self.iface.activeLayer().commitChanges()
                                            self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
                                            self.iface.setActiveLayer(self.vl)

                                        except Exception as e:
                                            exc_type, exc_obj, exc_tb = sys.exc_info()
                                            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                                            QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                                                     exc_tb.tb_lineno) + ' ' + str(e))
                                        return

                                    elif self.objectType == 'utility':
                                        try:
                                            self.vl = QgsProject.instance().mapLayersByName('utilities')[0]
                                            self.iface.setActiveLayer(self.vl)
                                            for f in self.vl.selectedFeatures():
                                                self.iface.actionToggleEditing().trigger()
                                                self.vl.deleteFeature(f.id())
                                                self.iface.activeLayer().commitChanges()
                                            self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
                                            self.iface.setActiveLayer(self.vl)

                                        except Exception as e:
                                            exc_type, exc_obj, exc_tb = sys.exc_info()
                                            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                                            QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                                                     exc_tb.tb_lineno) + ' ' + str(e))
                                        return

                                    elif self.objectType == 'easement':
                                        try:
                                            self.vl = QgsProject.instance().mapLayersByName('easements')[0]
                                            self.iface.setActiveLayer(self.vl)
                                            for f in self.vl.selectedFeatures():
                                                self.iface.actionToggleEditing().trigger()
                                                self.vl.deleteFeature(f.id())
                                                self.iface.activeLayer().commitChanges()
                                            self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
                                            self.iface.setActiveLayer(self.vl)

                                        except Exception as e:
                                            exc_type, exc_obj, exc_tb = sys.exc_info()
                                            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                                            QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                                                     exc_tb.tb_lineno) + ' ' + str(e))

                                    pass

                                else:

                                    self.vl = QgsProject.instance().mapLayersByName('supplementals')[0]
                                    self.iface.setActiveLayer(self.vl)

                                    for f in self.vl.selectedFeatures():
                                        self.vl.deleteFeature(f.id())

                                    self.iface.activeLayer().commitChanges()

                            self.resetLegend()

                        elif msg.clickedButton() is more:
                            try:
                                for a in self.iface.mainWindow().children():
                                    if a.objectName() == 'mActionDeselectAll':
                                        a.trigger()
                                        self.iface.actionSelectFreehand().trigger()
                                        self.iface.mapCanvas().selectionChanged.connect(self.feature_selected)
                            except Exception as e:
                                pass
                            return

                        elif msg.clickedButton() is cancel:
                            try:
                                self.iface.mapCanvas().selectionChanged.disconnect()
                            except Exception as e:
                                pass
                            return

                except Exception as e:
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                         "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                             exc_tb.tb_lineno) + ' ' + str(e))
                    return

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QgsMessageLog.logMessage('FAIL: ' + str(self.supp_type), 'sGIS', level=Qgis.Info)
            QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))
            return

    def genJobFromLine(self):

        # count features / merge if necessary.
        vLayer = self.iface.activeLayer()
        feats_count = vLayer.selectedFeatureCount()
        # QgsMessageLog.logMessage('XX FEATURE COUNT: ' + str(feats_count) + ' | ' + str(self.supp_type),
        #                          'sGIS', level=Qgis.Info)
        if self.poly == 0:

            self.iface.actionCopyFeatures().trigger()
            self.fields = self.iface.activeLayer().fields()

            # create tmp layer for buffer
            self.tmpLayer = QgsVectorLayer('MultiLineString?crs=EPSG:26919', 'tmp_buffer', 'memory')
            QgsProject.instance().addMapLayers([self.tmpLayer])
            self.iface.actionIdentify().trigger()
            self.vl = QgsProject.instance().mapLayersByName('tmp_buffer')[0]
            self.iface.setActiveLayer(self.vl)

            # add fields to tmp layer
            self.iface.actionToggleEditing()
            self.layerData = self.tmpLayer.dataProvider()
            self.layerData.addAttributes(self.fields)

            self.iface.activeLayer().commitChanges()

            # paste copied feature to tmp and buffer
            self.iface.actionToggleEditing().trigger()
            self.iface.actionPasteFeatures().trigger()

            if feats_count == 0:
                pass
            elif feats_count == 1:
                self.selectLastFeature(self.vl)
                pass
            else:
                self.iface.mainWindow().findChild(QAction, 'mActionMergeFeatures').trigger()

            self.bufferLINE('tmp_buffer')

            self.vl = QgsProject.instance().mapLayersByName('tmp_buffer')[0]
            self.iface.setActiveLayer(self.vl)
            self.iface.activeLayer().commitChanges()
            QgsProject.instance().removeMapLayer(self.vl.id())

            # copy selected feature from output layer
            self.vl = QgsProject.instance().mapLayersByName('output')[0]
            self.iface.setActiveLayer(self.vl)
            self.layerData = self.vl.dataProvider()
            self.iface.actionToggleEditing().trigger()
            self.layerData.addAttributes([QgsField("supp_type", QVariant.String)])
            self.iface.activeLayer().commitChanges()

            new_supp = self.vl.selectedFeatures()[0]
            self.updateAttribute(new_supp, 'supp_type', str(self.supp_type))

            self.iface.actionCopyFeatures().trigger()

            if self.supp_type == 'X':
                self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
            else:
                self.vl = QgsProject.instance().mapLayersByName('supplementals')[0]
            self.iface.setActiveLayer(self.vl)
            self.iface.actionToggleEditing().trigger()
            self.iface.actionPasteFeatures().trigger()
            self.iface.activeLayer().commitChanges()
            layer2 = QgsProject.instance().mapLayersByName('output')[0]
            QgsProject.instance().removeMapLayer(layer2.id())

        elif self.poly == 2:

            self.iface.actionCopyFeatures().trigger()
            self.fields = self.iface.activeLayer().fields()

            # create tmp layer for merge
            self.tmpLayer = QgsVectorLayer('MultiPolygon?crs=EPSG:102683', 'tmp_merge', 'memory')
            QgsProject.instance().addMapLayers([self.tmpLayer])
            self.iface.actionIdentify().trigger()

            # add fields to tmp layer
            self.layerData = self.tmpLayer.dataProvider()
            self.layerData.addAttributes(self.fields)

            self.layerData.addAttributes([QgsField("supp_type", QVariant.String)])
            self.tmpLayer.updateFields()

            self.iface.setActiveLayer(self.vl)
            self.iface.actionCopyFeatures().trigger()

            self.iface.setActiveLayer(self.tmpLayer)
            self.iface.actionToggleEditing().trigger()
            self.iface.actionPasteFeatures().trigger()

            if feats_count == 0:
                pass
            elif feats_count == 1:
                self.iface.activeLayer().commitChanges()
                self.selectLastFeature(self.tmpLayer)
                pass
            else:
                self.iface.mainWindow().findChild(QAction, 'mActionMergeFeatures').trigger()
                self.iface.activeLayer().commitChanges()
                self.selectLastFeature(self.tmpLayer)

            self.vl = self.tmpLayer

            new_supp = self.tmpLayer.selectedFeatures()[0]
            self.updateAttribute(new_supp, 'supp_type', str(self.supp_type))

            self.iface.actionCopyFeatures().trigger()

            self.vl = QgsProject.instance().mapLayersByName('supplementals')[0]
            self.iface.setActiveLayer(self.vl)

            self.iface.actionToggleEditing().trigger()
            self.iface.actionPasteFeatures().trigger()
            self.iface.activeLayer().commitChanges()

            QgsProject.instance().removeMapLayer(self.tmpLayer.id())

            self.iface.actionIdentify().trigger()

        else:
            # nothing to do - not a line
            pass

    def updateJobGL(self):

        if self.supp_type == 'X':
            self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
            self.iface.setActiveLayer(self.vl)

        else:
            self.vl = QgsProject.instance().mapLayersByName('supplementals')[0]
            self.iface.setActiveLayer(self.vl)

        try:
            self.selectLastFeature(self.vl)
            new_job = self.vl.selectedFeatures()[0]
            job_num = str(new_job["job_no"])

        except Exception as e:
            QgsMessageLog.logMessage('FAIL FAIL FAIL!', 'sGIS', level=Qgis.Info)
            return 1

        layerData = self.vl.dataProvider()

        try:
            lat_lon0 = new_job['lat_lon']
            lat_lonF = formatLL(lat_lon0)

        except Exception as e:
            QgsMessageLog.logMessage('FAIL @ lat_lon', 'sGIS', level=Qgis.Info)
            pass

        try:
            county = self.extract('county', 'COUNTY')
            layer2 = QgsProject.instance().mapLayersByName('output')[0]
            QgsProject.instance().removeMapLayer(layer2.id())
        except Exception as e:
            county = ''

        try:
            town = self.extract('metwp24P', 'TOWN')
            layer2 = QgsProject.instance().mapLayersByName('output')[0]
            QgsProject.instance().removeMapLayer(layer2.id())
        except Exception as e:
            town = ''

        try:
            zipCode = self.extract('Zip Codes', 'ZCTA5CE10')
            layer2 = QgsProject.instance().mapLayersByName('output')[0]
            QgsProject.instance().removeMapLayer(layer2.id())
        except Exception as e:
            zipCode = ''

        result = county + ' | ' + town + ' | ' + zipCode
        # QgsMessageLog.logMessage('GEOLOCATION: ' + result, 'sGIS', level=Qgis.Info)
        self.iface.activeLayer().commitChanges()

        if self.supp_type == 'X':
            self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
            self.iface.setActiveLayer(self.vl)
            self.selectLastFeature(self.vl)
        else:
            self.vl = QgsProject.instance().mapLayersByName('supplementals')[0]
            self.iface.setActiveLayer(self.vl)
            self.selectLastFeature(self.vl)

        new_job = self.vl.selectedFeatures()[0]
        job_num = str(new_job["job_no"])

        try:
            self.updateAttribute(new_job, 'lat_lon', lat_lonF)
        except Exception as e:
            pass

        self.updateAttribute(new_job, 'locus_addr', self.street)
        self.updateAttribute(new_job, 'town', town)
        self.updateAttribute(new_job, 'county', county)
        self.updateAttribute(new_job, 'zipcode', zipCode)
        # self.updateAttribute(new_job, 'map_bk_lot', address)
        self.updateAttribute(new_job, 'area', '0.00')
        self.updateAttribute(new_job, 'objectType', self.objectType)

        return 0

    def active_edit(self):

        QgsMessageLog.logMessage('active_edit started.', 'sGIS', level=Qgis.Info)

        if self.supp_type == 'X':
            self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
            self.iface.setActiveLayer(self.vl)
        else:
            self.vl = QgsProject.instance().mapLayersByName('supplementals')[0]
            self.iface.setActiveLayer(self.vl)

        self.selectLastFeature(self.vl)
        self.vl = self.iface.activeLayer()
        f = self.vl.selectedFeatures()[0]

        QGuiApplication.setOverrideCursor(Qt.ArrowCursor)

        result = self.iface.openFeatureForm(self.iface.activeLayer(), f, False, True)
        if result:
            pass
        else:
            QGuiApplication.restoreOverrideCursor()
            QgsMessageLog.logMessage('DEBUG: active_edit cancelled.', 'sGIS', level=Qgis.Info)
            return 0

        if self.supp_type == 'X':
            QgsMessageLog.logMessage('START ABUTTERS...', 'sGIS', level=Qgis.Info)
            self.abutters_dialog = sgis_abutters(self.iface)
            self.abutters_dialog.run()
            QgsMessageLog.logMessage('END ABUTTERS...', 'sGIS', level=Qgis.Info)
        else:
            pass

        QGuiApplication.restoreOverrideCursor()

        # if self.supp_type == 'X':
        #     self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        #     self.iface.setActiveLayer(self.vl)
        # else:
        #     self.vl = QgsProject.instance().mapLayersByName('supplementals')[0]
        #     self.iface.setActiveLayer(self.vl)

        self.iface.activeLayer().commitChanges()
        self.iface.messageBar().clearWidgets()
        return 1

    def setLineSourceLayers(self, msg, type):

        if self.poly == 1:

            self.objectType = 'poly'
            self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
            self.iface.setActiveLayer(self.vl)
            # QgsMessageLog.logMessage('DEBUG: IDd as poly.', 'sGIS', level=Qgis.Info)
            return self.objectType

        else:

            msg.setText('Which type of feature are we working with?')
            road = msg.addButton('Road', QMessageBox.AcceptRole)
            utility = msg.addButton('Utility', QMessageBox.AcceptRole)
            easement = msg.addButton('Access/Easement', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)

            msg.setDefaultButton(cancel)
            QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
            msg.exec_()
            msg.deleteLater()
            QGuiApplication.restoreOverrideCursor()

            if msg.clickedButton() is road:
                self.objectType = 'road'
                if self.type == 'new':
                    self.vl = QgsProject.instance().mapLayersByName('new_roads')[0]
                else:
                    self.vl = QgsProject.instance().mapLayersByName('ng911rdss')[0]
                self.iface.setActiveLayer(self.vl)

                return self.objectType

            elif msg.clickedButton() is utility:
                self.objectType = 'utility'
                self.vl = QgsProject.instance().mapLayersByName('utilities')[0]
                self.iface.setActiveLayer(self.vl)
                return self.objectType

            elif msg.clickedButton() is easement:
                self.objectType = 'easement'
                self.vl = QgsProject.instance().mapLayersByName('easements')[0]
                self.iface.setActiveLayer(self.vl)
                return self.objectType

            elif msg.clickedButton() is cancel:
                self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
                self.iface.setActiveLayer(self.vl)
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception:
                    return

    def updateAttribute(self, job, attVar, attVal):
        layerData = self.vl.dataProvider()
        self.iface.actionToggleEditing().trigger()
        idx = layerData.fieldNameIndex(attVar)
        self.vl.changeAttributeValue(job.id(), idx, attVal)
        self.vl.updateFields()
        self.iface.activeLayer().commitChanges()

    def bufferLINE(self, sInputLayer):

        Processing.initialize()
        inputL = QgsProject.instance().mapLayersByName(sInputLayer)[0]

        self.fb = QgsProcessingFeedback()
        self.context = QgsProcessingContext
        QgsMessageLog.logMessage(sInputLayer, 'sGIS', level=Qgis.Info)
        self.outputLayer = processing.run("native:buffer", {
            'INPUT': sInputLayer,
            'DISSOLVE': False,
            'DISTANCE': 5,
            'END_CAP_STYLE': 2,
            'JOIN_STYLE': 0,
            'MITER_LIMIT': 2,
            'SEGMENTS': 5,
            'OUTPUT': 'memory:',
            'PREDICATE': [0]}, feedback=self.fb)['OUTPUT']

        try:
            QgsProject.instance().addMapLayer(self.outputLayer)
            self.iface.setActiveLayer(self.outputLayer)
            self.selectLastFeature(self.outputLayer)
            return 1

        except Exception as e:

            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "EXCEPTIONz",
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))

    def extract(self, sInputLayer, sInputAttribute):

        Processing.initialize()
        inputL = QgsProject.instance().mapLayersByName(sInputLayer)[0]

        self.fb = QgsProcessingFeedback()
        self.context = QgsProcessingContext
        # QgsMessageLog.logMessage(sInputLayer, 'sGIS', level=Qgis.Info)

        if self.supp_type == 'X':
            iLayer = 'jobs'
        else:
            iLayer = 'supplementals' \
                     ''
        self.outputLayer = processing.run("native:extractbylocation", {
            'INPUT': sInputLayer,
            'INTERSECT': QgsProcessingFeatureSourceDefinition(
                iLayer, True),
            'OUTPUT': 'memory:output',
            'PREDICATE': [0]}, feedback=self.fb)['OUTPUT']

        # QgsMessageLog.logMessage(str(self.outputLayer), 'sGIS', level=Qgis.Info)

        try:
            QgsProject.instance().addMapLayer(self.outputLayer)
        except Exception as e:
            return
        try:
            self.selectLastFeature(self.outputLayer)
            source = self.iface.activeLayer().selectedFeatures()[0]
            output = source[sInputAttribute]
            return str(output)

        except Exception as e:
            return
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))

    def launchForm(self):
        QgsMessageLog.logMessage('Launching form for editing...', 'sGIS', level=Qgis.Info)

        if self.supp_type == 'X':
            self.vl = QgsProject.instance().mapLayersByName('contacts')[0]
            self.iface.setActiveLayer(self.vl)
            self.iface.actionToggleEditing().trigger()

            self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
            self.iface.setActiveLayer(self.vl)
            self.iface.actionToggleEditing().trigger()
            self.iface.actionIdentify().trigger()
        else:
            self.vl = QgsProject.instance().mapLayersByName('supplementals')[0]
            self.iface.setActiveLayer(self.vl)
            self.iface.actionToggleEditing().trigger()

    def resetLegend(self):
        root = QgsProject.instance().layerTreeRoot()
        for child in root.children():
            child.setExpanded(False)
        aGroup = root.findGroup('Surveyor')
        aGroup.setExpanded(True)

    def selectLastFeature(self, layer):

        # layer.setSubsetString('id > 1')
        # layer.setSubsetString("supp_type != 'X'")
        f2 = layer.getFeatures()
        fCount = layer.featureCount()
        fIds = []
        fIds = [f.id() for f in f2]
        fIds.sort()

        fId = [fIds[-1]]

        layer.selectByIds(fId)


class sgis_newParcelSupp(object):
    newJob = 0
    selComp = 0
    multiFeat = 0
    count = 0

    def __init__(self, iface):
        # save reference to the QGIS interface
        self.iface = iface

    def initGui(self, type):

        self.action = QAction("Create Supplemental (Parcel-based)", self.iface.mainWindow())
        self.action.triggered.connect(self.run)
        self.action.trigger()
        self.iface.mapCanvas().selectionChanged.connect(self.select_changed)

    def run(self, type):
        self.vl = QgsProject.instance().mapLayersByName('Parcels')[0]
        self.iface.setActiveLayer(self.vl)

        reply = QMessageBox.question(self.iface.mainWindow(), 'New Supplemental (Parcel-based)',
                                     'Click OK and select the correct parcel(s) for the new supplemental.',
                                     QMessageBox.Ok, QMessageBox.Cancel)
        if reply == QMessageBox.Ok:
            if self.selComp == 1:
                return
            else:
                self.newJob = 1
                if self.count == 0:
                    for a in self.iface.mainWindow().children():
                        if a.objectName() == 'mActionDeselectAll':
                            a.trigger()
                            QgsMessageLog.logMessage('FIRST RUN: Previously selected parcel(s) have been cleared.',
                                                     'sGIS', level=Qgis.Info)
                            QgsMessageLog.logMessage('PLAN creation starting...', 'sGIS', level=Qgis.Info)
                            self.iface.actionSelectFreehand().trigger()
                            self.count = self.count + 1
                else:
                    self.iface.actionSelect().trigger()

        else:
            QgsMessageLog.logMessage('DEBUG: PLAN creation cancelled.', 'sGIS', level=Qgis.Info)

        resetLegend(self)

    def select_changed(self):

        if self.newJob == 0:
            return
        else:
            try:

                vLayer = self.iface.activeLayer()
                feats_count = vLayer.selectedFeatureCount()

                f = vLayer.selectedFeatures()[0]
                msg = QMessageBox()
                msg.setWindowTitle('Selection')
                msg.setText(str(feats_count) + ' features have been selected. Continue?')
                create = msg.addButton('Continue', QMessageBox.AcceptRole)
                again = msg.addButton('Select Again', QMessageBox.AcceptRole)
                cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
                msg.setDefaultButton(create)
                msg.exec_()
                msg.deleteLater()

                QGuiApplication.restoreOverrideCursor()

                if msg.clickedButton() is create:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                    QgsMessageLog.logMessage('PLAN creation will begin for multiple Parcels:',
                                             'sGIS', level=Qgis.Info)

                    self.iface.actionCopyFeatures().trigger()
                    self.vl = QgsProject.instance().mapLayersByName('supplementals')[0]
                    self.iface.setActiveLayer(self.vl)
                    self.iface.actionToggleEditing().trigger()
                    self.iface.actionPasteFeatures().trigger()
                    self.iface.mainWindow().findChild(QAction, 'mActionMergeFeatures').trigger()

                    self.iface.actionIdentify().trigger()

                    # end merge features, launch form
                    self.iface.activeLayer().commitChanges()
                    layer = self.iface.activeLayer()
                    prov = layer.dataProvider()

                    idx = prov.fieldNameIndex('sid')

                    oid = layer.maximumValue(idx)
                    it = layer.getFeatures(QgsFeatureRequest().setFilterExpression(u'"sid" = {0}'.format(oid)))
                    # QgsMessageLog.logMessage('sid: ' + str(sid) + ' it: ' + str(it), 'sGIS', level=Qgis.Info)
                    for feature in it:
                        f = feature.id()
                        layer.select(f)

                    QgsMessageLog.logMessage('Launching form for merged feature...', 'sGIS', level=Qgis.Info)

                    self.iface.activeLayer().commitChanges()
                    self.iface.actionToggleEditing().trigger()

                    result = self.active_edit()
                    QgsMessageLog.logMessage('RESULT: ' + str(result), 'sGIS', level=Qgis.Info)

                    if result:
                        self.iface.activeLayer().commitChanges()
                    else:
                        self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                        self.iface.actionRollbackAllEdits().trigger()

                        for f in self.vl.selectedFeatures():
                            self.vl.deleteFeature(f.id())

                        self.iface.activeLayer().commitChanges()
                        return

                    self.vl = QgsProject.instance().mapLayersByName('supplementals')[0]

                    jobNo = self.vl.selectedFeatures()[0]
                    lat_lon = jobNo['lat_lon']
                    dd = float(lat_lon.split(',')[0])
                    dd2 = float(lat_lon.split(',')[1])

                    d = int(float(dd))
                    m = int(float((dd - d)) * 60)
                    s = (dd - d - m / 60) * 3600.00
                    z = round(s, 2)

                    if d >= 0:
                        lat = (str(abs(d)) + '°' + str(abs(m)) + "'" + str(abs(z)) + '"' + 'N')
                    else:
                        lat = (str(abs(d)) + '°' + str(abs(m)) + "'" + str(abs(z)) + '"' + 'S')

                    d = int(float(dd2))
                    m = int(float((dd2 - d)) * 60)
                    s = (dd2 - d - m / 60) * 3600.00
                    z = round(s, 2)

                    if d >= 0:
                        lon = (str(abs(d)) + '°' + str(abs(m)) + "'" + str(abs(z)) + '"' + 'E')
                    else:
                        lon = (str(abs(d)) + '°' + str(abs(m)) + "'" + str(abs(z)) + '"' + 'W')

                    lat_lon = str(lat) + '    ' + str(lon)
                    lyr = QgsProject.instance().mapLayersByName('supplementals')[0]
                    # start editing, change field value
                    self.iface.actionToggleEditing().trigger()
                    layerData = lyr.dataProvider()
                    idx3 = layerData.fieldNameIndex('lat_lon')
                    lyr.changeAttributeValue(jobNo.id(), idx3, lat_lon)
                    lyr.updateFields()
                    # end field update, save layer
                    self.iface.activeLayer().commitChanges()

                    self.iface.setActiveLayer(self.vl)

                    layer = self.iface.activeLayer()
                    prov = layer.dataProvider()

                    idx = prov.fieldNameIndex('sid')

                    oid = layer.maximumValue(idx)
                    it = layer.getFeatures(QgsFeatureRequest().setFilterExpression(u'"sid" = {0}'.format(oid)))

                    for feature in it:
                        f = feature.id()
                        layer.select(f)

                    jobNo = self.vl.selectedFeatures()[0]
                    kNo = jobNo["job_no"]
                    QgsMessageLog.logMessage('SuppNo: ' + str(kNo) + ' has been created and saved.',
                                             'sGIS', level=Qgis.Info)
                    lyr = QgsProject.instance().mapLayersByName('supplementals')[0]
                    self.iface.setActiveLayer(lyr)

                elif msg.clickedButton() is again:
                    # self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                    QGuiApplication.restoreOverrideCursor()
                    # return

                elif msg.clickedButton() is cancel:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                    for a in self.iface.mainWindow().children():
                        if a.objectName() == 'mActionDeselectAll':
                            a.trigger()
                            QgsMessageLog.logMessage('PLAN creation starting...', 'sGIS', level=Qgis.Info)
                            self.iface.actionIdentify().trigger()

                    QGuiApplication.restoreOverrideCursor()
                    return


            except Exception as e:
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception as e:
                    pass
                return

    def active_edit(self):

        try:

            self.vl = QgsProject.instance().mapLayersByName('supplementals')[0]
            self.iface.setActiveLayer(self.vl)
            self.iface.activeLayer().commitChanges()
            self.iface.actionToggleEditing().trigger()
            layer = self.iface.activeLayer()
            prov = layer.dataProvider()

            idx = prov.fieldNameIndex('sid')

            gid = layer.maximumValue(idx)
            it = layer.getFeatures(QgsFeatureRequest().setFilterExpression(u'"sid" = {0}'.format(gid)))

            for a in self.iface.mainWindow().children():
                if a.objectName() == 'mActionDeselectAll':
                    a.trigger()
                    QgsMessageLog.logMessage('CLEAR: Previously selected parcel(s) have been cleared.',
                                             'sGIS', level=Qgis.Info)

            for feature in it:
                f = feature.id()
                layer.select(f)

            f = layer.selectedFeatures()[0]

            QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
            self.iface.openFeatureForm(self.vl, f, False, True)
            QGuiApplication.restoreOverrideCursor()
            self.iface.activeLayer().commitChanges()
            self.iface.messageBar().clearWidgets()

            QgsMessageLog.logMessage('Saving changes...', 'sGIS', level=Qgis.Info)
            return "GOOD!"



        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))
            return


class sgis_editJob(object):

    def __init__(self, iface):
        self.iface = iface

    def initGui(self):

        self.action = QAction("Edit Job", self.iface.mainWindow())
        self.action.triggered.connect(self.run)
        self.action.trigger()
        self.iface.mapCanvas().selectionChanged.connect(self.select_changed)

    def run(self):
        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(self.vl)

        vLayer = self.iface.activeLayer()
        feats_count = vLayer.selectedFeatureCount()

        if feats_count == 0:
            msg = QMessageBox()
            msg.setWindowTitle('Selection')
            msg.setText('Choose the job you wish to edit and click OK to proceed.')
            cont = msg.addButton('Continue', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(cont)
            msg.exec_()
            msg.deleteLater()
            if msg.clickedButton() is cont:
                self.iface.actionSelect().trigger()
            else:
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception as e:
                    pass
                return
        else:
            parcel = self.iface.activeLayer().selectedFeatures()[0]
            jobNo = parcel["job_no"]

            msg = QMessageBox()
            msg.setWindowTitle('Selection')
            msg.setText(str(jobNo) + ' is selected.  Do you want to edit that selection or choose a '
                                           'new selection?')
            cont = msg.addButton('Continue', QMessageBox.AcceptRole)
            new = msg.addButton('New Selection', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(cont)
            msg.exec_()
            msg.deleteLater()

            if msg.clickedButton() is cont:
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception as e:
                    pass

                self.select_changed()
            elif msg.clickedButton() is new:
                msg = QMessageBox()
                msg.setWindowTitle('Selection')
                msg.setText('Choose the job you wish to edit and click OK to proceed.')
                cont = msg.addButton('Continue', QMessageBox.AcceptRole)
                cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
                msg.setDefaultButton(cont)
                msg.exec_()
                msg.deleteLater()
                if msg.clickedButton() is cont:
                    for a in self.iface.mainWindow().children():

                        if a.objectName() == 'mActionDeselectAll':
                            a.trigger()
                            QgsMessageLog.logMessage('FIRST RUN: Previous selection has been cleared.',
                                                     'sGIS', level=Qgis.Info)
                            QgsMessageLog.logMessage('Job editing starting...', 'sGIS', level=Qgis.Info)
                            self.iface.actionSelect().trigger()
                elif msg.clickedButton() is cancel:
                    try:
                        self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                    except Exception as e:
                        pass

            elif msg.clickedButton() is cancel:
                QgsMessageLog.logMessage('DEBUG: Job editing cancelled.', 'sGIS', level=Qgis.Info)
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception as e:
                    pass

    def select_changed(self):

        try:
            parcel = self.iface.activeLayer().selectedFeatures()[0]
            jobNo = parcel["job_no"]

            msg = QMessageBox()
            msg.setWindowTitle('SOURCE Job')
            msg.setText('JobNo: ' + jobNo + ' has been selected. Continue?')
            edit = msg.addButton('Select Job', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(edit)
            QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
            msg.exec_()
            msg.deleteLater()
            QGuiApplication.restoreOverrideCursor()
            if msg.clickedButton() is edit:

                QgsMessageLog.logMessage('Launching form...', 'sGIS', level=Qgis.Info)
                self.iface.actionToggleEditing().trigger()
                self.vl = QgsProject.instance().mapLayersByName('contacts')[0]
                self.iface.setActiveLayer(self.vl)
                self.iface.actionToggleEditing().trigger()
                self.active_edit()
                QgsMessageLog.logMessage('JobNo:' + parcel["job_no"] + ' has been modified and saved.',
                                         'sGIS', level=Qgis.Info)

                self.vl = QgsProject.instance().mapLayersByName('contacts')[0]
                self.iface.setActiveLayer(self.vl)
                self.iface.activeLayer().commitChanges()
                lyr = QgsProject.instance().mapLayersByName('jobs')[0]
                self.iface.setActiveLayer(lyr)
                self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                self.iface.actionIdentify().trigger()
                resetLegend(self)

            elif msg.clickedButton() is cancel:
                self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                QgsMessageLog.logMessage('DEBUG: Job editing cancelled.', 'sGIS', level=Qgis.Info)
                return
            else:
                    QgsMessageLog.logMessage('DEBUG: Job editing cancelled.', 'sGIS', level=Qgis.Info)
        except Exception as e:
            QgsMessageLog.logMessage('EXCEPTION: ' + str(e), 'sGIS', level=Qgis.Info)

    def active_edit(self):

        try:
            self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
            self.iface.setActiveLayer(self.vl)
            f = self.vl.selectedFeatures()[0]
            # change to jobs form

            lat_lon = f['lat_lon']
            if str(lat_lon) == 'NULL':
                lat_lon = ''
            else:
                lat_lon = lat_lon

            ll = len(lat_lon)
            # QgsMessageLog.logMessage('ll: ' + lat_lon + ' | ' + str(ll), 'sGIS', level=Qgis.Info)

            if ll <= 30:
                pass
            else:
                lat_lon = formatLL(lat_lon)

                #self.iface.actionToggleEditing().trigger()
                layerData = self.vl.dataProvider()
                idx = layerData.fieldNameIndex('lat_lon')
                self.vl.changeAttributeValue(f.id(), idx, lat_lon)
                self.vl.updateFields()
                #self.iface.activeLayer().commitChanges()

            QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
            self.iface.openFeatureForm(self.iface.activeLayer(), f, False, True)
            QGuiApplication.restoreOverrideCursor()
            QgsMessageLog.logMessage('llF: ' + lat_lon, 'sGIS', level=Qgis.Info)

            QgsMessageLog.logMessage('Saving changes...', 'sGIS', level=Qgis.Info)

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))
            return

        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(self.vl)
        self.iface.activeLayer().commitChanges()


class sgis_editSupp(object):

    def __init__(self, iface):
        # save reference to the QGIS interface
        self.iface = iface

    def initGui(self):

        self.action = QAction("Edit Supplemental", self.iface.mainWindow())
        self.action.triggered.connect(self.run)
        self.action.trigger()
        self.iface.mapCanvas().selectionChanged.connect(self.select_changed)

    def run(self):
        self.vl = QgsProject.instance().mapLayersByName('supplementals')[0]
        self.iface.setActiveLayer(self.vl)

        vLayer = self.iface.activeLayer()
        feats_count = vLayer.selectedFeatureCount()

        if feats_count == 0:
            msg = QMessageBox()
            msg.setWindowTitle('Selection')
            msg.setText('Choose the supplemental you wish to edit and click OK to proceed.')
            cont = msg.addButton('Continue', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(cont)
            msg.exec_()
            msg.deleteLater()
            if msg.clickedButton() is cont:
                self.iface.actionSelect().trigger()
                pass
            else:
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception as e:
                    pass
                return
        else:
            msg = QMessageBox()
            msg.setWindowTitle('Selection')
            msg.setText(str(feats_count) + ' feature(s) are selected. Do you want to edit that selection or choose a '
                                           'new selection?')
            cont = msg.addButton('Continue', QMessageBox.AcceptRole)
            new = msg.addButton('New Selection', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(cont)
            msg.exec_()
            msg.deleteLater()

            if msg.clickedButton() is cont:
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception as e:
                    pass

                self.select_changed()
            elif msg.clickedButton() is new:
                msg = QMessageBox()
                msg.setWindowTitle('Selection')
                msg.setText('Choose the supplemental you wish to edit and click OK to proceed.')
                cont = msg.addButton('Continue', QMessageBox.AcceptRole)
                cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
                msg.setDefaultButton(cont)
                msg.exec_()
                msg.deleteLater()
                if msg.clickedButton() is cont:
                    for a in self.iface.mainWindow().children():
                        if a.objectName() == 'mActionDeselectAll':
                            a.trigger()
                            QgsMessageLog.logMessage('FIRST RUN: Previous selection has been cleared.',
                                                     'sGIS', level=Qgis.Info)
                            QgsMessageLog.logMessage('Supplemental editing starting...', 'sGIS', level=Qgis.Info)
                            self.iface.actionSelect().trigger()
                elif msg.clickedButton() is cancel:
                    try:
                        self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                    except Exception as e:
                        pass

            elif msg.clickedButton() is cancel:
                QgsMessageLog.logMessage('DEBUG: Job editing cancelled.', 'sGIS', level=Qgis.Info)
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception as e:
                    pass

    def select_changed(self):
        try:
            self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
        except Exception as e:
            pass
        try:
            supp = self.iface.activeLayer().selectedFeatures()[0]
            suppNo = supp["job_no"]
            msg = QMessageBox()
            msg.setWindowTitle('Edit Supplemental')
            msg.setText('SuppNo: ' + suppNo + ' has been selected. Continue?')
            edit = msg.addButton('Edit', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(edit)
            QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
            msg.exec_()
            msg.deleteLater()
            QGuiApplication.restoreOverrideCursor()
            if msg.clickedButton() is edit:
                QgsMessageLog.logMessage('Supplemental editing will begin for: ' + suppNo,
                                         'sGIS', level=Qgis.Info)
                QgsMessageLog.logMessage('Launching form...', 'sGIS', level=Qgis.Info)
                self.iface.actionToggleEditing().trigger()
                self.active_edit()
                QgsMessageLog.logMessage('SuppNo:' + supp["job_no"] + ' has been modified and saved.',
                                         'sGIS', level=Qgis.Info)

                lyr = QgsProject.instance().mapLayersByName('supplementals')[0]
                self.iface.setActiveLayer(lyr)
                self.iface.actionIdentify().trigger()

            elif msg.clickedButton() is cancel:
                QgsMessageLog.logMessage('DEBUG: Supplemental editing cancelled.', 'sGIS', level=Qgis.Info)
                QGuiApplication.restoreOverrideCursor()
                return
            else:
                self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                QgsMessageLog.logMessage('DEBUG: Supplemental editing cancelled.', 'sGIS', level=Qgis.Info)
                QGuiApplication.restoreOverrideCursor()

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))
            QGuiApplication.restoreOverrideCursor()
            return

    def active_edit(self):

        try:
            self.vl = QgsProject.instance().mapLayersByName('supplementals')[0]
            self.iface.setActiveLayer(self.vl)
            f = self.vl.selectedFeatures()[0]

            QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
            result = self.iface.openFeatureForm(self.iface.activeLayer(), f, False, True)
            QgsMessageLog.logMessage('RESULT: ' + str(result), 'sGIS', level=Qgis.Info)

            QGuiApplication.restoreOverrideCursor()
            QgsMessageLog.logMessage('Saving changes...', 'sGIS', level=Qgis.Info)
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))
            QGuiApplication.restoreOverrideCursor()

            return

        self.vl = QgsProject.instance().mapLayersByName('supplementals')[0]
        self.iface.setActiveLayer(self.vl)
        self.iface.activeLayer().commitChanges()


class sgis_editPlan(object):

    def __init__(self, iface):
        # save reference to the QGIS interface
        self.iface = iface

    def initGui(self):

        self.action = QAction("Edit Plan", self.iface.mainWindow())
        self.action.triggered.connect(self.run)
        self.action.trigger()
        self.iface.mapCanvas().selectionChanged.connect(self.select_changed)

    def run(self):
        self.vl = QgsProject.instance().mapLayersByName('plans')[0]
        self.iface.setActiveLayer(self.vl)
        vLayer = self.iface.activeLayer()
        feats_count = vLayer.selectedFeatureCount()

        if feats_count == 0:
            msg = QMessageBox()
            msg.setWindowTitle('Selection')
            msg.setText('Choose the correct parcel for the plan you wish to edit and click OK to proceed.')
            cont = msg.addButton('Continue', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(cont)
            msg.exec_()
            msg.deleteLater()
            if msg.clickedButton() is cont:
                self.iface.actionSelect().trigger()
            else:
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception as e:
                    pass
                return
        else:
            msg = QMessageBox()
            msg.setWindowTitle('Selection')
            msg.setText(str(feats_count) + ' feature(s) are selected. Do you want to edit that selection or choose a '
                                           'new selection?')
            cont = msg.addButton('Continue', QMessageBox.AcceptRole)
            new = msg.addButton('New Selection', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(cont)
            msg.exec_()
            msg.deleteLater()

            if msg.clickedButton() is cont:
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception as e:
                    pass
                self.select_changed()
            elif msg.clickedButton() is new:
                msg = QMessageBox()
                msg.setWindowTitle('Selection')
                msg.setText('Choose the plan you wish to edit and click OK to proceed.')
                cont = msg.addButton('Continue', QMessageBox.AcceptRole)
                cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
                msg.setDefaultButton(cont)
                msg.exec_()
                msg.deleteLater()
                if msg.clickedButton() is cont:
                    for a in self.iface.mainWindow().children():

                        if a.objectName() == 'mActionDeselectAll':
                            a.trigger()
                            QgsMessageLog.logMessage('FIRST RUN: Previous selection has been cleared.',
                                                     'sGIS', level=Qgis.Info)
                            QgsMessageLog.logMessage('Plan editing starting...', 'sGIS', level=Qgis.Info)
                            self.iface.actionSelect().trigger()
                elif msg.clickedButton() is cancel:
                    try:
                        self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                    except Exception as e:
                        pass
            elif msg.clickedButton() is cancel:
                QgsMessageLog.logMessage('DEBUG: Plan editing cancelled.', 'sGIS', level=Qgis.Info)
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception as e:
                    pass

    def select_changed(self):

        try:
            parcel = self.iface.activeLayer().selectedFeatures()[0]
            planNo = parcel["plan_no"]
            msg = QMessageBox()
            msg.setWindowTitle('Last Chance')
            msg.setText('PlanNo: ' + planNo + ' will be edited. Continue?')
            edit = msg.addButton('Edit Plan', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(edit)
            QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
            msg.exec_()
            msg.deleteLater()
            QGuiApplication.restoreOverrideCursor()
            if msg.clickedButton() is edit:
                QgsMessageLog.logMessage('Plan editing will begin for: ' + planNo,
                                         'sGIS', level=Qgis.Info)
                QgsMessageLog.logMessage('Launching form...', 'sGIS', level=Qgis.Info)
                self.iface.actionToggleEditing().trigger()
                self.active_edit()
                QgsMessageLog.logMessage('PlanNo:' + parcel["plan_no"] + ' has been modified and saved.',
                                         'sGIS', level=Qgis.Info)

                lyr = QgsProject.instance().mapLayersByName('plans')[0]
                self.iface.setActiveLayer(lyr)
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception as e:
                    self.iface.actionIdentify().trigger()

            elif msg.clickedButton() is cancel:
                QgsMessageLog.logMessage('DEBUG: Plan editing cancelled.', 'sGIS', level=Qgis.Info)
                QGuiApplication.restoreOverrideCursor()
                return
            else:
                self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                QgsMessageLog.logMessage('DEBUG: Plan editing cancelled.', 'sGIS', level=Qgis.Info)
                QGuiApplication.restoreOverrideCursor()

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))
            QGuiApplication.restoreOverrideCursor()
            return

    def updateAttribute(self, job, attVar, attVal):
        layerData = self.vl.dataProvider()
        self.iface.actionToggleEditing().trigger()
        idx = layerData.fieldNameIndex(attVar)
        self.vl.changeAttributeValue(job.id(), idx, attVal)
        self.vl.updateFields()
        self.iface.activeLayer().commitChanges()


    def active_edit(self):

        try:
            self.vl = QgsProject.instance().mapLayersByName('plans')[0]
            self.iface.setActiveLayer(self.vl)
            f = self.vl.selectedFeatures()[0]
            l_l = f['lat_lon']
            try:
                self.iface.actionToggleEditing().trigger()
                lat_lon = formatLL(l_l)
                self.updateAttribute(f, 'lat_lon', lat_lon)
            except Exception as e:
                pass
            self.iface.actionToggleEditing().trigger()
            #change to plans form

            QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
            result = self.iface.openFeatureForm(self.iface.activeLayer(), f, False, True)
            # QgsMessageLog.logMessage('RESULT: ' + str(result), 'sGIS', level=Qgis.Info)

            QGuiApplication.restoreOverrideCursor()
            QgsMessageLog.logMessage('Saving changes...', 'sGIS', level=Qgis.Info)
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))
            QGuiApplication.restoreOverrideCursor()

            return

        self.vl = QgsProject.instance().mapLayersByName('plans')[0]
        self.iface.setActiveLayer(self.vl)
        self.iface.activeLayer().commitChanges()


class sgis_deleteFeature(object):

    def __init__(self, iface):
        self.iface = iface

    def initGui(self):

        self.action = QAction("Delete Feature", self.iface.mainWindow())
        self.action.triggered.connect(self.run)
        self.action.trigger()
        self.iface.mapCanvas().selectionChanged.connect(self.select_changed)

    def run(self):

        msg = QMessageBox()
        msg.setWindowTitle('DELETE Feature')
        msg.setText('Which type of feature do you want to delete?')
        job = msg.addButton('Job', QMessageBox.AcceptRole)
        plan = msg.addButton('Plan', QMessageBox.AcceptRole)
        supp = msg.addButton('Supplemental', QMessageBox.AcceptRole)
        cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
        msg.setDefaultButton(job)
        QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
        msg.exec_()
        msg.deleteLater()
        QGuiApplication.restoreOverrideCursor()

        if msg.clickedButton() is job:
            # do job stuff
            self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
            self.iface.setActiveLayer(self.vl)

            vLayer = self.iface.activeLayer()
            feats_count = vLayer.selectedFeatureCount()

            if feats_count == 0:
                msg = QMessageBox()
                msg.setWindowTitle('Selection')
                msg.setText('Choose the job you wish to delete and click OK to proceed.')
                cont = msg.addButton('Continue', QMessageBox.AcceptRole)
                cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
                msg.setDefaultButton(cont)
                msg.exec_()
                msg.deleteLater()
                if msg.clickedButton() is cont:
                    self.iface.actionSelect().trigger()
                else:
                    try:
                        self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                    except Exception as e:
                        pass
                    return
            else:
                msg = QMessageBox()
                msg.setWindowTitle('Selection')
                msg.setText(
                    str(feats_count) + ' feature(s) are selected. Do you want to delete the current selection '
                                       'or choose a new selection?')
                cont = msg.addButton('Continue', QMessageBox.AcceptRole)
                new = msg.addButton('New Selection', QMessageBox.AcceptRole)
                cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
                msg.setDefaultButton(cont)
                msg.exec_()
                msg.deleteLater()

                if msg.clickedButton() is cont:
                    try:
                        self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                    except Exception as e:
                        pass
                    self.select_changed()

                elif msg.clickedButton() is new:
                    msg = QMessageBox()
                    msg.setWindowTitle('Selection')
                    msg.setText('Choose the job you wish to delete and click OK to proceed.')
                    cont = msg.addButton('Continue', QMessageBox.AcceptRole)
                    cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
                    msg.setDefaultButton(cont)
                    msg.exec_()
                    msg.deleteLater()
                    if msg.clickedButton() is cont:
                        for a in self.iface.mainWindow().children():

                            if a.objectName() == 'mActionDeselectAll':
                                a.trigger()
                                QgsMessageLog.logMessage('FIRST RUN: Previous selection has been cleared.',
                                                         'sGIS', level=Qgis.Info)
                                QgsMessageLog.logMessage('Job deletion starting...', 'sGIS', level=Qgis.Info)
                                self.iface.actionSelect().trigger()
                    elif msg.clickedButton() is cancel:
                        try:
                            self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                        except Exception as e:
                            pass

                elif msg.clickedButton() is cancel:
                    QgsMessageLog.logMessage('DEBUG: Job deletion cancelled.', 'sGIS', level=Qgis.Info)
                    try:
                        self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                    except Exception as e:
                        pass
        elif msg.clickedButton() is plan:
            # do plan stuff
            self.vl = QgsProject.instance().mapLayersByName('plans')[0]
            self.iface.setActiveLayer(self.vl)

            vLayer = self.iface.activeLayer()
            feats_count = vLayer.selectedFeatureCount()

            if feats_count == 0:
                msg = QMessageBox()
                msg.setWindowTitle('Selection')
                msg.setText('Choose the plan you wish to delete and click OK to proceed.')
                cont = msg.addButton('Continue', QMessageBox.AcceptRole)
                cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
                msg.setDefaultButton(cont)
                msg.exec_()
                msg.deleteLater()
                if msg.clickedButton() is cont:
                    self.iface.actionSelect().trigger()
                else:
                    try:
                        self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                    except Exception as e:
                        pass
                    return
            else:
                msg = QMessageBox()
                msg.setWindowTitle('Selection')
                msg.setText(
                    str(feats_count) + ' feature(s) are selected. Do you want to delete the current selection '
                                       'or choose a selection?')
                cont = msg.addButton('Continue', QMessageBox.AcceptRole)
                new = msg.addButton('New Selection', QMessageBox.AcceptRole)
                cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
                msg.setDefaultButton(cont)
                msg.exec_()
                msg.deleteLater()

                if msg.clickedButton() is cont:
                    try:
                        self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                    except Exception as e:
                        pass
                    self.select_changed()

                elif msg.clickedButton() is new:
                    msg = QMessageBox()
                    msg.setWindowTitle('Selection')
                    msg.setText('Choose the plan you wish to delete and click OK to proceed.')
                    cont = msg.addButton('Continue', QMessageBox.AcceptRole)
                    cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
                    msg.setDefaultButton(cont)
                    msg.exec_()
                    msg.deleteLater()
                    if msg.clickedButton() is cont:
                        for a in self.iface.mainWindow().children():

                            if a.objectName() == 'mActionDeselectAll':
                                a.trigger()
                                QgsMessageLog.logMessage('FIRST RUN: Previous selection has been cleared.',
                                                         'sGIS', level=Qgis.Info)
                                QgsMessageLog.logMessage('Plan deletion starting...', 'sGIS', level=Qgis.Info)
                                self.iface.actionSelect().trigger()
                    elif msg.clickedButton() is cancel:
                        try:
                            self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                        except Exception as e:
                            pass

                elif msg.clickedButton() is cancel:
                    QgsMessageLog.logMessage('DEBUG: Plan deletion cancelled.', 'sGIS', level=Qgis.Info)
                    try:
                        self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                    except Exception as e:
                        pass
        elif msg.clickedButton() is supp:
            # do supp stuff
            self.vl = QgsProject.instance().mapLayersByName('supplementals')[0]
            self.iface.setActiveLayer(self.vl)

            vLayer = self.iface.activeLayer()
            feats_count = vLayer.selectedFeatureCount()

            if feats_count == 0:
                msg = QMessageBox()
                msg.setWindowTitle('Selection')
                msg.setText('Choose the supplemental you wish to delete and click OK to proceed.')
                cont = msg.addButton('Continue', QMessageBox.AcceptRole)
                cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
                msg.setDefaultButton(cont)
                msg.exec_()
                msg.deleteLater()
                if msg.clickedButton() is cont:
                    self.iface.actionSelect().trigger()
                else:
                    try:
                        self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                    except Exception as e:
                        pass
                    return
            else:
                msg = QMessageBox()
                msg.setWindowTitle('Selection')
                msg.setText(
                    str(feats_count) + ' feature(s) are selected. Do you want to delete the current selection '
                                       'or choose a selection?')
                cont = msg.addButton('Continue', QMessageBox.AcceptRole)
                new = msg.addButton('New Selection', QMessageBox.AcceptRole)
                cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
                msg.setDefaultButton(cont)
                msg.exec_()
                msg.deleteLater()

                if msg.clickedButton() is cont:
                    try:
                        self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                    except Exception as e:
                        pass
                    self.select_changed()

                elif msg.clickedButton() is new:
                    msg = QMessageBox()
                    msg.setWindowTitle('Selection')
                    msg.setText('Choose the supplemental you wish to delete and click OK to proceed.')
                    cont = msg.addButton('Continue', QMessageBox.AcceptRole)
                    cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
                    msg.setDefaultButton(cont)
                    msg.exec_()
                    msg.deleteLater()
                    if msg.clickedButton() is cont:
                        for a in self.iface.mainWindow().children():

                            if a.objectName() == 'mActionDeselectAll':
                                a.trigger()
                                QgsMessageLog.logMessage('FIRST RUN: Previous selection has been cleared.',
                                                         'sGIS', level=Qgis.Info)
                                QgsMessageLog.logMessage('Supplemental deletion starting...', 'sGIS', level=Qgis.Info)
                                self.iface.actionSelect().trigger()
                    elif msg.clickedButton() is cancel:
                        try:
                            self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                        except Exception as e:
                            pass

                elif msg.clickedButton() is cancel:
                    QgsMessageLog.logMessage('DEBUG: Supplemental deletion cancelled.', 'sGIS', level=Qgis.Info)
                    try:
                        self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                    except Exception as e:
                        pass
        elif msg.clickedButton() is cancel:
            QgsMessageLog.logMessage('DEBUG: Supplemental deletion cancelled.', 'sGIS', level=Qgis.Info)
            return

    def select_changed(self):
        try:
            self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
        except Exception as e:
            pass

        try:
            layerActive = self.iface.activeLayer()
            parcel = self.iface.activeLayer().selectedFeatures()[0]

            if layerActive.name() == 'jobs':
                jobNo = parcel["job_no"]
                msg = QMessageBox()
                msg.setWindowTitle('DELETE Job')
                msg.setText('JobNo: ' + jobNo + ' has been selected. Continue?\n\nNOTE: This action CANNOT be undone!')
                delete = msg.addButton('Delete Job', QMessageBox.AcceptRole)
                cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
                msg.setDefaultButton(delete)
                QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
                msg.exec_()
                msg.deleteLater()
                QGuiApplication.restoreOverrideCursor()

                if msg.clickedButton() is delete:
                    QgsMessageLog.logMessage('Deleting job...', 'sGIS', level=Qgis.Info)
                    self.iface.actionToggleEditing().trigger()
                    lyr = QgsProject.instance().mapLayersByName('jobs')[0]
                    self.iface.setActiveLayer(lyr)
                    lyr.deleteFeature(parcel.id())
                    self.iface.activeLayer().commitChanges()
                    self.iface.actionIdentify().trigger()
                    resetLegend(self)

                elif msg.clickedButton() is cancel:
                    QgsMessageLog.logMessage('DEBUG: Job deletion cancelled.', 'sGIS', level=Qgis.Info)
                    return
                else:
                        QgsMessageLog.logMessage('DEBUG: Job deletion cancelled.', 'sGIS', level=Qgis.Info)
            elif layerActive.name() == 'plans':
                planNo = parcel["plan_no"]
                msg = QMessageBox()
                msg.setWindowTitle('DELETE Plan')
                msg.setText('PlanNo: ' + planNo + ' has been selected. Continue?\n\nNOTE: This action CANNOT be undone!')
                delete = msg.addButton('Delete Plan', QMessageBox.AcceptRole)
                cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
                msg.setDefaultButton(delete)
                QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
                msg.exec_()
                msg.deleteLater()
                QGuiApplication.restoreOverrideCursor()

                if msg.clickedButton() is delete:
                    QgsMessageLog.logMessage('Deleting plan...', 'sGIS', level=Qgis.Info)
                    self.iface.actionToggleEditing().trigger()
                    lyr = QgsProject.instance().mapLayersByName('plans')[0]
                    self.iface.setActiveLayer(lyr)
                    lyr.deleteFeature(parcel.id())
                    self.iface.activeLayer().commitChanges()
                    self.iface.actionIdentify().trigger()
                    resetLegend(self)

                elif msg.clickedButton() is cancel:
                    QgsMessageLog.logMessage('DEBUG: Plan deletion cancelled.', 'sGIS', level=Qgis.Info)
                    return
                else:
                    QgsMessageLog.logMessage('DEBUG: Plan deletion cancelled.', 'sGIS', level=Qgis.Info)

            elif layerActive.name() == 'supplementals':
                suppNo = parcel["supp_no"]
                msg = QMessageBox()
                msg.setWindowTitle('DELETE Supplemental')
                msg.setText('SuppNo: ' + suppNo + ' has been selected. Continue?\n\nNOTE: This action CANNOT be undone!')
                delete = msg.addButton('Delete Supplemental', QMessageBox.AcceptRole)
                cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
                msg.setDefaultButton(delete)
                QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
                msg.exec_()
                msg.deleteLater()
                QGuiApplication.restoreOverrideCursor()

                if msg.clickedButton() is delete:
                    QgsMessageLog.logMessage('Deleting supplemental...', 'sGIS', level=Qgis.Info)
                    self.iface.actionToggleEditing().trigger()
                    lyr = QgsProject.instance().mapLayersByName('supplementals')[0]
                    self.iface.setActiveLayer(lyr)
                    lyr.deleteFeature(parcel.id())
                    self.iface.activeLayer().commitChanges()
                    self.iface.actionIdentify().trigger()
                    resetLegend(self)

                elif msg.clickedButton() is cancel:
                    QgsMessageLog.logMessage('DEBUG: Supp deletion cancelled.', 'sGIS', level=Qgis.Info)
                    return
                else:
                    QgsMessageLog.logMessage('DEBUG: Supp deletion cancelled.', 'sGIS', level=Qgis.Info)
        except Exception as e:
            QgsMessageLog.logMessage('EXCEPTION: ' + str(e), 'sGIS', level=Qgis.Info)
            resetLegend(self)


class sgis_abutters(object):
    newJob = 0
    selComp = 0
    multiFeat = 0
    count = 0

    def __init__(self, iface):
        self.iface = iface

    def run(self):

        self.set_abutters()

    def set_abutters(self):

        vLayer = self.iface.activeLayer()
        self.iface.actionCopyFeatures().trigger()
        self.newJob = 0
        self.fields = self.iface.activeLayer().fields()

        self.tmpLayer = QgsVectorLayer('MultiPolygon?crs=EPSG:102683', 'tmp_buffer', 'memory')
        QgsProject.instance().addMapLayers([self.tmpLayer])
        self.iface.actionIdentify().trigger()
        self.vl = QgsProject.instance().mapLayersByName('tmp_buffer')[0]
        self.iface.setActiveLayer(self.vl)

        self.iface.actionToggleEditing().trigger()
        self.layerData = self.tmpLayer.dataProvider()
        self.layerData.addAttributes(self.fields)
        self.iface.activeLayer().commitChanges()

        self.iface.actionToggleEditing().trigger()
        self.iface.actionPasteFeatures().trigger()
        self.iface.activeLayer().commitChanges()

        Buffer_only_selected_features = True

        if vLayer.selectedFeatures() and Buffer_only_selected_features is True:
            features = vLayer.selectedFeatures()

        else:
            features = vLayer.selectedFeatures()

        lyr = self.iface.activeLayer()

        feat = self.iface.activeLayer().selectedFeatures()[0]

        buff = feat.geometry().buffer(35, 5)
        lyr.dataProvider().changeGeometryValues({feat.id(): buff})

        self.layerData = lyr.dataProvider()
        self.iface.actionToggleEditing().trigger()
        self.layerData.addAttributes([QgsField("referrer", QVariant.String)])
        self.layerData.addAttributes([QgsField("referrerj", QVariant.String)])
        self.layerData.addAttributes([QgsField("mailingaddress", QVariant.String)])
        self.layerData.addAttributes([QgsField("bookpage", QVariant.String)])
        self.iface.activeLayer().commitChanges()

        self.vl = QgsProject.instance().mapLayersByName('tmp_buffer')[0]
        self.iface.setActiveLayer(self.vl)

        lyr = self.iface.activeLayer()

        self.iface.actionToggleEditing().trigger()
        idx = self.layerData.fieldNameIndex('referrer')
        idx2 = self.layerData.fieldNameIndex('referrerj')
        idx3 = self.layerData.fieldNameIndex('objectid')

        feat = self.iface.activeLayer().selectedFeatures()[0]

        lyr.changeAttributeValue(feat.id(), idx, feat['sid'])
        lyr.changeAttributeValue(feat.id(), idx2, feat['job_no'])
        lyr.changeAttributeValue(feat.id(), idx3, feat['sid'])
        lyr.updateFields()
        self.iface.activeLayer().commitChanges()

        feat = self.iface.activeLayer().selectedFeatures()[0]
        referrer = feat['referrer']
        referrerJ = feat['job_no']
        # mailingaddress = feat['mailingaddress']
        self.iface.actionToggleEditing().trigger()
        field_ids = []
        fieldnames = {'objectid', 'map_bk_lot', 'town', 'county', 'prop_loc', 'referrer', 'referrerj', 'owner1',
                      'addressee', 'own_street1', 'own_street2', 'own_city', 'own_state', 'own_zip', 'locusaddress',
                      'mailingaddress', 'rawdeeds', 'formattedaddress', 'geocode', 'state_id'}

        for field in lyr.fields():
            if field.name() not in fieldnames:
                idx = self.layerData.fieldNameIndex(field.name())
                field_ids.append(idx)

        # self.layerData.deleteAttributes(field_ids)
        self.tmpLayer.updateFields()
        self.iface.activeLayer().commitChanges()

        layer1 = QgsProject.instance().mapLayersByName('Parcels')[0]
        layer2 = QgsProject.instance().mapLayersByName('tmp_buffer')[0]
        layer3 = QgsProject.instance().mapLayersByName('abutters')[0]
        
        self.fb = QgsProcessingFeedback()
        self.context = QgsProcessingContext

        QgsMessageLog.logMessage('START SBL...', 'sGIS', level=Qgis.Info)
        processing.run("native:selectbylocation", {
            'INPUT': layer1,
            'PREDICATE': [0],
            'INTERSECT': layer2, 'METHOD': 0}, feedback=self.fb)

        QgsMessageLog.logMessage('END SBL...', 'sGIS', level=Qgis.Info)

        self.iface.setActiveLayer(layer1)
        self.iface.actionCopyFeatures().trigger()
        self.iface.setActiveLayer(layer3)
        self.iface.actionToggleEditing().trigger()
        self.iface.actionPasteFeatures().trigger()
        self.iface.activeLayer().commitChanges()

        cLayer = self.iface.activeLayer()

        self.iface.actionToggleEditing().trigger()
        for f in cLayer.selectedFeatures():

            self.layerData = self.iface.activeLayer().dataProvider()
            idx = self.layerData.fieldNameIndex('referrer')
            idx2 = self.layerData.fieldNameIndex('referrerj')

            if str(f['referrer']) == 'NULL':
                self.iface.activeLayer().changeAttributeValue(f.id(), idx,
                                                              str(referrer), True)
            if str(f['referrerj']) == 'NULL':
                self.iface.activeLayer().changeAttributeValue(f.id(), idx2,
                                                              str(referrerJ), True)
            if str(f['objectid']) == referrer:
                cLayer.deleteFeature(f.id())

            else:
                pass

        self.iface.activeLayer().commitChanges()
        QgsProject.instance().removeMapLayer(layer2.id())
        self.iface.activeLayer().removeSelection()

        self.vl = QgsProject.instance().mapLayersByName('Parcels')[0]
        self.iface.setActiveLayer(self.vl)
        self.iface.activeLayer().removeSelection()

        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(self.vl)
        canvas = self.iface.mapCanvas()
        canvas.zoomToSelected(self.vl)

        ex = canvas.extent()
        ex.scale(2.0)
        canvas.setExtent(ex)
        canvas.refresh()
        resetLegend(self)
        self.iface.messageBar().clearWidgets()

        return


class sgis_pgTEST(object):

    def __init__(self, iface):
        self.iface = iface

    def run(self):

        cLayer = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(cLayer)
        # self.get_credentials(cLayer)
        conns = self.getConnections()
        QgsMessageLog.logMessage('conns: ' + str(conns), 'sGIS', level=Qgis.Info)
        self.getSQL()

        resetLegend(self)

    def get_credentials(self, lyr):
        connInfo = QgsDataSourceUri(lyr.dataProvider().dataSourceUri()).connectionInfo()
        (success, user, passwd) = QgsCredentials.instance().get(connInfo, None, None)
        # Put the credentials back (for yourself and the provider), as QGIS removes it when you "get" it
        if success:
            QgsCredentials.instance().put(connInfo, user, passwd)
            QgsMessageLog.logMessage('user: ' + str(user) + ' | pass: ' + str(passwd), 'sGIS', level=Qgis.Info)

    def getSQL(self):
        base64_pw = dbPass
        base64_bytes = base64_pw.encode('ascii')
        pw_bytes = base64.b64decode(base64_bytes)
        pw = pw_bytes.decode('ascii')
        # QgsMessageLog.logMessage('user: postgres | pass: ' + str(pw), 'sGIS', level=Qgis.Info)

    def getConnections(self):
        s = QSettings()
        s.beginGroup("PostgreSQL/connections")
        currentConnections = s.childGroups()
        s.endGroup()
        return currentConnections

    def setConnection(self,conn):
        s = QSettings()
        s.beginGroup("PostgreSQL/connections/"+conn)
        currentKeys = s.childKeys()
        # print "keys: ", currentKeys
        QgsMessageLog.logMessage('keys: ' + str(currentKeys), 'sGIS', level=Qgis.Info)
        PSQLDatabase=s.value("database", "" )
        PSQLHost=s.value("host", "" )
        PSQLUsername=s.value("username", "" )
        PSQLPassword=s.value("password", "" )
        PSQLPort=s.value("port", "" )
        PSQLService=s.value("service", "" )
        s.endGroup()
        self.db = QSqlDatabase.addDatabase("QPSQL")
        self.db.setHostName(PSQLHost)
        self.db.setDatabaseName(PSQLDatabase)
        self.db.setUserName(PSQLUsername)
        self.db.setPassword(PSQLPassword)
        ok = self.db.open()
        if not ok:
            error = "Database Error: %s" % self.db.lastError().text()
        else:
            error=""


class sgis_printFolderLabel(object):
    def __init__(self, iface):
        # save reference to the QGIS interface
        self.iface = iface

    def __call__(self):
        self.action = QAction("PFL", self.iface.mainWindow())
        self.action.triggered.connect(self.run)
        self.action.trigger()

    def initGui(self):

        icon = QIcon(os.path.dirname(__file__) + "/icons/sgis_voronoi.png")

        self.action = QAction(icon, "PFL", self.iface.mainWindow())
        self.action.triggered.connect(self.run)
        self.action.trigger()
        self.iface.removeToolBarIcon(self.action)

    def run(self):
        import datetime
        year = datetime.datetime.today().strftime('%Y')
        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(self.vl)

        try:
            attribs = self.iface.activeLayer().selectedFeatures()[0]
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "No Selection",
                                 "Please ensure that you have a parcel selected\nand attempt to "
                                 "generate the output again.\n\n"
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))
            return

        clientLast = attribs["folder_name"]
        clientFirst = attribs["folder_name"]
        clientName = attribs["client_name"]
        folderName = attribs["folder_name"]
        folderType = attribs["folder_type"]
        primaryContact = attribs["primary_contact"]
        jobNo = attribs["job_no"]
        jobYear = '20' + jobNo[:2]

        if jobYear == year:
            year = year
        else:
            year = jobYear

        addr = attribs["locus_addr"]
        if str(addr) == 'NULL':
            addr = ''
        else:
            addr = addr

        town = attribs["town"]

        try:

            map_bk_lot = attribs["map_bk_lot"]
            mbl = map_bk_lot.split('-')
            mbLen = len(mbl)

            if mbLen == 1:
                map_bk_lot = map_bk_lot
            elif mbLen == 2:
                map_bk_lot = 'Map ' + mbl[0].lstrip('0') + ', Lot ' + mbl[1].lstrip('0')
            elif mbLen == 3:
                map_bk_lot = 'Map ' + mbl[0].lstrip('0') + ', Lot ' + mbl[1].lstrip('0') + '-' + mbl[2].lstrip('0')

        except Exception as e:
            map_bk_lot = 'N/A'

        jobType = attribs["job_type"]
        jobSubType = attribs['jobSubtype']
        
        if str(jobSubType) == 'NULL':
            jobSubType = ''
        else:
            jobSubType = ' | ' + jobSubType
        
        date_due = attribs["date_due"]

        jobFolders(jobNo)
 
        path = os.path.join(jobsPath, year, jobNo)  # need to programattically grab year
        jipath = os.path.join(path, "Job_Info")

        from openpyxl import load_workbook

        fPath = os.path.dirname(os.path.abspath(__file__)) + '\\GIS_templates.xlsx'
        wb = load_workbook(fPath)

        for s in range(len(wb.sheetnames)):
            if wb.sheetnames[s] == 'face':
                break
        wb.active = s

        sheets = wb.sheetnames

        for s in sheets:

            if s != 'face':
                sheet_name = wb[s]
                wb.remove(sheet_name)
        ws = wb.active

        try:
            ws['A1'] = folderName
            ws['A2'] = ''
            ws['B1'] = folderType
            if str(folderName) != str(clientName):
                ws['A2'] = clientName
                if str(clientName) != str(primaryContact):
                    ws['A2'] = 'c/o: ' + primaryContact
                    ws['A6'] = addr
                    ws['A7'] = town.upper()
                    ws['A10'] = map_bk_lot
                    ws['A12'] = jobType + jobSubType
                    ws['A15'] = jobNo
                    if str(date_due) == '1900-01-01':
                        ws['A20'] = ''
                    else:
                        ws['A20'] = date_due
                else:
                    ws['A3'] = ''
                    ws['A6'] = addr
                    ws['A7'] = town.upper()
                    ws['A10'] = map_bk_lot
                    ws['A12'] = jobType + jobSubType
                    ws['A15'] = jobNo
                    if str(date_due) == '1900-01-01':
                        ws['A20'] = ''
                    else:
                        ws['A20'] = date_due
                if str(clientName) != str(primaryContact):
                    ws['A3'] = 'c/o: ' + primaryContact
                else:
                    ws['A3'] = ''
            else:
                if str(clientName) != str(primaryContact):
                    ws['A2'] = 'c/o: ' + primaryContact
                    ws['A6'] = addr
                    ws['A7'] = town.upper()
                    ws['A10'] = map_bk_lot
                    ws['A12'] = jobType + jobSubType
                    ws['A15'] = jobNo
                    if str(date_due) == '1900-01-01':
                        ws['A20'] = ''
                    else:
                        ws['A20'] = date_due
                else:
                    ws['A3'] = ''
                    ws['A6'] = addr
                    ws['A7'] = town.upper()
                    ws['A10'] = map_bk_lot
                    ws['A12'] = jobType + jobSubType
                    ws['A15'] = jobNo
                    if str(date_due) == '1900-01-01':
                        ws['A20'] = ''
                    else:
                        ws['A20'] = date_due
                        
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                 "Required input missing: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))
            return

        facefile = str(jipath) + "\\" + jobNo + "_FolderFaceLabel_" + datetime.datetime.today().strftime(
            '%Y.%m.%d') + ".xlsx"
        QgsMessageLog.logMessage('Saving file: ' + facefile + '...', 'sGIS', level=Qgis.Info)
        wb.save(facefile)
        resetLegend(self)


class sgis_printYellowSheet(object):
    def __init__(self, iface):
        # save reference to the QGIS interface
        self.iface = iface

    def initGui(self):

        self.action = QAction("PYS", self.iface.mainWindow())
        self.action.triggered.connect(self.run)
        self.action.trigger()

    def run(self):
        import datetime
        year = datetime.datetime.today().strftime('%Y')
        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(self.vl)

        try:
            attribs = self.iface.activeLayer().selectedFeatures()[0]
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "No Selection",
                                 "Please ensure that you have a parcel selected\nand attempt to "
                                 "generate the output again.\n\n"
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))
            return

        date_requested = attribs["date_requested"]
        date_due = attribs["date_due"]
        jobNo = attribs["job_no"]
        revNo = attribs["rev_no"]
        folder_name = attribs["folder_name"]
        mail_addr = attribs["contact_addr"]
        locus_addr = attribs["locus_addr"]
        phone_mobile = attribs["phone_mobile"]
        phone_home = attribs["phone_home"]
        phone_work = attribs["phone_work"]
        extension = attribs["extension"]
        email_primary = attribs["email_primary"]
        email_secondary = attribs["email_secondary"]
        lowtide = attribs["lowtide"]
        zoning_district = attribs["zdistrict"]
        jobYear = '20' + jobNo[:2]

        if jobYear == year:
            year = year
        else:
            year = jobYear

        if lowtide == 't':
            lowtide = 'YES'
        else:
            lowtide = ''

        lowtide_hrs = attribs["lowtide_hrs"]

        if lowtide == 'YES':
            pass
        else:
            lowtide_hrs = ''

        town = attribs["town"]
        try:
            map_bk_lot = attribs["map_bk_lot"]
            mbl = map_bk_lot.split('-')
            mbLen = len(mbl)

            if mbLen == 1:
                map_bk_lot = map_bk_lot
            elif mbLen == 2:
                map_bk_lot = 'Map ' + mbl[0].lstrip('0') + ', Lot ' + mbl[1].lstrip('0')
            elif mbLen == 3:
                map_bk_lot = 'Map ' + mbl[0].lstrip('0') + ', Lot ' + mbl[1].lstrip('0') + '-' + mbl[2].lstrip('0')

        except Exception as e:
            map_bk_lot = 'N/A'

        jobType = attribs["job_type"]
        jobSubType = attribs["jobSubtype"]

        if str(jobSubType) == 'NULL':
            jobSubType = ''
        else:
            jobSubType = jobSubType

        folderName = attribs["folder_name"]
        clientName = attribs["client_name"]
        folderType = attribs["folder_type"]
        clientRole = attribs["client_role"]
        primaryContact = attribs["primary_contact"]

        rate_rs = ''
        rate_cad = ''
        rate_fw = ''
        rate_misc = ''

        date_recorded = attribs["date_recorded"]
        recorded_by = attribs["recorded_by"]
        planbook_page = attribs["planbook_page"]

        job_desc = attribs["job_desc"]

        path = os.path.join(jobsPath, year, jobNo)
        jipath = os.path.join(path, "Job_Info")

        if not os.path.exists(path):
            os.makedirs(path)

        if not os.path.exists(jipath):
            os.makedirs(jipath)

        from openpyxl import load_workbook

        fPath = os.path.dirname(os.path.abspath(__file__)) + '\\GIS_templates.xlsx'
        wb = load_workbook(fPath)

        sheets = wb.sheetnames
        for s in sheets:

            if s != 'yellow':
                sheet_name = wb.get_sheet_by_name(s)
                wb.remove_sheet(sheet_name)

        for s in range(len(wb.sheetnames)):
            if wb.sheetnames[s] == 'yellow':
                break

        wb.active = s

        try:
            ws2 = wb.active
            ws2['B1'] = date_requested
            ws2['I1'] = jobNo
            ws2['I2'] = revNo

            if str(clientName) == str(folderName):
                pass
            else:
                from openpyxl.styles import PatternFill
            
            ws2['B4'] = folderName + ' (' + folderType + ')'
            ws2['B5'] = clientName
            ws2['B6'] = locus_addr
            ws2['F4'] = mail_addr
            ws2['B7'] = town
            ws2['B8'] = map_bk_lot
            ws2['F6'] = phone_mobile
            ws2['B9'] = clientRole
            ws2['B10'] = folderType
            ws2['C1'] = jobType
            ws2['C3'] = jobSubType
            ws2['B11'] = lowtide
            ws2['B12'] = lowtide_hrs
            ws2['C18'] = rate_rs
            ws2['C17'] = rate_cad
            ws2['C16'] = rate_fw
            ws2['C19'] = rate_misc

            if str(email_secondary) == 'NULL':
                ws2['F10'] = ''
            else:
                ws2['F10'] = email_secondary
            if str(email_primary) == 'NULL':
                ws2['F9'] = ''
            else:
                ws2['F9'] = email_primary
            if str(phone_work) == 'NULL':
                ws2['F7'] = ''
            else:
                ws2['F7'] = str(phone_work) + ' ' + str(extension)
            if str(phone_home) == 'NULL':
                ws2['F8'] = ''
            else:
                ws2['F8'] = phone_home
            if str(date_recorded) == '1900-01-01':
                ws2['B32'] = ''
            else:
                ws2['B32'] = date_recorded
            if str(recorded_by) == 'NULL':
                ws2['B33'] = ''
            else:
                ws2['B33'] = recorded_by
            if str(planbook_page) == 'NULL':
                ws2['B34'] = ''
            else:
                ws2['B34'] = planbook_page
            if str(zoning_district) == 'NULL':
                ws2['B36'] = ''
            else:
                ws2['B36'] = zoning_district
            if str(date_due) == '1900-01-01':
                ws2['B2'] = ''
            else:
                ws2['B2'] = date_due

            if str(job_desc) == 'NULL':
                ws2['B40'] = ''
            else:
                ws2['A39'] = job_desc

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                 "REQUIRED FIELDS ARE EMPTY - check job and try again.\n\nDetails: " + str(
                                     exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))
            return

        yellowfile = str(jipath) + "\\" + jobNo + "_YellowSheet_" + datetime.datetime.today().strftime(
            '%Y.%m.%d') + ".xlsx"
        QgsMessageLog.logMessage('Saving file: ' + yellowfile + '...', 'sGIS', level=Qgis.Info)
        wb.save(yellowfile)
        resetLegend(self)


class sgis_label_dialog(QDialog, Ui_sgis_label_form):
    dValue = 4

    def __init__(self, iface):
        QDialog.__init__(self)
        self.iface = iface
        self.setupUi(self)

        buttonBox = self.buttonBox

        self.dValue = self.dial.value()
        buttonBox.accepted.connect(partial(self.print_label, self.dValue))
        buttonBox.rejected.connect(self.finished)

    def dv(self):
        self.dValue = self.dial.value()
        QgsMessageLog.logMessage('Dial Value : ' + str(self.dValue) + '...', 'sGIS', level=Qgis.Info)

    def print_label(self, dv):
        self.dValue = self.dial.value()
        import datetime
        dv = self.dValue
        QgsMessageLog.logMessage('Labels Used: ' + str(dv) + '...', 'sGIS', level=Qgis.Info)

        year = datetime.datetime.today().strftime('%Y')
        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(self.vl)
        attribs = self.iface.activeLayer().selectedFeatures()[0]
        clientLast = attribs["client_name"]
        clientFirst = attribs["client_name"]

        folderName = attribs["folder_name"]
        folderType = attribs["folder_type"]
        jobNo = attribs["job_no"]
        jobYear = '20' + jobNo[:2]

        if jobYear == year:
            year = year
        else:
            year = jobYear

        addr = attribs["locus_addr"]
        town = attribs["town"]

        try:

            map_bk_lot = attribs["map_bk_lot"]
            mbl = map_bk_lot.split('-')
            mbLen = len(mbl)

            if mbLen == 1:
                map_bk_lot = map_bk_lot
            elif mbLen == 2:
                map_bk_lot = 'Map ' + mbl[0].lstrip('0') + ', Lot ' + mbl[1].lstrip('0')
            elif mbLen == 3:
                map_bk_lot = 'Map ' + mbl[0].lstrip('0') + ', Lot ' + mbl[1].lstrip('0') + '-' + mbl[2].lstrip('0')

        except Exception as e:
            map_bk_lot = 'N/A'

        jobType = attribs["job_type"]
        jobSubType = attribs["jobSubtype"]


        path = os.path.join(jobsPath, year, jobNo)
        jipath = os.path.join(path, "Job_Info")
        QgsMessageLog.logMessage('Checking output folder: ' + path + '...', 'sGIS', level=Qgis.Info)
        if not os.path.exists(path):
            os.makedirs(path)
        QgsMessageLog.logMessage('Checking output folder: ' + jipath + '...', 'sGIS', level=Qgis.Info)
        if not os.path.exists(jipath):
            os.makedirs(jipath)

        from openpyxl import load_workbook

        fPath = os.path.dirname(os.path.abspath(__file__)) + '\\GIS_templates.xlsx'
        wb = load_workbook(fPath)
        sheets = wb.sheetnames
        for s in sheets:

            if s != 'label':
                sheet_name = wb.get_sheet_by_name(s)
                wb.remove_sheet(sheet_name)

        for s in range(len(wb.sheetnames)):
            if wb.sheetnames[s] == 'label':
                break

        wb.active = s
        ws3 = wb.active

        cv1 = folderName + '   ' + map_bk_lot
        cv2 = addr + ", " + town + "     " + jobType  # + " | " + jobSubType

        if dv == 0:
            ws3['A19'] = cv1
            ws3['A20'] = cv2
        elif dv == 1:
            ws3['A16'] = cv1
            ws3['A17'] = cv2
        elif dv == 2:
            ws3['A13'] = cv1
            ws3['A14'] = cv2
        elif dv == 3:
            ws3['A10'] = cv1
            ws3['A11'] = cv2
        elif dv == 4:
            ws3['A7'] = cv1
            ws3['A8'] = cv2
        elif dv == 5:
            ws3['A4'] = cv1
            ws3['A5'] = cv2
        elif dv == 6:
            ws3['A1'] = cv1
            ws3['A2'] = cv2
        elif dv == 7:
            QMessageBox.critical(self.iface.mainWindow(), "HEY, GENIUS...",
                                 "Your sheet of labels is empty - get a new one.  Derp.")
            return
        else:
            pass

        labelfile = str(jipath) + "\\" + jobNo + "_FolderLabel_" + datetime.datetime.today().strftime(
            '%Y.%m.%d') + ".xlsx"
        QgsMessageLog.logMessage('Saving file: ' + labelfile + '...', 'sGIS', level=Qgis.Info)
        wb.save(labelfile)
        resetLegend(self)

    def finished(self, **kwargs):
        self.done(1)


class sgis_printMapTable(object):

    def __init__(self, iface):
        self.iface = iface

    def initGui(self):

        self.action = QAction("PMT", self.iface.mainWindow())
        self.action.triggered.connect(self.run)
        self.action.trigger()

    def run(self):

        cfg0 = 0
        cfg1 = 1

        layerActive = self.iface.activeLayer()

        try:
            feat = self.iface.activeLayer().selectedFeatures()[0]

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "No Selection",
                                 "Please ensure that you have a parcel selected\nand attempt to "
                                 "generate the output again.\n\n"
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))
            return

        fid = feat.id()
        key = str(feat['map_bk_lot'])
        if key == 'NULL':
            key = str(feat['sid'])
        
        self.iface.setActiveLayer(layerActive)

        self.getRelatedWork(feat, cfg0)

        if layerActive.name() == 'jobs':
            import datetime
            relW = self.updateJobRelated(feat)
            year = datetime.datetime.today().strftime('%Y')

            try:
                attribs = self.iface.activeLayer().selectedFeatures()[0]
            except Exception as e:
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                QMessageBox.critical(self.iface.mainWindow(), "NO SELECTION!",
                                     "Please ensure that you have a parcel selected\nand attempt to "
                                     "generate the output again.\n\n"
                                     "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                         exc_tb.tb_lineno) + ' ' + str(e))
                return

            jobNo = attribs["job_no"]
            jobYear = '20' + jobNo[:2]

            if jobYear == year:
                year = year
            else:
                year = jobYear

            path = os.path.join(jobsPath, year, jobNo)
            jipath = os.path.join(path, "Job_Info")
            if not os.path.exists(path):
                os.makedirs(path)
            if not os.path.exists(jipath):
                os.makedirs(jipath)

            from openpyxl import load_workbook

            fPath = os.path.dirname(os.path.abspath(__file__)) + '\\GIS_templates.xlsx'
            wb = load_workbook(fPath)

            for s in range(len(wb.sheetnames)):
                if wb.sheetnames[s] == 'maptable':
                    break
            wb.active = s
            sheets = wb.sheetnames
            ws = wb.active

            clientName = attribs["client_name"]
            folderName = attribs["folder_name"]
            addr = attribs["locus_addr"]
            town = attribs["town"]

            try:

                map_bk_lot = attribs["map_bk_lot"]
                ogMap = map_bk_lot
                map_bk_lotO = attribs["map_bk_lot"]
                mbl = map_bk_lot.split('-')
                mbLen = len(mbl)

                # NEED TO HANDLE MORE THAN 3 SECTIONS OF MBL
                if mbLen == 1:
                    map_bk_lot = map_bk_lot
                elif mbLen == 2:
                    map_bk_lot = 'Map ' + mbl[0].lstrip('0') + ', Lot ' + mbl[1].lstrip('0')
                elif mbLen == 3:
                    map_bk_lot = 'Map ' + mbl[0].lstrip('0') + ', Lot ' + mbl[1].lstrip('0') + '-' + mbl[2].lstrip('0')

            except Exception as e:
                map_bk_lot = 'N/A'

            jobType = attribs["job_type"]
            jobSubType = attribs["jobSubtype"]

            if str(jobSubType) == 'NULL':
                jobSubType = ''
            else:
                jobSubType = jobSubType

            revNo = attribs["rev_no"]
            county = attribs["county"]
            state = attribs["state"]
            perimeter = attribs["sPerimeter"]
            area = attribs["area"]
            planbook_page = attribs["planbook_page"]
            referrerJ = attribs['job_no']
            zipCode = attribs['zipcode']

            if str(zipCode) == 'NULL':
                zipCode = ''
            else:
                zipCode = zipCode

            lat_lon = attribs['lat_lon']

            if str(clientName) == 'NULL':
                clientName = 'N/A'
            else:
                clientName = clientName

            if str(county) == 'NULL':
                county = 'N/A'
            else:
                county = county

            try:
                ws['A1'] = folderName
                ws['D2'] = town.upper()
                ws['E2'] = county
                ws['F2'] = state
                ws['A3'] = 'Job#: ' + str(jobNo)
                ws['B3'] = 'Rev#: ' + str(revNo)
                ws['C3'] = 'Type: ' + str(jobType)  # + ' | ' + str(jobSubType)
                ws['D3'] = 'Perimeter: ' + str(perimeter) + ' L.Ft'
                ws['E3'] = 'Area: ' + str(area) + ' Ac.'
                ws['F3'] = 'Centroid: ' + str(lat_lon)
                ws['D4'] = zipCode
                ws['A6'] = map_bk_lot
                ws['B6'] = addr
                ws['E6'] = folderName
                ws['B9'] = str(relW)

                layer3 = QgsProject.instance().mapLayersByName('abutters')[0]
                exp = QgsExpression(u'"referrerj" = \'%s\'' % (jobNo))
                request = QgsFeatureRequest(exp)
                request.setSubsetOfAttributes(['referrerj'], layer3.fields())
                request.setFlags(QgsFeatureRequest.NoGeometry)

                aNo = 0
                startCell = 13
                startCellb = 14
                startCellp = 15
                startCelld = 16

                for f in layer3.getFeatures(request):

                    if str(f['map_bk_lot']) == key:
                        owner1 = str(f['owner1'])
                        if owner1 == 'NULL':
                            owner1 = ''

                        formattedaddress = str(f['formattedaddress'])
                        if formattedaddress == 'NULL':
                            formattedaddress = ''
                        locusaddress = str(f['locusaddress'])
                        
                        if locusaddress == 'NULL':
                            locusaddress = ''

                        ownInfo = 'OWNER: ' + owner1 # + ' | BOOKPAGE: ' + str(f['bookpage'])
                        formattedaddress = 'MAIL: ' + formattedaddress
                        alldeeds = str(f['rawdeeds'])
                        if alldeeds == 'NULL':
                            alldeeds = ''

                        ws['A7'] = locusaddress
                        if len(ownInfo) >= 150:
                            ws.row_dimensions[7].height = 45
                            ws['B10'].alignment = Alignment(horizontal='left', vertical='center', text_rotation=0,
                                                    wrap_text=True, shrink_to_fit=False, indent=0)
                        ws['B7'] = ownInfo
                        if len(formattedaddress) >= 150:
                            ws.row_dimensions[10].height = 45
                            ws['B8'].alignment = Alignment(horizontal='left', vertical='center', text_rotation=0,
                                                    wrap_text=True, shrink_to_fit=False, indent=0)
                        ws['B8'] = formattedaddress
                        if len(alldeeds) >= 150:
                            ws.row_dimensions[10].height = 45
                            ws['B10'].alignment = Alignment(horizontal='left', vertical='center', text_rotation=0,
                                                    wrap_text=True, shrink_to_fit=False, indent=0)
                        ws['A10'] = 'All Deeds:'
                        ws['B10'] = alldeeds

                        pass
                    else:

                        QgsMessageLog.logMessage('abutter found: ' + str(f['map_bk_lot']), 'sGIS', level=Qgis.Info)

                        aNo += 1
                        c1 = 'A' + str(startCell)
                        c2 = 'B' + str(startCell)
                        c3 = 'E' + str(startCell)
                        c4 = 'B' + str(startCellp)
                        c5 = 'B' + str(startCellb)
                        c6 = 'A' + str(startCellb)
                        c7 = 'A' + str(startCelld)
                        c8 = 'B' + str(startCelld)

                        map_bk_lotP = f['map_bk_lot']
                        mbl = map_bk_lotP.split('-')
                        mbLen = len(mbl)

                        if mbLen == 1:
                            map_bk_lotP = map_bk_lotP
                        elif mbLen == 2:
                            map_bk_lotP = 'Map ' + mbl[0].lstrip('0') + ', Lot ' + mbl[1].lstrip('0')
                        elif mbLen == 3:
                            map_bk_lotP = 'Map ' + mbl[0].lstrip('0') + ', Lot ' + mbl[1].lstrip('0') + '-' + mbl[2].lstrip(
                                '0')

                        ws[c1] = str(map_bk_lotP)

                        if str(f['locusaddress']) == 'NULL':
                            ownInfo = ''
                            bookpage = ''
                            formattedaddress = ''
                            locusaddress = ''
                            alldeeds = ''
                            if str(f['bookpage']) == 'NULL':
                                bookpage = ''
                                formattedaddress = ''
                                locusaddress = ''
                                alldeeds = ''
                            else:
                                pass
                        else:
                            owner1 = str(f['owner1'])
                            if owner1 == 'NULL':
                                owner1 = ''

                            ownInfo = 'OWNER: ' + owner1  # + ' | BOOKPAGE: ' + str(f['bookpage'])
                            formattedaddress = 'MAIL: ' + str(f['formattedaddress'])
                            locusaddress = str(f['locusaddress'])
                            alldeeds = str(f['rawdeeds'])
                            # bookpage = str(f['bookpage'])
                        try:
                            if len(ownInfo) > 0:
                                if len(ownInfo) >= 115:
                                    ws.row_dimensions[startCell].height = 30
                                    ws[c2].alignment = Alignment(horizontal='left', vertical='center',
                                                                    text_rotation=0, wrap_text=True,
                                                                 shrink_to_fit=False, indent=0)
                                    ws.merge_cells(start_row=startCell, start_column=2, end_row=startCell, end_column=8)
                                if len(formattedaddress) >= 115:
                                    ws.row_dimensions[startCellb].height = 30
                                    ws[c5].alignment = Alignment(horizontal='left', vertical='center',
                                                                    text_rotation=0, wrap_text=True,
                                                                 shrink_to_fit=False, indent=0)
                                    ws.merge_cells(start_row=startCellb, start_column=2, end_row=startCellb, end_column=8)
                                ws[c2] = ownInfo
                                ws[c7] = 'All Deeds:'
                                if len(alldeeds) >= 115:
                                    ws.row_dimensions[startCelld].height = 30
                                    ws[c5].alignment = Alignment(horizontal='left', vertical='center',
                                                                    text_rotation=0,
                                                                    wrap_text=True, shrink_to_fit=False, indent=0)
                                    ws.merge_cells(start_row=startCelld, start_column=2, end_row=startCelld, end_column=8)
                                ws[c8] = alldeeds

                            else:
                                ws[c2] = ''
                        except Exception as e:
                            ws[c2] = ''
                            exc_type, exc_obj, exc_tb = sys.exc_info()
                            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                            QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION!",
                                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                                     exc_tb.tb_lineno) + ' ' + str(e))
                            return

                        layer3.selectByIds([f.id()])
                        self.iface.setActiveLayer(layer3)
                        sA = self.iface.activeLayer().selectedFeatures()[0]
                        oid = sA['gid']
                        QgsMessageLog.logMessage('ABUTTER SELECTED: ' + str(oid), 'sGIS', level=Qgis.Info)

                        # try:
                        self.getRelatedWork(sA, cfg0)
                        jobNo = str(jobNo)
                        # QgsMessageLog.logMessage('jobNo: ' + str(jobNo), 'sGIS', level=Qgis.Info)
                        relAW = self.updateAbutterRelated(sA, jobNo)

                        QgsMessageLog.logMessage('RELATED AW: ' + str(relAW), 'sGIS', level=Qgis.Info)
                        if len(relAW) >= 115:
                            ws.row_dimensions[startCellp].height = 30
                            ws[c4].alignment = Alignment(horizontal='left', vertical='center',
                                                         text_rotation=0,
                                                         wrap_text=True, shrink_to_fit=False, indent=0)
                        ws[c4] = relAW
                        ws[c5] = formattedaddress
                        ws[c6] = locusaddress
                        ws[c7] = 'All Deeds:'
                        ws[c8] = alldeeds
                        startCell += 4
                        startCellb += 4
                        startCellp += 4
                        startCelld += 4
                        relAW = ''

            except Exception as e:
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                QMessageBox.critical(self.iface.mainWindow(), "ERROR!",
                                     "Please ensure that you have a parcel selected\nand attempt to "
                                     "generate the output again.\n\n"
                                     "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                         exc_tb.tb_lineno) + ' ' + str(e))

            delRow = startCell
            ws.delete_rows(delRow, 200)

            # QgsMessageLog.logMessage('cell/aNo: ' + str(startCell) + '/' + str(aNo), 'sGIS',level=Qgis.Info)

            for s in sheets:

                if s != 'maptable':
                    sheet_name = wb.get_sheet_by_name(s)
                    wb.remove_sheet(sheet_name)

            # Save the file
            mtfile = str(jipath) + "\\" + jobNo + "_MapTable_" + datetime.datetime.today().strftime(
                '%Y.%m.%d') + ".xlsx"
            QgsMessageLog.logMessage('Saving files: ' + mtfile + '...', 'sGIS', level=Qgis.Info)
            try:
                wb.save(mtfile)

            except Exception as e:
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                QMessageBox.critical(self.iface.mainWindow(), "File Open",
                                     "Please ensure that you do not have today's mapTable open in Excel\nand attempt to "
                                     "generate the output again.\n\n"
                                     "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(exc_tb.tb_lineno) + ' ' + str(e))
                                                 
            year = datetime.datetime.today().strftime('%Y')
            
            jobYear = '20' + jobNo[:2]

            if jobYear == year:
                year = year
            else:
                year = jobYear
            
            path = os.path.join(jobsPath, year, jobNo)  # need to programattically grab year
            jipath = os.path.join(path, "Job_Info")
            
            yfile = str(jipath) + "\\" + jobNo + "_YellowSheet_" + datetime.datetime.today().strftime(
                '%Y.%m.%d') + ".xlsx"
            pfile = str(jipath) + "\\" + jobNo + "_YellowSheet_" + datetime.datetime.today().strftime(
                '%Y.%m.%d') + ".pdf"
            pfile2 = str(jipath) + "\\" + jobNo + "_MapTable_" + datetime.datetime.today().strftime(
                '%Y.%m.%d') + ".pdf"    
            outfile = str(jipath) + "\\" + jobNo+ "_YellowSheet&MapTable_" + datetime.datetime.today().strftime(
                '%Y.%m.%d') + ".pdf"

                                         
            #convert to PDF
            excel = client.Dispatch("Excel.Application")
            excel.Interactive = False
            excel.Visible = False
            excel.DisplayAlerts = False
            sheets = excel.Workbooks.Open(yfile)
            wks = sheets.Worksheets[0]
            wks.ExportAsFixedFormat(0,pfile)
            sheets.Close(False)
            sheets = None
            excel.Quit()
            excel = None

            excel2 = client.Dispatch("Excel.Application")
            excel2.Interactive = False
            excel2.Visible = False
            excel2.DisplayAlerts = False
            sheets2 = excel2.Workbooks.Open(mtfile)
            wks2 = sheets2.Worksheets[0]
            wks2.ExportAsFixedFormat(0,pfile2)
            wks2 = None
            sheets2.Close(True)
            sheets2 = None
            excel2.Quit()
            excel2 = None
            
            filenames = [pfile, pfile2]
            merger = PdfWriter()
            for filename in filenames:
                merger.append(filename)
            merger.write(outfile)
            merger.close()
            merger = None

            QgsMessageLog.logMessage('Saving duplex and flushing temporary files: ' + outfile + '...', 'sGIS', level=Qgis.Info)
                        
            self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
            self.iface.setActiveLayer(self.vl)    
            target_field = 'job_no' 
            condition = jobNo                   
            
            #re-select job_no by attribute
            self.vl.selectByExpression(f"\"{target_field}\"='{condition}'")          
            self.iface.messageBar().clearWidgets()
                       
            try:
                os.remove(yfile)
                os.remove(pfile)
                os.remove(mtfile)
                os.remove(pfile2)
                
            except Exception as e:
                pass
                
            return

        layerRelated = QgsProject.instance().mapLayersByName('relatedwork')[0]
        layerRelated.setSubsetString('id > 1')

        self.vl.dataProvider().forceReload()
        self.iface.mapCanvas().refresh()
        self.iface.setActiveLayer(layerActive)
        layerActive.selectByIds([fid])
        resetLegend(self)

    def identAbutters(self):

        aNo = 0

        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(self.vl)

        try:
            attribs = self.iface.activeLayer().selectedFeatures()[0]

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "No Selection",
                                 "Please ensure that you have a parcel selected\nand attempt to "
                                 "generate the output again.\n\n"
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))
            return

        jobNo = attribs["job_no"]
        map_bk_lot = attribs["map_bk_lot"]

    def getRelatedWork(self, feature, cfg):
        if cfg == 0:

            layer = self.iface.activeLayer()

            if layer.name() not in ('jobs', 'abutters'):
                return

            layerJobs = QgsProject.instance().mapLayersByName('jobs')[0]
            layerPlans = QgsProject.instance().mapLayersByName('plans')[0]
            layerRelated = QgsProject.instance().mapLayersByName('relatedwork')[0]
            layerAbutters = QgsProject.instance().mapLayersByName('abutters')[0]
            layerSupps = QgsProject.instance().mapLayersByName('supplementals')[0]

            self.fb = QgsProcessingFeedback()
            self.context = QgsProcessingContext

            if self.iface.activeLayer().name() == 'jobs':
                jobNo = feature['job_no']
                map_bk_lot = feature['map_bk_lot']
                town = feature['town']
                if str(map_bk_lot) == 'NULL':
                    pass
                else:
                    # QgsMessageLog.logMessage('MAPBKLOT: ' + str(map_bk_lot), 'sGIS', level=Qgis.Info)

                    self.iface.actionToggleEditing().trigger()
                    layerRelated.setSubsetString(u'"job_no" = \'%s\'' % (jobNo))
                    listOfIds = [feat.id() for feat in layerRelated.getFeatures()]
                    layerRelated.dataProvider().deleteFeatures(listOfIds)
                    self.iface.activeLayer().commitChanges()
                    self.iface.setActiveLayer(layerJobs)
                    layerRelated.setSubsetString('id > 1')

                    lId = self.iface.activeLayer().id()
                    varI = QgsProcessingFeatureSourceDefinition(str(lId), True)

                    # QgsMessageLog.logMessage('lId: ' + str(lId) + ' ' + str(varI), 'sGIS', level=Qgis.Info)

                    try:
                        processing.runAndLoadResults("qgis:intersection",
                                                 {'INPUT': varI,
                                                  'INPUT_FIELDS': ['job_no', 'map_bk_lot'],
                                                  'OUTPUT': 'memory:tmp_related',
                                                  'OVERLAY': layerPlans,
                                                  'OVERLAY_FIELDS': ['plan_no', 'job', 'old_plan']
                                                  }, feedback=self.fb)
                        self.addRelated(0)
                    except Exception as e:
                        pass
                    # getRelated JOBS for JOB feature
                    try:
                        processing.runAndLoadResults("qgis:intersection",
                                                     {'INPUT': varI,
                                                      'INPUT_FIELDS': ['job_no', 'map_bk_lot'],
                                                      'OUTPUT': 'memory:tmp_related',
                                                      'OVERLAY': layerJobs,
                                                      'OVERLAY_FIELDS': ['plan_no', 'job', 'old_plan']
                                                      }, feedback=self.fb)
                        self.addRelated(0)
                    except Exception as e:
                        pass
                    # getRelated SUPPS for JOB feature
                    try:

                        processing.runAndLoadResults("qgis:intersection",
                                                     {'INPUT': varI,
                                                      'INPUT_FIELDS': ['job_no', 'map_bk_lot'],
                                                      'OUTPUT': 'memory:tmp_related',
                                                      'OVERLAY': layerSupps,
                                                      'OVERLAY_FIELDS': ['plan_no', 'job', 'old_plan']
                                                      }, feedback=self.fb)
                        self.addRelated(0)
                    except Exception as e:
                        pass

            elif self.iface.activeLayer().name() == 'abutters':
                lId = self.iface.activeLayer().id()
                varI = QgsProcessingFeatureSourceDefinition(str(lId), True)
                mbl = feature['map_bk_lot']

                self.iface.actionToggleEditing().trigger()

                # NEED TO CHECK POTENTIAL TOWN OVERLAP ISSUES. add TOWN condition when getting RELATED for ABUTTER?
                layerRelated.setSubsetString(u'"map_bk_lot" = \'%s\'' % (mbl))
                listOfIds = [feat.id() for feat in layerRelated.getFeatures()]
                # QgsMessageLog.logMessage('LIST OF IDs: ' + str(listOfIds), 'sGIS', level=Qgis.Info)

                layerRelated.dataProvider().deleteFeatures(listOfIds)
                self.iface.activeLayer().commitChanges()
                self.iface.setActiveLayer(layerAbutters)
                layerRelated.setSubsetString('id > 1')

                # getRelated PLANS for ABUTTER feature
                processing.runAndLoadResults("qgis:intersection",
                                             {'INPUT': varI,
                                              'INPUT_FIELDS': ['job_no', 'map_bk_lot', 'town_parcels'],
                                              'OUTPUT': 'memory:tmp_related',
                                              'OVERLAY': layerPlans,
                                              'OVERLAY_FIELDS': ['plan_no', 'job', 'old_plan']
                                              }, feedback=self.fb)

                self.addRelated(0)

                processing.runAndLoadResults("qgis:intersection",
                                             {'INPUT': varI,
                                              'INPUT_FIELDS': ['job_no', 'map_bk_lot', 'town'],
                                              'OUTPUT': 'memory:tmp_related',
                                              'OVERLAY': layerJobs,
                                              'OVERLAY_FIELDS': ['plan_no', 'job', 'old_plan']
                                              }, feedback=self.fb)

                self.addRelated(0)
                processing.runAndLoadResults("qgis:intersection",
                                             {'INPUT': varI,
                                              'INPUT_FIELDS': ['job_no', 'map_bk_lot', 'town'],
                                              'OUTPUT': 'memory:tmp_related',
                                              'OVERLAY': layerSupps,
                                              'OVERLAY_FIELDS': ['plan_no', 'job', 'old_plan']
                                              }, feedback=self.fb)
                self.addRelated(0)
                self.addRelatedAbutterJobs(mbl)

            else:
                pass
        else:
            pass

    def addRelated(self, cfg):
        # try:

        layerTmpRelated = QgsProject.instance().mapLayersByName('Intersection')[0]
        layerTmpRelated.selectAll()
        self.iface.actionCopyFeatures().trigger()
        self.vl = QgsProject.instance().mapLayersByName('relatedwork')[0]
        self.iface.setActiveLayer(self.vl)
        self.iface.actionToggleEditing().trigger()

        self.iface.actionIdentify().trigger()
        self.iface.actionPasteFeatures().trigger()
        self.iface.activeLayer().commitChanges()
        self.iface.messageBar().clearWidgets()

        QgsProject.instance().removeMapLayer(layerTmpRelated.id())
        self.vl.dataProvider().forceReload()
        self.iface.mapCanvas().refresh()
        self.iface.messageBar().clearWidgets()

    def addRelatedAbutterJobs(self, map):
        # try:
        layerJobs = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(layerJobs)
        layerJobs.setSubsetString(u'"map_bk_lot" = \'%s\'' % (map))
        layerJobs.selectAll()

        # QgsMessageLog.logMessage('JOB MAP FILTER: ' + str(map), 'sGIS', level=Qgis.Info)
        #return
        self.iface.actionCopyFeatures().trigger()
        self.vl = QgsProject.instance().mapLayersByName('relatedwork')[0]
        self.iface.setActiveLayer(self.vl)
        self.iface.actionToggleEditing().trigger()
        self.iface.actionIdentify().trigger()
        self.iface.actionPasteFeatures().trigger()
        self.iface.activeLayer().commitChanges()

        layerJobs.setSubsetString("\"supp_type\" = 'X'")

        self.vl.dataProvider().forceReload()
        self.iface.mapCanvas().refresh()

    def showFeatureCount(layers):

        layer = layers[0]
        if layer.type() == QgsMapLayer.VectorLayer:
            root = QgsProject.instance().layerTreeRoot()
            myLayerNode = root.findLayer(layer.id())
            myLayerNode.setCustomProperty("showFeatureCount", True)

    def updateJobRelated(self, feat):

        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(self.vl)
        jobNo = feat['job_no']
        self.town = feat['town']
        layerRelated = QgsProject.instance().mapLayersByName('relatedwork')[0]
        relatedWork = []
        plans = []
        plans2 = []
        ppval = ''
        pval = ''
        pNo = 0

        layerRelated.setSubsetString(u'"job_no" = \'%s\'' % (jobNo))
        exp = QgsExpression(u'"job_no" = \'%s\'' % (jobNo))
        request = QgsFeatureRequest(exp)
        request.setSubsetOfAttributes(['job_no', 'old_plan'], layerRelated.fields())
        request.setFlags(QgsFeatureRequest.NoGeometry)

        for feat in layerRelated.getFeatures(request):

            pFinal = ''
            oldPlan = feat['old_plan']
            pval = oldPlan

            if str(jobNo) in (str(pval)):
                pass
            else:

                if str(ppval) == str(pval):
                    pval = ''

                elif str(pval) == 'NULL':
                    pval = ''
                else:
                    pval = pval
                    plans.append(pval)

            pNo += 1
            ppval = pval
            # QgsMessageLog.logMessage('UJR_plans: ' + str(plans), 'sGIS', level=Qgis.Info)

        feat = self.iface.activeLayer().selectedFeatures()[0]

        if len(plans) >= 1:
            plans.sort(reverse=True)
            pFinal = str(plans)
            pFinal = pFinal.strip('[')
            pFinal = pFinal.strip(']')
            pFinal = pFinal.replace('\'', '')

            self.iface.actionToggleEditing().trigger()
            layerData = self.vl.dataProvider()
            idx3 = layerData.fieldNameIndex('old_plan_no')
            self.vl.changeAttributeValue(feat.id(), idx3, str(pFinal))
            self.vl.updateFields()
            self.iface.activeLayer().commitChanges()

            plans = []
            pval = ''
            ppval = ''

        else:
            pFinal = ''
        layerRelated.setSubsetString('id > 1')
        # END getAllRelated for selected JOB
        return pFinal

    def updateAbutterRelated(self, feat, jobNo):

        self.vl = QgsProject.instance().mapLayersByName('abutters')[0]
        self.iface.setActiveLayer(self.vl)
        mbl = feat['map_bk_lot']
        layerRelated = QgsProject.instance().mapLayersByName('relatedwork')[0]
        relatedWork = []
        plans = []
        plans2 = []
        ppval = ''
        pval = ''
        pNo = 0

        layerRelated.setSubsetString(u'"map_bk_lot" = \'%s\'' % (mbl))
        exp = QgsExpression(u'"map_bk_lot" = \'%s\'' % (mbl))
        request = QgsFeatureRequest(exp)
        request.setSubsetOfAttributes(['old_plan'], layerRelated.fields())
        request.setFlags(QgsFeatureRequest.NoGeometry)

        for feat in layerRelated.getFeatures(request):

            oldPlan = feat['old_plan']
            fid = feat['id']
            town = feat['town']
            town_parcels = feat['town_parcels']

            if str(town_parcels) == 'NULL':
                town = town
            else:
                town = town_parcels
               
            if str(town) != str(self.town):
                pval = 'NULL'
            else:
                pval = oldPlan

            if str(jobNo) in str(pval):
                QgsMessageLog.logMessage('jobNo: ' + str(jobNo) + ' - MATCH, IGNORE!!!',
                                         'sGIS', level=Qgis.Info)
                pval = ''

            if str(ppval) == str(pval):
                pval = ''

            plen = len(plans)

            if str(pval) == 'NULL':
                pval = ''

                QgsMessageLog.logMessage('NULL pval: ' + str(pval), 'sGIS', level=Qgis.Info)
            else:
                if pval in (plans):
                    pass
                else:
                    pval = pval
                    plans.append(pval)

            pNo += 1
            ppval = pval
            QgsMessageLog.logMessage('UAR_plans: ' + str(plans), 'sGIS', level=Qgis.Info)

        feat = self.iface.activeLayer().selectedFeatures()[0]

        if len(plans) >= 1:
            plans.sort(reverse=True)
            pFinal = str(plans)
            pFinal = pFinal.strip('[')
            pFinal = pFinal.strip(']')
            pFinal = pFinal.replace('\'', '')

            if str(pFinal) == '':
                pFinal = 'N/A'

            plans = []

        else:
            plans = []
            pFinal = 'N/A'
            pass

        layerRelated.setSubsetString('id > 1')
        layerRelated.setSubsetString(u'"map_bk_lot" = \'%s\'' % (mbl))
        exp = QgsExpression(u'"map_bk_lot" = \'%s\'' % (mbl))
        request = QgsFeatureRequest(exp)
        request.setSubsetOfAttributes(['gid'], layerRelated.fields())
        request.setFlags(QgsFeatureRequest.NoGeometry)
        return pFinal


class sgis_printContacts(object):
    def __init__(self, iface):
        # save reference to the QGIS interface
        self.iface = iface

    def initGui(self):

        self.action = QAction("PC", self.iface.mainWindow())
        self.action.triggered.connect(self.run)
        self.action.trigger()

    def run(self):
        self.identContacts()
        #resetLegend(self)

    def identContacts(self):
        import datetime
        year = datetime.datetime.today().strftime('%Y')

        aNo = 0

        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(self.vl)

        try:
            attribs = self.iface.activeLayer().selectedFeatures()[0]

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "No Selection",
                                 "Please ensure that you have a parcel selected\nand attempt to "
                                 "generate the output again.\n\n"
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))
            return

        jobNo = attribs["job_no"]
        jobYear = '20' + jobNo[:2]

        if jobYear == year:
            year = year
        else:
            year = jobYear

        jobID = attribs["sid"]

        path = os.path.join(jobsPath, year, jobNo)
        jipath = os.path.join(path, "Job_Info")
        QgsMessageLog.logMessage('Checking output folder: ' + path + '...', 'sGIS', level=Qgis.Info)
        if not os.path.exists(path):
            os.makedirs(path)
        QgsMessageLog.logMessage('Checking output folder: ' + jipath + '...', 'sGIS', level=Qgis.Info)
        if not os.path.exists(jipath):
            os.makedirs(jipath)

        from openpyxl import load_workbook

        fPath = os.path.dirname(os.path.abspath(__file__)) + '\\GIS_templates.xlsx'
        wb = load_workbook(fPath)

        # grab the correct worksheet
        for s in range(len(wb.sheetnames)):
            if wb.sheetnames[s] == 'contacts':
                break
        wb.active = s
        sheets = wb.sheetnames
        ws = wb.active

        for s in sheets:
            if s != 'contacts':
                sheet_name = wb.get_sheet_by_name(s)
                wb.remove_sheet(sheet_name)

        layer3 = QgsProject.instance().mapLayersByName('contacts')[0]
        exp = QgsExpression(u'"jobs_id" = \'%s\'' % (jobID))
        request = QgsFeatureRequest(exp)
        request.setSubsetOfAttributes(['jobs_id'], layer3.fields())
        request.setFlags(QgsFeatureRequest.NoGeometry)

        startCell = 3

        for f in layer3.getFeatures(request):

            try:
                aNo += 1
                ws['B2'] = str(jobNo) + ' Contacts:'               

                c1 = 'B' + str(startCell)
                c2 = 'B' + str(startCell + 2)
                c3 = 'C' + str(startCell + 3)
                c4 = 'C' + str(startCell + 4)
                c5 = 'C' + str(startCell + 2)
                c6 = 'E' + str(startCell + 3)
                c7 = 'E' + str(startCell + 4)
                c8 = 'G' + str(startCell + 3)
                c9 = 'G' + str(startCell + 4)
                c10 = 'I' + str(startCell + 3)

                ws[c1] = str(f["contact_type"])
                ws[c2] = str(f["contact_name"]) #A5
                ws[c3] = str(f["primary_contact"])
                ws[c4] = str(f["secondary_contact"])
                ws[c5] = str(f["contact_addr"])
                ws[c6] = str(f["phone_mobile"])
                ws[c7] = str(f["email_primary"])
                ws[c8] = str(f["phone_work"]) + ' ' + str(f["extension"])
                ws[c9] = str(f["email_secondary"])
                ws[c10] = str(f["phone_home"])

                startCell += 5

            except Exception as e:
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                     "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                         exc_tb.tb_lineno) + ' ' + str(e))
                return

        if aNo == 1:
            delRow = startCell
            ws.delete_rows(delRow, 100)
        else:
            delRow = startCell
            ws.delete_rows(delRow, 100)
        QgsMessageLog.logMessage(str(aNo) + ' contacts found.', 'sGIS', level=Qgis.Info)
        
        # Save the file
        cfile = str(jipath) + "\\" + jobNo + "_Contacts_" + datetime.datetime.today().strftime(
            '%Y.%m.%d') + ".xlsx"
        pfile = str(jipath) + "\\" + jobNo + "_Contacts_" + datetime.datetime.today().strftime(
            '%Y.%m.%d') + ".pdf"
        QgsMessageLog.logMessage('Saving file: ' + cfile + '...', 'sGIS', level=Qgis.Info)
        
                
        #save XLSX
        wb.save(cfile)


class sgis_printEstimates(object):

    def __init__(self, iface):
        # save reference to the QGIS interface
        self.iface = iface

    def initGui(self):
        self.action = QAction("P.EST", self.iface.mainWindow())
        self.action.triggered.connect(self.run)
        self.action.trigger()

    def run(self):
        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(self.vl)

        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font
        from openpyxl.styles.differential import DifferentialStyle
        from openpyxl.formatting.rule import Rule

        efile = os.path.join(estimatesPath)
        wb = openpyxl.Workbook()
        ws = wb.active

        grey_fill = PatternFill(bgColor="DDDDDD")
        dxf = DifferentialStyle(fill=grey_fill)
        r = Rule(type="expression", dxf=dxf, stopIfTrue=True)
        r.formula = ['$I1="A"']
        ws.conditional_formatting.add("A1:I9999", r)

        ws['A1'] = 'job_no'
        ws['B1'] = 'rev_no'
        ws['C1'] = 'job_type'
        ws['D1'] = 'client_name'
        ws['E1'] = 'town'
        ws['F1'] = 'locus_addr'
        ws['G1'] = 'folder_name'
        ws['H1'] = 'date_requested'
        ws['I1'] = 'active'

        hdrfont = Font(size=12, bold=True)
        for cell in ws["1:1"]:
            cell.font = hdrfont

        def get_jobNo(j):
            return j['job_no']

        def makeRow(active, j, ws):
            job_no = str(j["job_no"])
            rev_no = str(j["rev_no"])
            job_type = str(j["job_type"])
            client_name = str(j["client_name"])
            town = str(j["town"])
            locus_addr = str(j["locus_addr"])
            folder_name = str(j["folder_name"])
            date_requested = str(j["date_requested"])
            row = job_no + "|" + rev_no + "|" + job_type + "|" + client_name \
                  + "|" + town + "|" + locus_addr + "|" + folder_name + "|" + date_requested + "|" + active
            return row

        def writeActives(active, j, ws):
            if active == 'A':
                row = makeRow(active, j, ws)
                # QgsMessageLog.logMessage('row: ' + row + '...', 'sGIS', level=Qgis.Info)
                row = list(row.split("|"))
                ws.append(row)
            else:
                pass

        def writeInactives(active, j, ws):
            if active == 'A':
                pass
            else:
                row = makeRow(active, j, ws)
                # QgsMessageLog.logMessage('row: ' + row + '...', 'sGIS', level=Qgis.Info)
                row = list(row.split("|"))
                ws.append(row)

        exT = QgsExpression(u'"active" = \'%s\'' % ('True'))
        exF = QgsExpression(u'"active" = \'%s\'' % ('False'))
        requestT = QgsFeatureRequest(exT)
        requestT.setFlags(QgsFeatureRequest.NoGeometry)
        requestF = QgsFeatureRequest(exF)
        requestF.setFlags(QgsFeatureRequest.NoGeometry)

        activeJobs = sorted(self.vl.getFeatures(requestT), key=get_jobNo)
        inactiveJobs = sorted(self.vl.getFeatures(requestF), key=get_jobNo)

        for j in activeJobs:
            estimate = str(j["estimate"])
            if estimate == 'False':
                pass
            else:
                active = str(j["active"])
                if active == 'True':
                    active = 'A'
                    writeActives(active, j, ws)
                else:
                    active = ''
                    pass

        for j in inactiveJobs:
            estimate = str(j["estimate"])
            if estimate == 'False':
                pass
            else:
                active = str(j["active"])
                if active == 'True':
                    active = 'A'
                    pass
                else:
                    active = ''
                    writeInactives(active, j, ws)

        for column_cells in ws.columns:
            length = max(len(cell.value) + 5 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length

        QgsMessageLog.logMessage('Saving file: ' + efile + '...', 'sGIS', level=Qgis.Info)
        wb.save(efile)


class sgis_printMapView(object):

    def __init__(self, iface):
        # save reference to the QGIS interface
        self.iface = iface

    def initGui(self):
        self.action = QAction("PMV", self.iface.mainWindow())
        self.action.triggered.connect(self.run)
        self.action.trigger()

    def run(self):
        import datetime

        jstd = os.path.dirname(os.path.realpath(__file__)) + '/qml/jobs_std.qml'
        jprn =  os.path.dirname(os.path.realpath(__file__)) + '/qml/jobs_print.qml'
        astd =  os.path.dirname(os.path.realpath(__file__)) + '/qml/abutters_std.qml'
        aprn =  os.path.dirname(os.path.realpath(__file__)) + '/qml/abutters_print.qml'

        year = datetime.datetime.today().strftime('%Y')
        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(self.vl)

        try:
            attribs = self.iface.activeLayer().selectedFeatures()[0]
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "No Selection",
                                 "Please ensure that you have a parcel selected\nand attempt to "
                                 "generate the output again.\n\n"
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))
            return

        jobNo = attribs["job_no"]
        jobType = attribs["job_type"]
        jobSubType = attribs["jobSubtype"]
        county = attribs["county"]
        jobYear = '20' + jobNo[:2]

        ids = [attribs.id()]

        if jobYear == year:
            year = year
        else:
            year = jobYear

        folder_name = attribs["folder_name"]

        path = os.path.join(jobsPath, year, jobNo)
        jipath = os.path.join(path, "Job_Info")

        QgsMessageLog.logMessage('Checking output folder: ' + path + '...', 'sGIS', level=Qgis.Info)
        if not os.path.exists(path):
            os.makedirs(path)
        QgsMessageLog.logMessage('Checking output folder: ' + jipath + '...', 'sGIS', level=Qgis.Info)
        if not os.path.exists(jipath):
            os.makedirs(jipath)

        cfile = str(jipath) + "\\" + jobNo + "_MapView_" + datetime.datetime.today().strftime(
            '%Y.%m.%d') + ".pdf"
     
        # enable abutters print style | filter: reffererj = job_no
        self.vl = QgsProject.instance().mapLayersByName('abutters')[0]
        self.vl.loadNamedStyle(aprn)
        self.vl.setSubsetString('"referrerj"=\'%s\'' % jobNo)

        # enable jobs print style | filter: job_no = job_no

        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.vl.selectAll()
        self.clone_layer = processing.run("native:saveselectedfeatures", {'INPUT': self.vl, 'OUTPUT': 'memory:'})['OUTPUT']
        self.vl.removeSelection()
        QgsProject.instance().addMapLayer(self.clone_layer)

        self.ol = QgsProject.instance().mapLayersByName('output')[0]

        self.vl.loadNamedStyle(jprn)
        self.vl.setSubsetString('"job_no"=\'%s\'' % jobNo)
        currentScale = self.iface.mapCanvas().scale()

        # generate output
        self.make_pdf(cfile, jobNo, folder_name, jobType, county, cfile)

        # enable abutters standard style | filter: n/a
        self.vl = QgsProject.instance().mapLayersByName('abutters')[0]
        self.vl.loadNamedStyle(astd)
        self.vl.setSubsetString('')

        # enable jobs standard style | filter: n/a
        # QgsMessageLog.logMessage('qmlPath: ' + str(qmlPath) + '...', 'sGIS', level=Qgis.Info)
        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.vl.loadNamedStyle(jstd)
        self.vl.setSubsetString('"supp_type"=\'%s\'' % 'X')

        # disable all type-specific layers:
        self.toggleLayer('HAT_2015', 0)
        self.toggleLayer('75ft_SB_HAT_2015', 0)
        # self.toggleLayer('S_Fld_Haz_Ar', 0)
        # self.toggleLayer('S_Fld_Haz_Ar VE', 0)

        layer = QgsProject.instance().mapLayersByName('output')[0]
        QgsProject.instance().removeMapLayer(layer.id())
        self.iface.messageBar().clearWidgets()
        resetLegend(self)
        
        self.iface.setActiveLayer(self.vl)
        self.vl.selectByIds(ids)
        self.iface.mapCanvas().zoomScale(currentScale)
        self.vl.triggerRepaint()
        self.reset()
        resetLegend(self)

    def spin(self, seconds):
        import time
        """Pause for set amount of seconds, replaces time.sleep so program doesn't stall"""
        time_end = time.time() + seconds
        while time.time() < time_end:
            QGuiApplication.processEvents()

    def toggleLayer(self, layer, status):
        # QgsMessageLog.logMessage('TOGGLE: ' + layer + '...', 'sGIS', level=Qgis.Info)
        try:
            lyr = QgsProject.instance().mapLayersByName(layer)[0]
            if status == 0:
                QgsProject.instance().layerTreeRoot().findLayer(lyr.id()).setItemVisibilityChecked(False)
            else:
                QgsProject.instance().layerTreeRoot().findLayer(lyr.id()).setItemVisibilityChecked(True)
        except Exception as e:
            pass

    def make_pdf(self, cf, jn, fn, jt, county, cfile):

        # QGuiApplication.setOverrideCursor(Qt.WaitCursor)
        if str(fn) == 'NULL':
            fn = 'Folder Name'
        else:
            fn = fn

        jobType = jt

        # enable layers based on job_type
        if jt == 'BRSDP':
            jt = 'SDP'
        else:
            jt = jt

        if jt == 'SDP':
            self.toggleLayer('HAT_2015', 1)
            self.toggleLayer('75ft_SB_HAT_2015', 1)
            self.toggleLayer('S_FIRM_PAN', 1)

        elif jt == 'MIS':
            self.toggleLayer('S_Fld_Haz_Ar', 1)
            self.toggleLayer('S_Fld_Haz_Ar VE', 1)
            self.toggleLayer('S_FIRM_PAN', 1)

        else:
            pass

        title = str(jn) + ' (' + str(jobType) + ')' + ' | ' + str(fn)

        project = QgsProject.instance()
        l = QgsPrintLayout(project)
        l.initializeDefaults()
        l.setUnits(QgsUnitTypes.LayoutMillimeters)
        page = l.pageCollection().pages()[0]

        msg = QMessageBox()
        msg.setWindowTitle('Select Orientation')
        msg.setText('Select the desired orientation for the mapView:')
        portrait = msg.addButton('Portrait', QMessageBox.AcceptRole)
        landscape = msg.addButton('Landscape', QMessageBox.AcceptRole)
        cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
        msg.setDefaultButton(portrait)
        QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
        msg.exec_()
        msg.deleteLater()
        QGuiApplication.restoreOverrideCursor()

        if msg.clickedButton() is portrait:
            paperSize = self.setPaperSizePortrait()
            QgsMessageLog.logMessage('Saving file: ' + cfile + '...', 'sGIS', level=Qgis.Info)
        elif msg.clickedButton() is landscape:
            paperSize = self.setPaperSizeLandscape()
            QgsMessageLog.logMessage('Saving file: ' + cfile + '...', 'sGIS', level=Qgis.Info)
        else:
            QgsMessageLog.logMessage('DEBUG: MapView Export CANCELLED.', 'sGIS', level=Qgis.Info)
            return

        page.setPageSize(QgsLayoutSize(paperSize[0], paperSize[1]))

        lm = 10  # left margin
        tm = 30  # upper margin
        bm = 65  # lower margin

        refSize = paperSize[0]
        if paperSize[1] < refSize:
            refSize = paperSize[1]

        # add map
        x, y = lm, tm
        w, h = paperSize[0] - 2 * lm, paperSize[1] - bm

        theMap = QgsLayoutItemMap(l)
        theMap.updateBoundingRect()
        theMap.setRect(QRectF(x, y, w, h))
        theMap.setPos(x, y)
        theMap.setFrameEnabled(True)
        # theMap.setScale(4000)
        
        theMap.setLayers(project.mapThemeCollection().masterVisibleLayers())  # remember ANNOTATION!
        theMap.setExtent(self.iface.mapCanvas().extent())
        theMap.attemptSetSceneRect(QRectF(x, y, w, h))
        l.addItem(theMap)

        titleFont = QFont("Arial", 20)
        titleFont.setBold(True)

        titleLabel = QgsLayoutItemLabel(l)
        titleLabel.setText(title)
        titleLabel.setPos(lm, 10)
        titleLabel.setFont(titleFont)
        titleLabel.adjustSizeToText()
        l.addItem(titleLabel)

        scaleBar = QgsLayoutItemScaleBar(l)
        scaleBar.setLinkedMap(theMap)
        scaleBar.applyDefaultSettings()
        scaleBar.applyDefaultSize(QgsUnitTypes.DistanceFeet)

        scaleBar.setNumberOfSegmentsLeft(0)
        scaleBar.setNumberOfSegments(5)
        scaleBar.update()
        scaleBar.setPos(lm, tm + (paperSize[1] - bm) + 10)

        l.addItem(scaleBar)

        exporter = QgsLayoutExporter(l)
        pdf_settings = exporter.PdfExportSettings()  # dpi?
        exporter.exportToPdf(cf, pdf_settings)
        
        import datetime
        year = datetime.datetime.today().strftime('%Y')
        jobYear = '20' + jn[:2]
        if jobYear == year:
            year = year
        else:
            year = jobYear
            
        path = os.path.join(jobsPath, year, jn)  # need to programattically grab year
        jipath = os.path.join(path, "Job_Info")
        
        cfile = str(jipath) + "\\" + jn + "_Contacts_" + datetime.datetime.today().strftime(
            '%Y.%m.%d') + ".xlsx"
        pfile = str(jipath) + "\\" + jn + "_Contacts_" + datetime.datetime.today().strftime(
            '%Y.%m.%d') + ".pdf"
        mvfile = str(jipath) + "\\" + jn + "_MapView_" + datetime.datetime.today().strftime(
            '%Y.%m.%d') + ".pdf"
        outfile = str(jipath) + "\\" + jn+ "_MapView&Contacts_" + datetime.datetime.today().strftime(
            '%Y.%m.%d') + ".pdf"

        #convert to PDF
        excel = client.Dispatch("Excel.Application")
        excel.Interactive = False
        excel.Visible = False
        excel.DisplayAlerts = False
        
        sheets = excel.Workbooks.Open(cfile)
        wks = sheets.Worksheets[0]
        wks.ExportAsFixedFormat(0,pfile)
        sheets.Close(False)
        sheets = None
        excel.Quit()
        excel = None
        
        filenames = [mvfile, pfile]
        merger = PdfWriter()
        for filename in filenames:
            merger.append(filename)
        merger.write(outfile)
        merger.close()
        merger = None
        
        try:
            os.remove(cfile)
            os.remove(pfile)
            os.remove(mvfile)
        
        except Exception as e:
            pass

    def setPaperSizePortrait(self):

        longSide = 279
        shortSide = 216
        width = shortSide
        height = longSide

        return width, height

    def setPaperSizeLandscape(self):

        longSide = 216
        shortSide = 279
        width = shortSide
        height = longSide

        return width, height

    def reset(self):

        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(self.vl)
        jform = os.path.dirname(os.path.realpath(__file__)) + '/forms/jobs.ui'
        jpy = os.path.dirname(os.path.realpath(__file__)) + '/forms/jobs_init.py'
        form_config = self.iface.activeLayer().editFormConfig()
        form_config.setUiForm(jform)
        form_config.setInitFilePath(jpy)
        self.iface.activeLayer().setEditFormConfig(form_config)

        self.vl = QgsProject.instance().mapLayersByName('plans')[0]
        self.iface.setActiveLayer(self.vl)
        jform = os.path.dirname(os.path.realpath(__file__)) + '/forms/plans.ui'
        jpy = os.path.dirname(os.path.realpath(__file__)) + '/forms/plans_init.py'
        form_config.setUiForm(jform)
        form_config.setInitFilePath(jpy)
        self.iface.activeLayer().setEditFormConfig(form_config)

        self.vl = QgsProject.instance().mapLayersByName('contacts')[0]
        self.iface.setActiveLayer(self.vl)
        jform = os.path.dirname(os.path.realpath(__file__)) + '/forms/contacts.ui'
        jpy = os.path.dirname(os.path.realpath(__file__)) + '/forms/contacts_init.py'
        form_config.setUiForm(jform)
        form_config.setInitFilePath(jpy)
        self.iface.activeLayer().setEditFormConfig(form_config)

        self.vl = QgsProject.instance().mapLayersByName('supplementals')[0]
        self.iface.setActiveLayer(self.vl)
        jform = os.path.dirname(os.path.realpath(__file__)) + '/forms/supplementals.ui'
        jpy = os.path.dirname(os.path.realpath(__file__)) + '/forms/supplementals_init.py'
        form_config.setUiForm(jform)
        form_config.setInitFilePath(jpy)
        self.iface.activeLayer().setEditFormConfig(form_config)
        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(self.vl)


class sgis_printSiteMap(object):

    def __init__(self, iface):
        # save reference to the QGIS interface
        self.iface = iface

    def initGui(self):
        self.action = QAction("PMV", self.iface.mainWindow())
        self.action.triggered.connect(self.run)
        self.action.trigger()

    def run(self):
        import datetime
        year = datetime.datetime.today().strftime('%Y')
        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(self.vl)

        try:
            attribs = self.iface.activeLayer().selectedFeatures()[0]
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "No Selection",
                                 "Please ensure that you have a parcel selected\nand attempt to "
                                 "generate the output again.\n\n"
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))
            return

        ids = [attribs.id()]

        jobNo = attribs["job_no"]
        jobYear = '20' + jobNo[:2]

        if jobYear == year:
            year = year
        else:
            year = jobYear

        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(self.vl)
        self.vl.setSubsetString('"job_no"=\'%s\'' % jobNo)

        path = os.path.join(jobsPath, year, jobNo)
        dwgpath = os.path.join(path, "dwg")

        QgsMessageLog.logMessage('Checking output folder: ' + path + '...', 'sGIS', level=Qgis.Info)

        if not os.path.exists(path):
            os.makedirs(path)

        if not os.path.exists(dwgpath):
            os.makedirs(dwgpath)

        ifile = str(dwgpath) + "\\" + "site.jpg"

        QgsMessageLog.logMessage('Saving file: ' + ifile + '...', 'sGIS', level=Qgis.Info)

        std =  os.path.dirname(os.path.realpath(__file__)) + '/qml/jobs_std.qml'
        prn =  os.path.dirname(os.path.realpath(__file__)) + '/qml/jobs_print.qml'
        sMap =  os.path.dirname(os.path.realpath(__file__)) + '/qml/jobs_siteMap.qml'

        currentScale = self.iface.mapCanvas().scale()

        layersNames = []
        for i in self.iface.mapCanvas().layers():
            layersNames.append(str(i.name()))

        for l in layersNames:
            self.toggleLayer(l, 0)

        self.toggleLayer('USA_Topo_Maps', 1)
        # QgsMessageLog.logMessage('USA_Topo_Maps...', 'sGIS', level=Qgis.Info)
        self.toggleLayer('jobs', 1)
        self.toggleLayer('ng911rdss', 1)

        QgsProject.instance().layerTreeRoot().findGroup('IMAGERY').setItemVisibilityChecked(1)
        QgsProject.instance().layerTreeRoot().findGroup('State Orthos').setItemVisibilityChecked(1)

        self.iface.mapCanvas().zoomScale(8000)

        self.vl.loadNamedStyle(prn)
        self.vl.triggerRepaint()
        self.make_jpg(dwgpath, jobNo)

        self.vl.loadNamedStyle(std)

        for l in layersNames:
            self.toggleLayer(l, 1)

        self.toggleLayer('USA_Topo_Maps', 0)
        self.vl.triggerRepaint()
        self.iface.mapCanvas().zoomScale(currentScale)
        self.vl.setSubsetString('"supp_type"=\'%s\'' % 'X')
        self.vl.selectByIds(ids)
        self.reset()
        resetLegend(self)
        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(self.vl)

    def make_jpg(self, cf, jn):

        project = QgsProject.instance()
        l = QgsPrintLayout(project)
        l.initializeDefaults()
        l.setUnits(QgsUnitTypes.LayoutMillimeters)
        page = l.pageCollection().pages()[0]

        paperSize = self.setPaperSizeLandscape()
        page.setPageSize(QgsLayoutSize(paperSize[0], paperSize[1]))

        lm = 10  # left margin
        tm = 10  # upper margin
        bm = 20  # lower margin

        refSize = paperSize[0]
        if paperSize[1] < refSize:
            refSize = paperSize[1]

        # add map
        x, y = lm, tm
        w, h = paperSize[0] - 2 * lm, paperSize[1] - bm

        theMap = QgsLayoutItemMap(l)
        theMap.updateBoundingRect()
        theMap.setRect(QRectF(x, y, w, h))
        theMap.setPos(x, y)
        theMap.setFrameEnabled(True)

        theMap.setLayers(project.mapThemeCollection().masterVisibleLayers())  # remember ANNOTATION!
        theMap.setExtent(self.iface.mapCanvas().extent())
        theMap.attemptSetSceneRect(QRectF(x, y, w, h))
        l.addItem(theMap)

        exporter = QgsLayoutExporter(l)
        jpg_settings = exporter.ImageExportSettings()  # dpi?

        res = exporter.exportToImage(cf + '\\site.jpg', jpg_settings)
        if res != QgsLayoutExporter.Success:
            res = exporter.exportToImage(cf + '\\site-1.jpg', jpg_settings)

    def toggleLayer(self, layer, status):
        # QgsMessageLog.logMessage('TOGGLE: ' + layer + '...', 'sGIS', level=Qgis.Info)
        lyr = QgsProject.instance().mapLayersByName(layer)[0]
        if status == 0:
            QgsProject.instance().layerTreeRoot().findLayer(lyr.id()).setItemVisibilityChecked(False)
        else:
            QgsProject.instance().layerTreeRoot().findLayer(lyr.id()).setItemVisibilityChecked(True)

    def setPaperSizePortrait(self):

        longSide = 297
        shortSide = 210
        width = shortSide
        height = longSide

        return width, height

    def setPaperSizeLandscape(self):

        longSide = 210
        shortSide = 297
        width = shortSide
        height = longSide

        return width, height

    def reset(self):

        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(self.vl)
        jform = os.path.dirname(os.path.realpath(__file__)) + '/forms/jobs.ui'
        jpy = os.path.dirname(os.path.realpath(__file__)) + '/forms/jobs_init.py'
        form_config = self.iface.activeLayer().editFormConfig()
        form_config.setUiForm(jform)
        form_config.setInitFilePath(jpy)
        self.iface.activeLayer().setEditFormConfig(form_config)

        self.vl = QgsProject.instance().mapLayersByName('plans')[0]
        self.iface.setActiveLayer(self.vl)
        jform = os.path.dirname(os.path.realpath(__file__)) + '/forms/plans.ui'
        jpy = os.path.dirname(os.path.realpath(__file__)) + '/forms/plans_init.py'
        form_config = self.iface.activeLayer().editFormConfig()
        form_config.setUiForm(jform)
        form_config.setInitFilePath(jpy)
        self.iface.activeLayer().setEditFormConfig(form_config)

        self.vl = QgsProject.instance().mapLayersByName('contacts')[0]
        self.iface.setActiveLayer(self.vl)
        jform = os.path.dirname(os.path.realpath(__file__)) + '/forms/contacts.ui'
        jpy = os.path.dirname(os.path.realpath(__file__)) + '/forms/contacts_init.py'
        form_config = self.iface.activeLayer().editFormConfig()
        form_config.setUiForm(jform)
        form_config.setInitFilePath(jpy)
        self.iface.activeLayer().setEditFormConfig(form_config)

        self.vl = QgsProject.instance().mapLayersByName('supplementals')[0]
        self.iface.setActiveLayer(self.vl)
        jform = os.path.dirname(os.path.realpath(__file__)) + '/forms/supplementals.ui'
        jpy = os.path.dirname(os.path.realpath(__file__)) + '/forms/supplementals_init.py'
        form_config.setUiForm(jform)
        form_config.setInitFilePath(jpy)
        self.iface.activeLayer().setEditFormConfig(form_config)
        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(self.vl)

class sgis_search(object):

    def __init__(self, iface):
        self.iface = iface

    def run(self):
        eMenu = self.iface.mainWindow()

        vLayer = self.iface.activeLayer()
        oLayer = vLayer

        if vLayer.name() == 'jobs':

            # !!! this no longer works in v3.40.1 - need to revisit hiding certain columns during search.
            # !!! issue being tracked in gitHub: xxxxxxxx
            
            pLayer = QgsProject.instance().mapLayersByName('jobs')[0]
            cLayer = QgsProject.instance().mapLayersByName('contacts')[0]
            pLayer = pLayer.id()
            cLayer = cLayer.id()

            self.killRelation(pLayer, cLayer)
            
            cols = ('job_no','rev_no','job_type','jobSubtype','map_bk_lot','old_plan_no','job_desc','client_name','folder_name',
                    'client_role','contact_type','contact_addr','client_name','locus_addr','town','planbook_page',
                    'estimate','active','pins_set','date_recorded', 'completed')

            columns = vLayer.fields()
            editor_widget_setup = QgsEditorWidgetSetup('Hidden', {})

            for c in columns:
                if str(c.name()) in cols:
                    pass
                else:
                    pass
                    # fieldIndex = vLayer.fields().indexFromName(c.name())
                    # vLayer.setEditorWidgetSetup(fieldIndex, editor_widget_setup)

        elif vLayer.name() == 'supplementals':
            
            cols = ('job_no','rev_no','job_type','map_bk_lot','old_plan_no','job_desc','client_name','folder_name',
                    'client_role','contact_type','contact_addr','client_name','locus_addr','town','planbook_page',
                    'estimate','active','pins_set','date_recorded','old_plan','job','folder_type','supp_type',
                    'document_subtype','design_type','map_type','map_subtype','pls_no','author', 'completed')

            columns = vLayer.fields()
            editor_widget_setup = QgsEditorWidgetSetup('Hidden', {})

            for c in columns:
                if str(c.name()) in cols:
                    pass
                else:
                    pass
                    # fieldIndex = vLayer.fields().indexFromName(c.name())
                    # vLayer.setEditorWidgetSetup(fieldIndex, editor_widget_setup)

        elif vLayer.name() == 'plans':
            cols = ('plan_no','map_bk_lot','name','address','town','county','job','date','surveyor',
                    'plan_type','size_no')

            columns = vLayer.fields()
            editor_widget_setup = QgsEditorWidgetSetup('Hidden', {})

            for c in columns:
                if str(c.name()) in cols:
                    pass
                else:
                    pass
                    # fieldIndex = vLayer.fields().indexFromName(c.name())
                    # vLayer.setEditorWidgetSetup(fieldIndex, editor_widget_setup)

        elif vLayer.name() == 'ng911rdss':
            cols = ('RDNAME','STREETNAME','PREDIR','SUFFIX','POSTDIR','TOWN','CITY','RCOUNTY','ROUTE_NUM','ONEWAY',
                    'SPEED','RDCLASS')

            columns = vLayer.fields()
            editor_widget_setup = QgsEditorWidgetSetup('Hidden', {})

            for c in columns:
                if str(c.name()) in cols:
                    pass
                else:
                    pass
                    # fieldIndex = vLayer.fields().indexFromName(c.name())
                    # vLayer.setEditorWidgetSetup(fieldIndex, editor_widget_setup)

        elif vLayer.name() == 'Parcels':
            cols = ('town', 'county', 'map_bk_lot','locus_address', 'owner1','rawdeeds', 'proplocnum', 'prop_loc')
            columns = vLayer.fields()
            editor_widget_setup = QgsEditorWidgetSetup('Hidden', {})

            for c in columns:
                if str(c.name()) in cols:
                    pass
                else:
                    pass
                    # fieldIndex = vLayer.fields().indexFromName(c.name())
                    # vLayer.setEditorWidgetSetup(fieldIndex, editor_widget_setup)
                    
        elif vLayer.name() == 'parcels_aux':
            cols = ('town', 'county', 'map_bk_lot','locus_address', 'owner1','rawdeeds', 'proplocnum', 'prop_loc')
            columns = vLayer.fields()
            editor_widget_setup = QgsEditorWidgetSetup('Hidden', {})

            for c in columns:
                if str(c.name()) in cols:
                    pass
                else:
                    pass
                    # fieldIndex = vLayer.fields().indexFromName(c.name())
                    # vLayer.setEditorWidgetSetup(fieldIndex, editor_widget_setup)

        else:
            pass

        form_config = self.iface.activeLayer().editFormConfig()
        #form_config.setInitCodeSource(1)
        self.iface.activeLayer().setEditFormConfig(form_config)

        try:
            for a in eMenu.findChildren(QAction, 'mActionSelectByForm'):
                a.trigger()
        except Exception as e:
            pass

        try:
            self.setRelation(pLayer,cLayer)
        except Exception as e:
            pass

        self.iface.setActiveLayer(oLayer)
        resetLegend(self)

    def setRelation(self, pLayer, cLayer):
        rel = QgsRelation()
        rel.setReferencingLayer(cLayer)
        rel.setReferencedLayer(pLayer)
        rel.addFieldPair('jobs_id', 'sid')
        rel.setId('fk_jobs_contacts')
        rel.setName('Job Contacts')
        # rel.isValid() # It will only be added if it is valid. If not, check the ids and field names
        QgsProject.instance().relationManager().addRelation(rel)

    def killRelation(self, pLayer, cLayer):
        rel = QgsRelation()
        rel.setReferencingLayer(cLayer)
        rel.setReferencedLayer(pLayer)
        rel.addFieldPair('jobs_id', 'sid')
        rel.setId('fk_jobs_contacts')
        rel.setName('Job Contacts')
        # rel.isValid() # It will only be added if it is valid. If not, check the ids and field names
        QgsProject.instance().relationManager().removeRelation(rel)


class sgis_bulkMapExport(object):
    newJob = 0
    selComp = 0
    multiFeat = 0
    count = 0

    def __init__(self, iface):
        self.iface = iface

    def initGui(self):

        self.action = QAction("&Export map_bk_lot by Layer", self.iface.mainWindow())
        self.action.triggered.connect(self.run)
        self.action.trigger()
        self.iface.mapCanvas().selectionChanged.connect(self.select_changed)

    def run(self):
        self.vl = QgsProject.instance().mapLayersByName('Parcels')[0]
        self.iface.setActiveLayer(self.vl)

        reply = QMessageBox.question(self.iface.mainWindow(), 'Select Feature(s)',
                                     'Click OK and select the correct parcel(s) for the new job.',
                                     QMessageBox.Ok, QMessageBox.Cancel)
        if reply == QMessageBox.Ok:
            if self.selComp == 1:
                return
            else:
                self.newJob = 1
                if self.count == 0:
                    for a in self.iface.mainWindow().children():
                        if a.objectName() == 'mActionDeselectAll':
                            a.trigger()
                            QgsMessageLog.logMessage('FIRST RUN: Previously selected parcel(s) have been cleared.',
                                                     'sGIS', level=Qgis.Info)
                            QgsMessageLog.logMessage('Selection starting...', 'sGIS', level=Qgis.Info)
                            self.iface.actionSelectFreehand().trigger()
                            self.count = self.count + 1
                else:
                    self.iface.actionSelect().trigger()

        else:
            QgsMessageLog.logMessage('DEBUG: Job selection cancelled.', 'sGIS', level=Qgis.Info)

    def select_changed(self):

        if self.newJob == 0:
            return
        else:
            vLayer = self.iface.activeLayer()
            feats_count = vLayer.selectedFeatureCount()

            msg = QMessageBox()
            msg.setWindowTitle('Selection')
            msg.setText(str(feats_count) + ' features have been selected. Continue?')
            create = msg.addButton('Continue', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(create)
            msg.exec_()
            msg.deleteLater()

            if msg.clickedButton() is create:
                self.iface.actionCopyFeatures().trigger()

                self.newJob = 0
                self.tmpLayer = QgsVectorLayer('MultiPolygon', 'temp', 'memory')
                QgsProject.instance().addMapLayers([self.tmpLayer])
                self.iface.setActiveLayer(self.tmpLayer)
                self.iface.actionIdentify().trigger()
                self.iface.actionToggleEditing().trigger()
                self.layerData = self.tmpLayer.dataProvider()
                self.layerData.addAttributes(
                    [QgsField("gid", QVariant.String), QgsField("map_bk_lot", QVariant.String)])
                self.iface.activeLayer().commitChanges()
                self.iface.actionToggleEditing().trigger()

                QgsMessageLog.logMessage('Pasting features to ' + str(self.tmpLayer.name()) + '...', 'sGIS',
                                         level=Qgis.Info)
                self.iface.actionPasteFeatures().trigger()
                self.iface.activeLayer().commitChanges()


            elif msg.clickedButton() is cancel:
                self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                return
            self.active_edit(self.tmpLayer)

    def active_edit(self, layer):

        self.iface.setActiveLayer(layer)
        self.features = self.iface.activeLayer().selectedFeatures()
        self.allmaps = ''
        for feature in layer.getFeatures():
            map_bk_lot = feature["map_bk_lot"]
            map_bk_lot = '\'' + map_bk_lot + '\''
            self.allmaps = str(self.allmaps) + str(map_bk_lot) + ','

        self.allmaps = self.allmaps[:-1]
        QgsMessageLog.logMessage('allmaps: ' + self.allmaps, 'sGIS',
                                 level=Qgis.Info)
        pyperclip.copy(self.allmaps)

        QgsMessageLog.logMessage('All map_bk_lot values copied to clipboard...paste as you will.', 'sGIS',
                                 level=Qgis.Info)
        QgsProject.instance().removeMapLayer(layer.id())

        def finished(self):
            self.done(1)


class sgis_planImportXLSX(object):

    def __init__(self, iface):
        self.iface = iface

    def initGui(self):

        self.action = QAction("&Plan Import from .XLSX", self.iface.mainWindow())
        self.action.triggered.connect(self.run)
        self.action.trigger()

    def run(self):
        self.vl = QgsProject.instance().mapLayersByName('Parcels')[0]
        self.iface.setActiveLayer(self.vl)
        cLayer = self.iface.mapCanvas().currentLayer()

        import openpyxl
        #if retained, reqork required
        path = "C:\\sGIS\\WIP.xlsx"
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active

        max_col = sheet_obj.max_column
        max_row = sheet_obj.max_row
        planData = []
        allPlans = []
        error_count = 0

        for j in range(2, max_row + 1):

            for i in range(1, max_col + 1):
                cell_obj = sheet_obj.cell(row=j, column=i)
                planData.append(str(cell_obj.value))
            allPlans.append(planData)
            planData = []

        for plan in allPlans:

            idValue = plan[0]
            size_num = plan[1]
            file_num = plan[2]
            plan_no = plan[3]
            name = plan[4]
            address = plan[5]
            townLA = plan[6]
            job = plan[7]
            date = plan[8]
            if str(date) == 'None':
                date = '1900/01/01'
            initials = plan[9]
            map = plan[10]
            lot = plan[11]
            map_bk_lot = plan[12]
            surveyor = plan[13]
            notes = plan[14]
            cd_no = plan[15]
            latitude = plan[16]
            longitude = plan[17]
            sTown = self.getTown(townLA)

            maps = map_bk_lot.split(",")

            cLayer.setSubsetString("town = '{}' ".format(sTown))
            it = cLayer.getFeatures()
            ids = [i.id() for i in it]

            for x in range(0, len(maps)):

                try:
                    it2 = cLayer.getFeatures(
                        QgsFeatureRequest().setFilterExpression(u'"map_bk_lot" = {0}'.format(maps[x])))
                    newIds = [i2.id() for i2 in it2]
                    idsToSel = list(set(ids).intersection(newIds))

                    cLayer.selectByIds(idsToSel, 1)
                    features = cLayer.selectedFeatures()
                    provider = cLayer.dataProvider()
                    spIndex = QgsSpatialIndex()

                    feat = QgsFeature()
                    fit = provider.getFeatures()

                    while fit.nextFeature(feat):
                        spIndex.insertFeature(feat)

                    try:
                        geometry = features[0].geometry()
                    except Exception as e:
                        return

                    centroid = geometry.centroid().asPoint()
                    nearestIds = spIndex.nearestNeighbor(centroid, 1)

                    geom = None
                    for feat in features:

                        for f in nearestIds:
                            if feat['objectid'] == f:
                                pass
                            else:
                                zipcode = feat['zipcode']
                                county = feat['county']
                                l_l = feat['lat_lon']

                        if geom == None:
                            geom = feat.geometry()

                        else:
                            geom = geom.combine(feat.geometry())

                except Exception as e:
                    self.vl = QgsProject.instance().mapLayersByName('Parcels')[0]
                    self.vl.setSubsetString('gid > 1')
                    self.vl.selectByIds([])
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                         "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                             exc_tb.tb_lineno) + ' ' + str(e))
                    QGuiApplication.restoreOverrideCursor()
                    return

            cLayer.selectByIds([])

            self.vl = QgsProject.instance().mapLayersByName('plans')[0]
            self.iface.setActiveLayer(self.vl)
            self.iface.actionToggleEditing().trigger()
            feature = QgsFeature()

            try:

                sourceCrs = QgsCoordinateReferenceSystem(102683)
                destCrs = QgsCoordinateReferenceSystem(102683)
                tr = QgsCoordinateTransform(sourceCrs, destCrs, QgsProject.instance())
                geom.transform(tr)
                feature.setGeometry(geom)

                dataProvider = self.vl.dataProvider()
                dataProvider.addFeature(feature)
                self.iface.activeLayer().commitChanges()
                self.vl.dataProvider().forceReload()
                self.iface.mapCanvas().refresh()
                self.selectLastFeature()

                newF = self.vl.selectedFeatures()[0]
                self.iface.actionToggleEditing().trigger()

                idx = dataProvider.fieldNameIndex('idValue')
                self.vl.changeAttributeValue(newF.id(), idx, int(idValue))
                idx = dataProvider.fieldNameIndex('size_no')
                self.vl.changeAttributeValue(newF.id(), idx, size_num)
                idx = dataProvider.fieldNameIndex('file_no')
                self.vl.changeAttributeValue(newF.id(), idx, file_num)
                idx = dataProvider.fieldNameIndex('plan_no')
                self.vl.changeAttributeValue(newF.id(), idx, plan_no)
                idx = dataProvider.fieldNameIndex('name')
                self.vl.changeAttributeValue(newF.id(), idx, name)
                idx = dataProvider.fieldNameIndex('address')
                self.vl.changeAttributeValue(newF.id(), idx, address)
                idx = dataProvider.fieldNameIndex('town')
                self.vl.changeAttributeValue(newF.id(), idx, townLA)
                idx = dataProvider.fieldNameIndex('town_Parcels')
                self.vl.changeAttributeValue(newF.id(), idx, sTown)
                idx = dataProvider.fieldNameIndex('job')
                self.vl.changeAttributeValue(newF.id(), idx, job)
                idx = dataProvider.fieldNameIndex('date')
                self.vl.changeAttributeValue(newF.id(), idx, date)
                idx = dataProvider.fieldNameIndex('initials')
                self.vl.changeAttributeValue(newF.id(), idx, initials)
                idx = dataProvider.fieldNameIndex('map')
                self.vl.changeAttributeValue(newF.id(), idx, map)
                idx = dataProvider.fieldNameIndex('lot')
                self.vl.changeAttributeValue(newF.id(), idx, lot)
                idx = dataProvider.fieldNameIndex('map_bk_lot')
                self.vl.changeAttributeValue(newF.id(), idx, map_bk_lot)
                idx = dataProvider.fieldNameIndex('surveyor')
                self.vl.changeAttributeValue(newF.id(), idx, surveyor)
                idx = dataProvider.fieldNameIndex('notes')
                self.vl.changeAttributeValue(newF.id(), idx, notes)
                idx = dataProvider.fieldNameIndex('cd_no')
                self.vl.changeAttributeValue(newF.id(), idx, cd_no)
                idx = dataProvider.fieldNameIndex('latitude')
                self.vl.changeAttributeValue(newF.id(), idx, latitude)
                idx = dataProvider.fieldNameIndex('longitude')
                self.vl.changeAttributeValue(newF.id(), idx, longitude)
                self.vl.updateFields()
                self.iface.activeLayer().commitChanges()
                geom = None

            except Exception as e:

                error_count = error_count + 1
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                     "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                         exc_tb.tb_lineno) + ' ' + str(e))
                QgsMessageLog.logMessage('ERROR: | ' + str(plan_no) + ' with ' + str(maps[x]), 'sGIS',
                                         level=Qgis.Info)
                self.iface.actionToggleEditing().trigger()
                QGuiApplication.restoreOverrideCursor()

        self.vl = QgsProject.instance().mapLayersByName('Parcels')[0]
        self.vl.setSubsetString('gid > 1')
        self.vl.selectByIds([])

        QgsMessageLog.logMessage('DONE.', 'sGIS', level=Qgis.Info)

    def getTown(self, town):

        if town == 'ARROW':
            town = 'Arrowsic'
        elif town == 'AUGUSTA':
            town = 'Augusta'
        elif town == 'ALNA':
            town = 'Alna'
        elif town == 'EDG':
            town = 'Edgecomb'
        elif town == 'EDGE':
            town = 'Edgecomb'
        elif town == 'BBY':
            town = 'Boothbay'
        elif town == 'BBH':
            town = 'Boothbay Harbor'
        elif town == 'WBATH':
            town = 'West Bath'
        elif town == 'SPT':
            town = 'Southport'
        elif town == 'BRISTOL':
            town = 'Bristol'
        elif town == 'BOWD':
            town = 'Bowdoinham'
        elif town == 'ELIOT':
            town = 'Eliot'
        elif town == 'BRUN':
            town = 'Brunswick'
        elif town == 'CAMDEN':
            town = 'Camden'
        elif town == 'BRIS':
            town = 'Bristol'
        elif town == 'CHRCOVE':
            town = 'South Bristol'
        elif town == 'S.BRIS':
            town = 'South Bristol'
        elif town == 'SBRIS':
            town = 'South Bristol'
        elif town == 'BATH':
            town = 'Bath'
        elif town == 'PHIP':
            town = 'Phippsburg'
        elif town == 'HARPS':
            town = 'Harpswell'
        elif town == 'BRUNS':
            town = 'Brunswick'
        elif town == 'FALMTH':
            town = 'Falmouth'
        elif town == 'FRNSHP':
            town = 'Friendship'
        elif town == 'FRPT':
            town = 'Freeport'
        elif town == 'GEO':
            town = 'Georgetown'
        elif town == 'GLDSBORO':
            town = 'Gouldsboro'
        elif town == 'Milbridge':
            town = 'Milbridge'
        elif town == 'MONMTH':
            town = 'Monmouth'
        elif town == 'NEWHRBR':
            town = 'Bristol'
        elif town == 'POWNL':
            town = 'Pownal'
        elif town == 'PORTLAND':
            town = 'Portland'
        elif town == 'READFIELD':
            town = 'Readfield'
        elif town == 'RICH':
            town = 'Richmond'
        elif town == 'RKPT':
            town = 'Rockport'
        elif town == 'RNDPND':
            town = 'Bristol'
        elif town == 'CUMB':
            town = 'Cumberland'
        elif town == 'St.GEO':
            town = 'Saint George'
        elif town == 'STGEO':
            town = 'Saint George'
        elif town == 'WALPOLE':
            town = 'Walpole'
        elif town == 'WATERBORO':
            town = 'Waterboro'
        elif town == 'WHTF':
            town = 'Whitefield'
        elif town == 'WTVL':
            town = 'Waterville'
        elif town == 'YAR':
            town = 'Yarmouth'
        elif town == 'WISC':
            town = 'Wiscasset'
        elif town == 'TOPS':
            town = 'Topsham'
        elif town == 'WOOL':
            town = 'Woolwich'
        elif town == 'WALDO':
            town = 'Waldoboro'
        elif town == 'Rnd Pond':
            town = 'Round Pond'
        elif town == 'JEFF':
            town = 'Jefferson'
        elif town == 'GEOTN':
            town = 'Georgetown'
        elif town == 'NOBLE':
            town = 'Nobleboro'
        elif town == 'NOB':
            town = 'Nobleboro'
        elif town == 'OWLHD':
            town = 'Owls Head'
        elif town == 'Bowdoin':
            pass
        elif town == 'DRES':
            town = 'Dresden'
        elif town == 'DAM':
            town = 'Damariscotta'
        elif town == 'NEWC':
            town = 'Newcastle'
        elif town == 'WESTP':
            town = 'Westport Island'
        elif town == 'WEST':
            town = 'Westport Island'
        elif town == 'BREM':
            town = 'Bremen'
        elif town == 'GARD':
            town = 'Gardiner'
        elif town == 'MONHE':
            town = 'Monhegan'
        elif town == 'POPHAM':
            town = 'Popham'
        elif town == 'BRIST':
            town = 'Bristol'
        elif town == 'ROCK':
            town = 'Rockland'
        elif town == 'LISBONFALLS':
            town = 'Lisbon'
        elif town == 'WHIT':
            town = 'Whiting'
        elif town == 'WATERBORO':
            town = 'Waterboro'
        else:
            pass
        return town

    def selectLastFeature(self):

        f2 = self.vl.getFeatures()
        fCount = self.vl.featureCount()

        fIds = []
        fIds = [f.id() for f in f2]
        fIds.sort()

        fId = [fIds[-1]]
        self.vl.selectByIds(fId)


class sgis_suppImportXLSX(object):

    def __init__(self, iface):
        self.iface = iface

    def initGui(self):

        self.action = QAction("&Supplemental Import from .XLSX", self.iface.mainWindow())
        self.action.triggered.connect(self.run)
        self.action.trigger()

    def run(self):
        global idsToSel
        self.vl = QgsProject.instance().mapLayersByName('Parcels')[0]
        self.iface.setActiveLayer(self.vl)
        cLayer = self.iface.mapCanvas().currentLayer()

        import openpyxl
        # if retained, rework required
        path = "C:\\sGIS\\WIP.xlsx"
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active

        max_col = sheet_obj.max_column
        max_row = sheet_obj.max_row
        planData = []
        allPlans = []
        error_count = 0

        for j in range(2, max_row + 1):

            for i in range(1, max_col + 1):
                cell_obj = sheet_obj.cell(row=j, column=i)
                planData.append(str(cell_obj.value))
            allPlans.append(planData)
            planData = []

        for plan in allPlans:

            idValue = plan[0]
            size_num = plan[1]
            file_num = plan[2]
            job_no = plan[3]
            name = plan[4]
            address = plan[5]
            townLA = plan[6]
            job_type = plan[7]
            date = plan[8]
            QgsMessageLog.logMessage(str(date), 'sGIS', level=Qgis.Info)
            if str(date) == 'None':
                date = '1900/01/01'

            initials = plan[9]
            map = plan[10]
            lot = plan[11]
            map_bk_lot = plan[12]
            surveyor = plan[13]
            notes = plan[14]
            cd_no = plan[15]
            latitude = plan[16]
            longitude = plan[17]
            sTown = self.getTown(townLA)

            maps = map_bk_lot.split(",")

            cLayer.setSubsetString("town = '{}' ".format(sTown))
            it = cLayer.getFeatures()
            ids = [i.id() for i in it]

            for x in range(0, len(maps)):

                expr2 = QgsExpression("map_bk_lot = '{}' ".format(maps[x]))

                try:
                    it2 = cLayer.getFeatures(
                        QgsFeatureRequest().setFilterExpression(u'"map_bk_lot" = {0}'.format(maps[x])))
                    newIds = [i2.id() for i2 in it2]
                    idsToSel = list(set(ids).intersection(newIds))
                    cLayer.selectByIds(idsToSel, 1)
                    features = cLayer.selectedFeatures()
                    provider = cLayer.dataProvider()

                    spIndex = QgsSpatialIndex()

                    feat = QgsFeature()
                    fit = provider.getFeatures()

                    while fit.nextFeature(feat):
                        spIndex.insertFeature(feat)

                    geometry = features[0].geometry()
                    centroid = geometry.centroid().asPoint()

                    nearestIds = spIndex.nearestNeighbor(centroid, 1)

                    geom = None
                    for feat in features:

                        for f in nearestIds:
                            if feat['objectid'] == f:
                                pass
                            else:
                                zipcode = feat['zipcode']
                                county = feat['county']
                                l_l = feat['lat_lon']

                        if geom == None:
                            geom = feat.geometry()

                        else:
                            geom = geom.combine(feat.geometry())

                except Exception as e:
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                         "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                             exc_tb.tb_lineno) + ' ' + str(e))
                    QGuiApplication.restoreOverrideCursor()
                    return

            cLayer.selectByIds([])

            self.vl = QgsProject.instance().mapLayersByName('supplementals')[0]
            self.vl.setSubsetString('')
            self.iface.setActiveLayer(self.vl)
            feature = QgsFeature()
            sourceCrs = QgsCoordinateReferenceSystem(102683)
            destCrs = QgsCoordinateReferenceSystem(102683)
            tr = QgsCoordinateTransform(sourceCrs, destCrs, QgsProject.instance())
            geom.transform(tr)
            feature.setGeometry(geom)

            dataProvider = self.vl.dataProvider()
            dataProvider.addFeature(feature)
            self.iface.activeLayer().commitChanges()
            self.vl.dataProvider().forceReload()
            self.iface.mapCanvas().refresh()
            self.selectLastFeature()
            newF = self.vl.selectedFeatures()[0]

            try:
                lat_lon = formatLL(l_l)
            except Exception as e:
                lat_lon = ''
                pass

            self.iface.actionToggleEditing().trigger()

            sid = newF['sid']

            idx = dataProvider.fieldNameIndex('job_no')
            self.vl.changeAttributeValue(newF.id(), idx, job_no)
            idx = dataProvider.fieldNameIndex('client_name')
            self.vl.changeAttributeValue(newF.id(), idx, name)
            idx = dataProvider.fieldNameIndex('locus_addr')
            self.vl.changeAttributeValue(newF.id(), idx, address)
            idx = dataProvider.fieldNameIndex('town')
            self.vl.changeAttributeValue(newF.id(), idx, sTown)
            idx = dataProvider.fieldNameIndex('county')
            self.vl.changeAttributeValue(newF.id(), idx, county)
            idx = dataProvider.fieldNameIndex('zipcode')
            self.vl.changeAttributeValue(newF.id(), idx, zipcode)
            idx = dataProvider.fieldNameIndex('date_recorded')
            self.vl.changeAttributeValue(newF.id(), idx, date)
            idx = dataProvider.fieldNameIndex('map_bk_lot')
            self.vl.changeAttributeValue(newF.id(), idx, map_bk_lot)
            idx = dataProvider.fieldNameIndex('supp_type')
            self.vl.changeAttributeValue(newF.id(), idx, 'K')
            idx = dataProvider.fieldNameIndex('folder_type')
            self.vl.changeAttributeValue(newF.id(), idx, surveyor)
            idx = dataProvider.fieldNameIndex('job_desc')
            self.vl.changeAttributeValue(newF.id(), idx, 'Notes: ' + notes + ' | JOB: ' + job_type)
            idx = dataProvider.fieldNameIndex('job_type')
            self.vl.changeAttributeValue(newF.id(), idx, job_type)
            idx = dataProvider.fieldNameIndex('objectid')
            self.vl.changeAttributeValue(newF.id(), idx, idsToSel[0])
            idx = dataProvider.fieldNameIndex('plan_no')
            self.vl.changeAttributeValue(newF.id(), idx, job_no)
            idx = dataProvider.fieldNameIndex('old_plan')
            self.vl.changeAttributeValue(newF.id(), idx, job_no + ' (' + job_type + ')')
            idx = dataProvider.fieldNameIndex('lat_lon')
            self.vl.changeAttributeValue(newF.id(), idx, lat_lon)

            self.vl.updateFields()
            self.iface.activeLayer().commitChanges()
            self.vl.dataProvider().forceReload()
            self.iface.mapCanvas().refresh()
            geom = None

        self.vl = QgsProject.instance().mapLayersByName('supplementals')[0]
        self.iface.setActiveLayer(self.vl)
        self.vl.setSubsetString("supp_type != 'X'")
        cLayer.setSubsetString("gid > 1")

        QgsMessageLog.logMessage('DONE.', 'sGIS', level=Qgis.Info)

    def feature_added(self, featureAdded):
        layer = self.iface.activeLayer()
        layer.featureAdded.disconnect()
        layer.select(featureAdded)
        layer.commitChanges()
        self.selectLastFeature(layer)

    def getTown(self, town):
        if town == 'ARROW':
            town = 'Arrowsic'
        elif town == 'AUGUSTA':
            town = 'Augusta'
        elif town == 'ALNA':
            town = 'Alna'
        elif town == 'EDG':
            town = 'Edgecomb'
        elif town == 'EDGE':
            town = 'Edgecomb'
        elif town == 'BBY':
            town = 'Boothbay'
        elif town == 'BBH':
            town = 'Boothbay Harbor'
        elif town == 'WBATH':
            town = 'West Bath'
        elif town == 'SPT':
            town = 'Southport'
        elif town == 'BRISTOL':
            town = 'Bristol'
        elif town == 'BOWD':
            town = 'Bowdoinham'
        elif town == 'ELIOT':
            town = 'Eliot'
        elif town == 'BRUN':
            town = 'Brunswick'
        elif town == 'CAMDEN':
            town = 'Camden'
        elif town == 'BRIS':
            town = 'Bristol'
        elif town == 'CHRCOVE':
            town = 'South Bristol'
        elif town == 'S.BRIS':
            town = 'South Bristol'
        elif town == 'SBRIS':
            town = 'South Bristol'
        elif town == 'BATH':
            town = 'Bath'
        elif town == 'PHIP':
            town = 'Phippsburg'
        elif town == 'HARPS':
            town = 'Harpswell'
        elif town == 'BRUNS':
            town = 'Brunswick'
        elif town == 'FALMTH':
            town = 'Falmouth'
        elif town == 'FRNSHP':
            town = 'Friendship'
        elif town == 'FRPT':
            town = 'Freeport'
        elif town == 'GEO':
            town = 'Georgetown'
        elif town == 'GLDSBORO':
            town = 'Gouldsboro'
        elif town == 'Milbridge':
            town = 'Milbridge'
        elif town == 'MONMTH':
            town = 'Monmouth'
        elif town == 'NEWHRBR':
            town = 'Bristol'
        elif town == 'POWNL':
            town = 'Pownal'
        elif town == 'PORTLAND':
            town = 'Portland'
        elif town == 'READFIELD':
            town = 'Readfield'
        elif town == 'RICH':
            town = 'Richmond'
        elif town == 'RKPT':
            town = 'Rockport'
        elif town == 'RNDPND':
            town = 'Bristol'
        elif town == 'CUMB':
            town = 'Cumberland'
        elif town == 'St.GEO':
            town = 'Saint George'
        elif town == 'WALPOLE':
            town = 'Walpole'
        elif town == 'WATERBORO':
            town = 'Waterboro'
        elif town == 'WHTF':
            town = 'Whitefield'
        elif town == 'WTVL':
            town = 'Waterville'
        elif town == 'YAR':
            town = 'Yarmouth'
        elif town == 'WISC':
            town = 'Wiscasset'
        elif town == 'TOPS':
            town = 'Topsham'
        elif town == 'WOOL':
            town = 'Woolwich'
        elif town == 'WALDO':
            town = 'Waldoboro'
        elif town == 'Rnd Pond':
            town = 'Round Pond'
        elif town == 'JEFF':
            town = 'Jefferson'
        elif town == 'GEOTN':
            town = 'Georgetown'
        elif town == 'NOBLE':
            town = 'Nobleboro'
        elif town == 'NOB':
            town = 'Nobleboro'
        elif town == 'OWLHD':
            town = 'Owls Head'
        elif town == 'Bowdoin':
            pass
        elif town == 'DRES':
            town = 'Dresden'
        elif town == 'DAM':
            town = 'Damariscotta'
        elif town == 'NEWC':
            town = 'Newcastle'
        elif town == 'WESTP':
            town = 'Westport Island'
        elif town == 'WEST':
            town = 'Westport Island'
        elif town == 'BREM':
            town = 'Bremen'
        elif town == 'GARD':
            town = 'Gardiner'
        elif town == 'MONHE':
            town = 'Monhegan'
        elif town == 'POPHAM':
            town = 'Popham'
        elif town == 'BRIST':
            town = 'Bristol'
        elif town == 'ROCK':
            town = 'Rockland'
        elif town == 'LISBONFALLS':
            town = 'Lisbon'
        else:
            pass
        return town

    def selectLastFeature(self):

        f2 = self.vl.getFeatures()
        fCount = self.vl.featureCount()

        fIds = []
        fIds = [f.id() for f in f2]
        fIds.sort()

        fId = [fIds[-1]]
        self.vl.selectByIds(fId)


class sgis_jobImportXLSX(object):

    def __init__(self, iface):

        self.iface = iface

    def initGui(self):

        self.action = QAction("&Job Import from .XLSX", self.iface.mainWindow())
        self.action.triggered.connect(self.run)
        self.action.trigger()

    def run(self):

        self.vl = QgsProject.instance().mapLayersByName('Parcels')[0]
        self.iface.setActiveLayer(self.vl)
        cLayer = self.iface.mapCanvas().currentLayer()

        jLayer = QgsProject.instance().mapLayersByName('jobs')[0]
        # path = "C:\\sGIS\\jobsWIP.xlsx"

        try:
            qfd = QFileDialog()
            title = 'Open File'
            # if retained, rework required
            path = "C:\\s_GIS\\"
            f = QFileDialog.getOpenFileName(qfd, title, path)
            path = f[0]
            wb_obj = openpyxl.load_workbook(path)

        except Exception:
            return

        sheet_obj = wb_obj.active
        max_col = sheet_obj.max_column
        max_row = sheet_obj.max_row
        jobData = []
        allJobs = []
        allIds = []
        error_count = 0
        QgsMessageLog.logMessage('LOADING | ' + str(path) + '...', 'sGIS', level=Qgis.Info)
        recs = max_row - 1
        QgsMessageLog.logMessage('Records found: ' + str(recs), 'sGIS', level=Qgis.Info)

        if recs == 0:
            return

        for j in range(2, max_row + 1):

            for i in range(1, max_col + 1):
                cell_obj = sheet_obj.cell(row=j, column=i)
                # QgsMessageLog.logMessage(str(j) + ' ' + str(i), 'sGIS', level=Qgis.Info)
                jobData.append(str(cell_obj.value))

            allJobs.append(jobData)
            jobData = []

        for job in allJobs:

            job_no = job[0]
            name = job[1]
            address = job[2]
            town = job[3]
            sTown = self.getTown(town)
            date = job[4]
            notes = job[5]
            job_type = job[5]
            map_bk_lot = job[6]

            QgsMessageLog.logMessage('IMPORTING | ' + str(job_no) + '...', 'sGIS', level=Qgis.Info)

            try:
                sDate = datetime.datetime.strptime(date, '%m/%d/%Y')
                finalDate = '{:%Y-%m-%d}'.format(sDate)
                date = finalDate

            except Exception:
                pass

            jLayer.setSubsetString("job_no = '{}' ".format(job_no))
            jIt = jLayer.getFeatures()

            for j in jIt:

                if job_no == j['job_no']:
                    QgsMessageLog.logMessage('WARNING | ' + str(job_no) + ' ALREADY EXISTS in DB!', 'sGIS',
                                             level=Qgis.Info)
                    jLayer.setSubsetString("id > 1")

                    reply = QMessageBox.question(self.iface.mainWindow(), str(job_no) + ' ALREADY EXISTS!',
                                                 'A duplicate will be created in GIS.  Continue?',
                                                 QMessageBox.Ok, QMessageBox.Cancel)
                    if reply == QMessageBox.Ok:
                        jLayer.setSubsetString("id > 1")

                    else:
                        jLayer.setSubsetString("id > 1")
                        self.iface.setActiveLayer(jLayer)
                        QgsMessageLog.logMessage('CANCELLED | ' + str(job_no) + ' import process aborted.', 'sGIS',
                                                 level=Qgis.Info)
                        return

                else:
                    jLayer.setSubsetString("id > 1")
                    self.iface.setActiveLayer(jLayer)
                    return

            maps = map_bk_lot.split(",")

            cLayer.setSubsetString("town = '{}' ".format(sTown))
            it = cLayer.getFeatures()
            ids = [i.id() for i in it]

            for x in range(0, len(maps)):

                expr2 = QgsExpression("map_bk_lot = '{}' ".format(maps[x]))

                try:
                    it2 = cLayer.getFeatures(QgsFeatureRequest(expr2))
                    newIds = [i2.id() for i2 in it2]
                    idsToSel = list(set(ids).intersection(newIds))
                    cLayer.selectByIds(idsToSel, 1)
                    features = cLayer.selectedFeatures()

                    provider = cLayer.dataProvider()

                    spIndex = QgsSpatialIndex()  # create spatial index object

                    feat = QgsFeature()
                    fit = provider.getFeatures()  # gets all features in layer

                    # insert features to index
                    while fit.nextFeature(feat):
                        spIndex.insertFeature(feat)

                    geometry = features[0].geometry()
                    centroid = geometry.centroid().asPoint()

                    # QgsSpatialIndex.nearestNeighbor (QgsPoint point, int neighbors)
                    nearestIds = spIndex.nearestNeighbor(centroid, 1)  # we need only one neighbour
                    QgsMessageLog.logMessage('nearestIds | ' + str(nearestIds), 'sGIS',
                                             level=Qgis.Info)

                    geom = None
                    for feat in features:

                        for f in nearestIds:
                            if feat['objectid'] == f:
                                pass
                            else:
                                zipcode = feat['zipcode']
                                county = feat['county']
                                l_l = feat['lat_lon']

                        if geom == None:
                            geom = feat.geometry()

                        else:
                            geom = geom.combine(feat.geometry())

                except Exception as e:
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                         "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                             exc_tb.tb_lineno) + ' ' + str(e))
                    QGuiApplication.restoreOverrideCursor()
                    return

            cLayer.selectByIds([])

            self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
            self.iface.setActiveLayer(self.vl)
            self.iface.actionToggleEditing().trigger()
            feature = QgsFeature()

            try:

                sourceCrs = QgsCoordinateReferenceSystem(102683)
                destCrs = QgsCoordinateReferenceSystem(26919)
                tr = QgsCoordinateTransform(sourceCrs, destCrs, QgsProject.instance())
                geom.transform(tr)
                feature.setGeometry(geom)

                dataProvider = self.vl.dataProvider()
                layer = QgsProject.instance().mapLayersByName('jobs')[0]
                dataProvider.addFeature(feature)
                self.iface.activeLayer().commitChanges()
                self.vl.dataProvider().forceReload()
                self.iface.mapCanvas().refresh()

                self.selectLastFeature()
                newF = self.vl.selectedFeatures()[0]

                lat_lon = formatLL(l_l)

                self.iface.actionToggleEditing().trigger()

                sid = newF['sid']

                idx = dataProvider.fieldNameIndex('job_no')
                self.vl.changeAttributeValue(newF.id(), idx, job_no)
                idx = dataProvider.fieldNameIndex('locus_addr')
                self.vl.changeAttributeValue(newF.id(), idx, address)
                idx = dataProvider.fieldNameIndex('town')
                self.vl.changeAttributeValue(newF.id(), idx, sTown)
                idx = dataProvider.fieldNameIndex('county')
                self.vl.changeAttributeValue(newF.id(), idx, county)
                idx = dataProvider.fieldNameIndex('zipcode')
                self.vl.changeAttributeValue(newF.id(), idx, zipcode)
                idx = dataProvider.fieldNameIndex('date_requested')
                self.vl.changeAttributeValue(newF.id(), idx, date)
                idx = dataProvider.fieldNameIndex('map_bk_lot')
                self.vl.changeAttributeValue(newF.id(), idx, map_bk_lot)
                idx = dataProvider.fieldNameIndex('job_desc')
                self.vl.changeAttributeValue(newF.id(), idx, notes)
                idx = dataProvider.fieldNameIndex('job_type')
                self.vl.changeAttributeValue(newF.id(), idx, job_type)
                idx = dataProvider.fieldNameIndex('objectid')
                self.vl.changeAttributeValue(newF.id(), idx, idsToSel[0])
                idx = dataProvider.fieldNameIndex('plan_no')
                self.vl.changeAttributeValue(newF.id(), idx, job_no)
                idx = dataProvider.fieldNameIndex('old_plan')
                self.vl.changeAttributeValue(newF.id(), idx, job_no + ' (' + job_type + ')')
                idx = dataProvider.fieldNameIndex('lat_lon')
                self.vl.changeAttributeValue(newF.id(), idx, lat_lon)

                self.vl.updateFields()
                self.iface.activeLayer().commitChanges()

                self.abutters_dialog = sgis_abutters(self.iface)
                self.abutters_dialog.run()

                feature = QgsFeature()
                self.vl = QgsProject.instance().mapLayersByName('contacts')[0]
                self.iface.setActiveLayer(self.vl)
                dataProvider = self.vl.dataProvider()

                self.iface.actionToggleEditing().trigger()
                dataProvider.addFeature(feature)
                self.iface.activeLayer().commitChanges()
                self.iface.mapCanvas().refresh()

                self.selectLastFeature()
                newF = self.vl.selectedFeatures()[0]

                self.iface.actionToggleEditing().trigger()
                idx = dataProvider.fieldNameIndex('jobs_id')
                self.vl.changeAttributeValue(newF.id(), idx, sid)
                idx = dataProvider.fieldNameIndex('contact_name')
                self.vl.changeAttributeValue(newF.id(), idx, name)
                idx = dataProvider.fieldNameIndex('contact_type')
                self.vl.changeAttributeValue(newF.id(), idx, 'IMPORTED')
                idx = dataProvider.fieldNameIndex('contact_addr')
                self.vl.changeAttributeValue(newF.id(), idx, address)
                idx = dataProvider.fieldNameIndex('primary_contact')
                self.vl.changeAttributeValue(newF.id(), idx, name)
                idx = dataProvider.fieldNameIndex('client')
                self.vl.changeAttributeValue(newF.id(), idx, 't')
                idx = dataProvider.fieldNameIndex('folder')
                self.vl.changeAttributeValue(newF.id(), idx, 't')

                self.vl.updateFields()

                self.iface.activeLayer().commitChanges()
                self.vl.dataProvider().forceReload()
                self.iface.mapCanvas().refresh()
                geom = None

            except Exception as e:
                error_count = error_count + 1
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]

                QgsMessageLog.logMessage('ERROR: | ' + str(job_no) + ' with ' + str(maps[x]) + ' | ' + 'Details: '
                                         + str(exc_type) + ' ' + str(fname) + ' ' + str(exc_tb.tb_lineno)
                                         + ' ' + str(e), 'sGIS', level=Qgis.Info)

                self.iface.actionToggleEditing().trigger()
                QGuiApplication.restoreOverrideCursor()

        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(self.vl)
        cLayer.setSubsetString('gid > 1')
        self.vl.selectByIds([])

        QgsMessageLog.logMessage('DONE.', 'sGIS', level=Qgis.Info)

    def getTown(self, town):

        if town == 'Arrowsic':
            pass
        elif town == 'AUGUSTA':
            town = 'Augusta'
        elif town == 'EDG':
            town = 'Edgecomb'
        elif town == 'BBY':
            town = 'Boothbay'
        elif town == 'BBH':
            town = 'Boothbay Harbor'
        elif town == 'WBATH':
            town = 'West Bath'
        elif town == 'SPT':
            town = 'Southport'
        elif town == 'BRISTOL':
            town = 'Bristol'
        elif town == 'S.BRIS':
            town = 'South Bristol'
        elif town == 'BATH':
            town = 'Bath'
        elif town == 'PHIPPS':
            town = 'Phippsburg'
        elif town == 'HARPS':
            town = 'Harpswell'
        elif town == 'BRUNS':
            town = 'Brunswick'
        elif town == 'WISC':
            town = 'Wiscasset'
        elif town == 'TOPS':
            town = 'Topsham'
        elif town == 'WOOL':
            town = 'Woolwich'
        elif town == 'WALDO':
            town = 'Waldoboro'
        elif town == 'Rnd Pond':
            town = 'Round Pond'
        elif town == 'JEFF':
            town = 'Jefferson'
        elif town == 'GEOTN':
            town = 'Georgetown'
        elif town == 'Nobleboro':
            pass
        elif town == 'Bowdoin':
            pass
        elif town == 'DRES':
            town = 'Dresden'
        elif town == 'DAM':
            town = 'Damariscotta'
        elif town == 'NEWC':
            town = 'Newcastle'
        elif town == 'WESTP':
            town = 'Westport Island'
        elif town == 'BREM':
            town = 'Bremen'
        elif town == 'Gardiner':
            pass
        elif town == 'MONHE':
            town = 'Monhegan'
        elif town == 'POPHAM':
            town = 'Popham'
        elif town == 'BRIST':
            town = 'Bristol'
        else:
            pass
        return town

    def selectLastFeature(self):

        f2 = self.vl.getFeatures()
        fCount = self.vl.featureCount()

        fIds = []
        fIds = [f.id() for f in f2]
        fIds.sort()

        fId = [fIds[-1]]
        self.vl.selectByIds(fId)


class sgis_moveJob(object):

    def __init__(self, iface):

        self.iface = iface

    def initGui(self):

        self.action = QAction("moveJob", self.iface.mainWindow())
        self.action.triggered.connect(self.run)
        self.action.trigger()
        self.iface.mapCanvas().selectionChanged.connect(self.select_changed)

    def run(self):

        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(self.vl)
        vLayer = self.iface.activeLayer()
        feats_count = vLayer.selectedFeatureCount()

        if feats_count == 0:
            msg = QMessageBox()
            msg.setWindowTitle('SELECT Job')
            msg.setText('Click OK and select the SOURCE job you wish to move.')
            cont = msg.addButton('Continue', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(cont)
            msg.exec_()
            msg.deleteLater()
            if msg.clickedButton() is cont:
                self.iface.actionSelect().trigger()
            else:
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception as e:
                    pass
                return
        else:
            job = self.iface.activeLayer().selectedFeatures()[0]
            jobNo = job["job_no"]
            msg = QMessageBox()
            msg.setWindowTitle('SELECT Job')
            msg.setText(str(jobNo) + ' is selected. Do you want to move that selection or choose a new selection?')
            cont = msg.addButton('Continue', QMessageBox.AcceptRole)
            new = msg.addButton('New Selection', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(cont)
            msg.exec_()
            msg.deleteLater()

            if msg.clickedButton() is cont:
                self.iface.actionSelect().trigger()
                self.select_changed()
            elif msg.clickedButton() is new:
                msg = QMessageBox()
                msg.setWindowTitle('SELECT Job')
                msg.setText('Click OK and select the SOURCE job you wish to move.')
                cont = msg.addButton('Continue', QMessageBox.AcceptRole)
                cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
                msg.setDefaultButton(cont)
                msg.exec_()
                msg.deleteLater()
                if msg.clickedButton() is cont:
                    for a in self.iface.mainWindow().children():
                        if a.objectName() == 'mActionDeselectAll':
                            a.trigger()
                            QgsMessageLog.logMessage('FIRST RUN: Previous selection has been cleared.',
                                                     'sGIS', level=Qgis.Info)
                            QgsMessageLog.logMessage('Job move starting...', 'sGIS', level=Qgis.Info)
                            self.iface.actionSelect().trigger()
                elif msg.clickedButton() is cancel:
                    try:
                        self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                    except Exception as e:
                        pass
            elif msg.clickedButton() is cancel:
                QgsMessageLog.logMessage('DEBUG: Job move cancelled.', 'sGIS', level=Qgis.Info)
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception as e:
                    pass
        resetLegend(self)

    def select_changed(self):

        try:
            self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
        except Exception as e:
            pass

        try:
            job = self.iface.activeLayer().selectedFeatures()[0]
            jobNo = job["job_no"]

            msg = QMessageBox()
            msg.setWindowTitle('SOURCE Job')
            msg.setText('JobNo: ' + jobNo + ' has been selected. Continue?')
            edit = msg.addButton('Select Job', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(edit)
            QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
            msg.exec_()
            msg.deleteLater()
            QGuiApplication.restoreOverrideCursor()

            if msg.clickedButton() is edit:
                QgsMessageLog.logMessage('Job MOVE will begin for: ' + jobNo, 'sGIS', level=Qgis.Info)
                self.active()

            else:
                self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                QgsMessageLog.logMessage('DEBUG: Job MOVE cancelled.', 'sGIS', level=Qgis.Info)
                return

        except Exception as e:
            QgsMessageLog.logMessage('EXCEPTION: ' + str(e), 'sGIS', level=Qgis.Info)

    def active(self):

        try:
            self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
        except Exception as e:
            pass

        try:
            self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
            self.iface.setActiveLayer(self.vl)
            self.oldF = self.vl.selectedFeatures()[0]
            self.job_no = self.oldF['job_no']
            self.featureType = self.oldF['objectType']
            self.attrs = self.oldF.attributes()

            QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
            QGuiApplication.restoreOverrideCursor()
            QgsMessageLog.logMessage('SOURCE: ' + self.job_no + ' has been selected.', 'sGIS', level=Qgis.Info)

            self.vl = QgsProject.instance().mapLayersByName('Parcels')[0]
            self.iface.setActiveLayer(self.vl)

            reply = QMessageBox.question(self.iface.mainWindow(), 'DESTINATION:',
                                         'Click OK and select the destination for the moved job.',
                                         QMessageBox.Ok, QMessageBox.Cancel)

            if reply == QMessageBox.Ok:

                self.iface.mapCanvas().selectionChanged.connect(self.select_dest)
                self.iface.actionSelectFreehand().trigger()

            else:
                QgsMessageLog.logMessage('DEBUG: Job MOVE cancelled.', 'sGIS', level=Qgis.Info)
                self.iface.mapCanvas().selectionChanged.disconnect(self.select_dest)

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))
            self.iface.mapCanvas().selectionChanged.disconnect(self.select_dest)
            return

    def select_dest(self): # NEED TO GRAB ALL VARIABLES FROM ORIGINAL JOB

        try:
            self.iface.mapCanvas().selectionChanged.disconnect(self.select_dest)
        except Exception as e:
            pass

        vLayer = self.iface.activeLayer()
        feats_count = vLayer.selectedFeatureCount()

        if feats_count != 0:
            msg = QMessageBox()
            msg.setWindowTitle('DESTINATION')
            msg.setText(str(feats_count) + ' feature(s) have been selected. Continue?')
            edit = msg.addButton('Continue', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(edit)
            QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
            msg.exec_()
            msg.deleteLater()
            QGuiApplication.restoreOverrideCursor()

            if msg.clickedButton() is edit:
                pass

            else:
                QgsMessageLog.logMessage('DEBUG: Job MOVE cancelled.', 'sGIS', level=Qgis.Info)
                return
        else:
            return

        try:
            QgsMessageLog.logMessage('SELECT DEST ENTERED.', 'sGIS', level=Qgis.Info)
            passwd = getSQL(self)
            try:
                connection = psycopg2.connect(user=dbUser,
                                              password=passwd,
                                              host=dbServer,
                                              port=dbPort,
                                              database=db)

                cursor = connection.cursor()

                # print("id BEFORE updating record: ")
                sql_select_query = """SELECT last_value from brs_jobs_id_seq"""
                cursor.execute(sql_select_query)
                record = cursor.fetchone()
                for r in record:
                    if str(r) == '':
                        pass
                    else:
                        currval = r
                        QgsMessageLog.logMessage('currval BEFORE: ' + str(currval), 'sGIS', level=Qgis.Info)
                        # print(str(currval))

            except (Exception, psycopg2.Error) as error:
                QgsMessageLog.logMessage('PostgreSQL ERROR: ' + str(error), 'sGIS', level=Qgis.Info)

            finally:
                # closing database connection.
                if (connection):
                    cursor.close()
                    connection.close()
                    # print("PostgreSQL connection is closed")

            self.iface.actionCopyFeatures().trigger()
            self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
            self.iface.setActiveLayer(self.vl)
            self.iface.actionToggleEditing().trigger()
            self.iface.actionPasteFeatures().trigger()
            self.iface.mainWindow().findChild(QAction, 'mActionMergeFeatures').trigger()
            self.iface.activeLayer().commitChanges()
            self.iface.messageBar().clearWidgets()

            ## DONE WITH CREATION, SELECT NEW AND PROCEED WITH SWAP

            self.selectLastFeature()
            self.newF = self.vl.selectedFeatures()[0]

            job_no2 = self.newF['job_no']
            lat_lon = self.newF['lat_lon']
            job_id2 = self.newF['sid']

            try:
              ll = len(lat_lon)

              if ll <= 30:
                  pass
              else:
                  lat_lon = formatLL(lat_lon)
                  self.iface.actionToggleEditing().trigger()
                  layerData = self.vl.dataProvider()
                  idx = layerData.fieldNameIndex('lat_lon')
                  self.vl.changeAttributeValue(self.newF.id(), idx, lat_lon)
                  self.iface.activeLayer().commitChanges()
            except Exception as e:
                pass
                
            i = 0
            for a in self.oldF.fields():
                if str(self.oldF[a.name()]) == 'NULL':
                    i = i + 1
                    pass
                else:
                    if i in (0,2,13,14,79,82,83,84,85,86):
                        i = i + 1
                        pass
                    else:
                        self.vl.startEditing()
                        self.vl.changeAttributeValue(self.newF.id(), i, str(self.oldF[a.name()]))
                        self.vl.commitChanges()
                        i = i + 1
            try:

                self.job_sid = self.oldF['sid']
                job_no = self.oldF['job_no']
                self.job_id = self.oldF.id()
                self.vl = QgsProject.instance().mapLayersByName('contacts')[0]
                self.iface.setActiveLayer(self.vl)
                dataProvider = self.vl.dataProvider()
                self.vl.setSubsetString("jobs_id = '{}' ".format(self.job_sid))

                self.iface.activeLayer().selectAll()

                for x in self.vl.selectedFeatures():

                    c = self.vl.selectedFeatures()[0]
                    self.iface.actionToggleEditing().trigger()
                    idx = dataProvider.fieldNameIndex('jobs_id')
                    self.vl.changeAttributeValue(c.id(), idx, job_id2)
                    self.vl.updateFields()
                    self.iface.activeLayer().commitChanges()

                self.vl.setSubsetString('cid > 1')

            except Exception as e:
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                QMessageBox.critical(self.iface.mainWindow(), "CONTACTS EXCEPTION",
                                     "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                         exc_tb.tb_lineno) + ' ' + str(e))
                return

            self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
            self.iface.setActiveLayer(self.vl)
            passwd = getSQL(self)
            try:
                connection = psycopg2.connect(user=dbUser,
                                              password=passwd,
                                              host=dbServer,
                                              port=dbPort,
                                              database=db)

                cursor = connection.cursor()

                # Update single record now
                sql_update_query = """SELECT setval('brs_jobs_id_seq', """ + str(currval) + """, true); """
                cursor.execute(sql_update_query)
                connection.commit()
                sql_update_query = """DELETE from abutters where referrerj = '""" + str(job_no) + """';"""
                cursor.execute(sql_update_query)
                connection.commit()

                # count = cursor.rowcount

                sql_select_query = """SELECT last_value from brs_jobs_id_seq"""
                cursor.execute(sql_select_query)
                record = cursor.fetchone()
                for r in record:
                    if str(r) == '':
                        pass
                    else:
                        currval = r
                        # QgsMessageLog.logMessage('currval AFTER: ' + str(currval), 'sGIS', level=Qgis.Info)

            except (Exception, psycopg2.Error) as error:
                QgsMessageLog.logMessage('PostgreSQL ERROR: ' + str(error), 'sGIS', level=Qgis.Info)

            finally:
                # closing database connection.
                if (connection):
                    cursor.close()
                    connection.close()

            self.abutters_dialog = sgis_abutters(self.iface)
            self.abutters_dialog.run()

            self.vl.startEditing()
            self.vl.deleteFeature(self.job_id)
            self.vl.commitChanges()

            status = "The feature has been moved successfully."
            QMessageBox.information(None, "SUCCESS", str(status))

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))
            return

    def selectLastFeature(self):

        f2 = self.vl.getFeatures()
        fCount = self.vl.featureCount()

        fIds = []
        fIds = [f.id() for f in f2]
        fIds.sort()

        fId = [fIds[-1]]
        self.vl.selectByIds(fId)


class sgis_movePlan(object):

    def __init__(self, iface):

        self.iface = iface

    def initGui(self):

        self.action = QAction("move@Plan", self.iface.mainWindow())
        self.action.triggered.connect(self.run)
        self.action.trigger()
        self.iface.mapCanvas().selectionChanged.connect(self.select_changed)

    def run(self):

        self.vl = QgsProject.instance().mapLayersByName('plans')[0]
        self.iface.setActiveLayer(self.vl)
        vLayer = self.iface.activeLayer()
        feats_count = vLayer.selectedFeatureCount()

        if feats_count == 0:
            msg = QMessageBox()
            msg.setWindowTitle('SELECT Plan')
            msg.setText('Click OK and select the SOURCE plan you wish to move.')
            cont = msg.addButton('Continue', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(cont)
            msg.exec_()
            msg.deleteLater()
            if msg.clickedButton() is cont:
                self.iface.actionSelect().trigger()
            else:
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception as e:
                    pass
                return
        else:
            plan = self.iface.activeLayer().selectedFeatures()[0]
            planNo = plan["plan_no"]
            msg = QMessageBox()
            msg.setWindowTitle('SELECT Plan')
            msg.setText(str(planNo) + ' is selected. Do you want to move that selection or choose a new selection?')
            cont = msg.addButton('Continue', QMessageBox.AcceptRole)
            new = msg.addButton('New Selection', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(cont)
            msg.exec_()
            msg.deleteLater()

            if msg.clickedButton() is cont:
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception as e:
                    pass
                self.select_changed()
            elif msg.clickedButton() is new:
                msg = QMessageBox()
                msg.setWindowTitle('SELECT Plan')
                msg.setText('Click OK and select the SOURCE plan you wish to move.')
                cont = msg.addButton('Continue', QMessageBox.AcceptRole)
                cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
                msg.setDefaultButton(cont)
                msg.exec_()
                msg.deleteLater()
                if msg.clickedButton() is cont:
                    for a in self.iface.mainWindow().children():
                        if a.objectName() == 'mActionDeselectAll':
                            a.trigger()
                            QgsMessageLog.logMessage('FIRST RUN: Previous selection has been cleared.',
                                                     'sGIS', level=Qgis.Info)
                            QgsMessageLog.logMessage('Plan move starting...', 'sGIS', level=Qgis.Info)
                            self.iface.actionSelect().trigger()
                elif msg.clickedButton() is cancel:
                    try:
                        self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                    except Exception as e:
                        pass
            elif msg.clickedButton() is cancel:
                QgsMessageLog.logMessage('DEBUG: Plan move cancelled.', 'sGIS', level=Qgis.Info)
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception as e:
                    pass
        resetLegend(self)

    def select_changed(self):

        try:
            self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
        except Exception as e:
            pass

        try:
            parcel = self.iface.activeLayer().selectedFeatures()[0]
            planNo = parcel["plan_no"]

            msg = QMessageBox()
            msg.setWindowTitle('SOURCE Plan')
            msg.setText('PlanNo: ' + planNo + ' has been selected. Continue?')
            edit = msg.addButton('Continue', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(edit)
            QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
            msg.exec_()
            msg.deleteLater()
            QGuiApplication.restoreOverrideCursor()

            if msg.clickedButton() is edit:
                QgsMessageLog.logMessage('Plan MOVE will begin for: ' + planNo, 'sGIS', level=Qgis.Info)
                self.active()

            else:
                self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                QgsMessageLog.logMessage('DEBUG: Plan MOVE cancelled.', 'sGIS', level=Qgis.Info)
                return

        except Exception as e:
            QgsMessageLog.logMessage('EXCEPTION: ' + str(e), 'sGIS', level=Qgis.Info)

    def active(self):

        try:
            self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
        except Exception as e:
            pass

        try:
            self.vl = QgsProject.instance().mapLayersByName('plans')[0]
            self.iface.setActiveLayer(self.vl)
            self.oldF = self.vl.selectedFeatures()[0]
            self.job_no = self.oldF['plan_no']
            self.attrs = self.oldF.attributes()

            QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
            QGuiApplication.restoreOverrideCursor()
            QgsMessageLog.logMessage('SOURCE: ' + self.job_no + ' has been selected.', 'sGIS', level=Qgis.Info)

            self.vl = QgsProject.instance().mapLayersByName('Parcels')[0]
            self.iface.setActiveLayer(self.vl)

            reply = QMessageBox.question(self.iface.mainWindow(), 'DESTINATION:',
                                         'Click OK and select the destination for the moved plan.',
                                         QMessageBox.Ok, QMessageBox.Cancel)

            if reply == QMessageBox.Ok:

                self.iface.mapCanvas().selectionChanged.connect(self.select_dest)
                self.iface.actionSelectFreehand().trigger()

            else:
                QgsMessageLog.logMessage('DEBUG: Plan MOVE cancelled.', 'sGIS', level=Qgis.Info)
                self.iface.mapCanvas().selectionChanged.disconnect(self.select_dest)

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))
            self.iface.mapCanvas().selectionChanged.disconnect(self.select_dest)
            return

    def select_dest(self): # NEED TO GRAB ALL VARIABLES FROM ORIGINAL JOB

        try:
            self.iface.mapCanvas().selectionChanged.disconnect(self.select_dest)
        except Exception as e:
            pass

        vLayer = self.iface.activeLayer()
        feats_count = vLayer.selectedFeatureCount()

        if feats_count != 0:
            msg = QMessageBox()
            msg.setWindowTitle('DESTINATION')
            msg.setText(str(feats_count) + ' feature(s) have been selected. Continue?')
            edit = msg.addButton('Continue', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(edit)
            QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
            msg.exec_()
            msg.deleteLater()
            QGuiApplication.restoreOverrideCursor()

            if msg.clickedButton() is edit:
                pass

            else:
                QgsMessageLog.logMessage('DEBUG: Plan MOVE cancelled.', 'sGIS', level=Qgis.Info)
                return
        else:
            return

        try:
            # QgsMessageLog.logMessage('SELECT DEST ENTERED.', 'sGIS', level=Qgis.Info)
            self.iface.actionCopyFeatures().trigger()
            self.vl = QgsProject.instance().mapLayersByName('plans')[0]
            self.iface.setActiveLayer(self.vl)
            self.iface.actionToggleEditing().trigger()
            self.iface.actionPasteFeatures().trigger()

            self.iface.mainWindow().findChild(QAction, 'mActionMergeFeatures').trigger()
            self.iface.activeLayer().commitChanges()
            self.iface.messageBar().clearWidgets()

            ## DONE WITH CREATION, SELECT NEW AND PROCEED WITH SWAP

            feats_count = self.vl.selectedFeatureCount()

            if feats_count == 0:
                self.selectLastFeature()
                self.newF = self.vl.selectedFeatures()[0]
            else:
                self.newF = self.vl.selectedFeatures()[0]

            lat_lon = self.newF['lat_lon']

            # QgsMessageLog.logMessage('lat_lon: ' + str(lat_lon), 'sGIS', level=Qgis.Info)
            
            if str(lat_lon) == 'NULL':
                ll = 1
                pass
            else:
                ll = len(lat_lon)

            if ll <= 30:
                pass
            else:
                lat_lon = formatLL(lat_lon)
                self.iface.actionToggleEditing().trigger()
                layerData = self.vl.dataProvider()
                idx = layerData.fieldNameIndex('lat_lon')
                self.vl.changeAttributeValue(self.newF.id(), idx, lat_lon)
                # self.vl.updateFields()
                self.iface.activeLayer().commitChanges()

            i = 0
            for a in self.oldF.fields():
                if str(self.oldF[a.name()]) == 'NULL':
                    i = i + 1
                    pass
                else:
                    # QgsMessageLog.logMessage('i: ' + str(i) + ' | '+ str(self.oldF[a.name()]), 'sGIS', level=Qgis.Info)
                    if i in (0,1,2,4,11,13):
                        # QgsMessageLog.logMessage('pass: ' + str(i), 'sGIS', level=Qgis.Info)
                        i = i + 1
                        pass
                    elif i == 8:
                        self.vl.startEditing()
                        self.vl.changeAttributeValue(self.newF.id(), i, self.oldF[a.name()])
                        self.vl.commitChanges()
                        # QgsMessageLog.logMessage('SET: ' + str(i) + ' ' + str(self.oldF[a.name()]), 'sGIS', level=Qgis.Info)
                        i = i + 1
                    elif i == 15:
                        self.vl.startEditing()
                        self.vl.changeAttributeValue(self.newF.id(), i, self.oldF[a.name()])
                        self.vl.commitChanges()
                        i = i + 1
                    else:
                        # self.updateAttribute(self.newF, i, str(self.oldF[a.name()]))
                        self.vl.startEditing()
                        self.vl.changeAttributeValue(self.newF.id(), i, str(self.oldF[a.name()]))
                        self.vl.commitChanges()
                        i = i + 1

            self.vl = QgsProject.instance().mapLayersByName('plans')[0]
            self.iface.setActiveLayer(self.vl)
            self.iface.actionIdentify().trigger()

            self.oldF_id = self.oldF.id()
            self.vl.startEditing()
            self.vl.deleteFeature(self.oldF_id)
            self.vl.commitChanges()

            status = "The feature has been moved successfully."
            QMessageBox.information(None, "SUCCESS", str(status))

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))
            return


class sgis_moveSupp(object):

    def __init__(self, iface):

        self.iface = iface

    def initGui(self):

        self.action = QAction("move@Supp", self.iface.mainWindow())
        self.action.triggered.connect(self.run)
        self.action.trigger()
        self.iface.mapCanvas().selectionChanged.connect(self.select_changed)

    def run(self):

        self.vl = QgsProject.instance().mapLayersByName('supplementals')[0]
        self.iface.setActiveLayer(self.vl)
        vLayer = self.iface.activeLayer()
        feats_count = vLayer.selectedFeatureCount()

        if feats_count == 0:
            msg = QMessageBox()
            msg.setWindowTitle('SELECT Supplemental')
            msg.setText('Click OK and select the SOURCE supplemental you wish to move.')
            cont = msg.addButton('OK', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(cont)
            msg.exec_()
            msg.deleteLater()
            if msg.clickedButton() is cont:
                self.iface.actionSelect().trigger()
            else:
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception as e:
                    pass
                return
        else:
            supp = self.iface.activeLayer().selectedFeatures()[0]
            suppNo = supp["supp_no"]
            msg = QMessageBox()
            msg.setWindowTitle('SELECT Supplemental')
            msg.setText(str(suppNo) + ' is selected. Do you want to move that selection or choose a new selection?')
            cont = msg.addButton('Continue', QMessageBox.AcceptRole)
            new = msg.addButton('New Selection', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(cont)
            msg.exec_()
            msg.deleteLater()

            if msg.clickedButton() is cont:
                self.select_changed()
            elif msg.clickedButton() is new:
                msg = QMessageBox()
                msg.setWindowTitle('SELECT Supplemental')
                msg.setText('Click OK and select the SOURCE supplemental you wish to move.')
                cont = msg.addButton('Continue', QMessageBox.AcceptRole)
                cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
                msg.setDefaultButton(cont)
                msg.exec_()
                msg.deleteLater()
                if msg.clickedButton() is cont:
                    for a in self.iface.mainWindow().children():
                        if a.objectName() == 'mActionDeselectAll':
                            a.trigger()
                            QgsMessageLog.logMessage('Supp move starting...', 'sGIS', level=Qgis.Info)
                            self.iface.actionSelect().trigger()
                elif msg.clickedButton() is cancel:
                    try:
                        self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                    except Exception as e:
                        pass
            elif msg.clickedButton() is cancel:
                QgsMessageLog.logMessage('DEBUG: Supp move cancelled.', 'sGIS', level=Qgis.Info)
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception as e:
                    pass
        resetLegend(self)

    def select_changed(self):

        try:
            self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
        except Exception as e:
            pass

        try:
            parcel = self.iface.activeLayer().selectedFeatures()[0]
            jobNo = parcel["job_no"]

            msg = QMessageBox()
            msg.setWindowTitle('SOURCE Supplemental')
            msg.setText('SuppNo: ' + jobNo + ' has been selected. Continue?')
            edit = msg.addButton('Select Supplemental', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(edit)
            QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
            msg.exec_()
            msg.deleteLater()
            QGuiApplication.restoreOverrideCursor()

            if msg.clickedButton() is edit:
                QgsMessageLog.logMessage('Supp MOVE will begin for: ' + jobNo, 'sGIS', level=Qgis.Info)
                self.active()

            else:
                self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                QgsMessageLog.logMessage('DEBUG: Supp MOVE cancelled.', 'sGIS', level=Qgis.Info)
                return

        except Exception as e:
            QgsMessageLog.logMessage('EXCEPTION: ' + str(e), 'sGIS', level=Qgis.Info)

    def active(self):

        try:
            self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
        except Exception as e:
            pass

        try:
            self.vl = QgsProject.instance().mapLayersByName('supplementals')[0]
            self.iface.setActiveLayer(self.vl)
            self.oldF = self.vl.selectedFeatures()[0]
            self.job_no = self.oldF['job_no']
            self.oldSuppType = self.oldF['supp_type']
            self.oldID = self.oldF['id']
            self.attrs = self.oldF.attributes()

            QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
            QGuiApplication.restoreOverrideCursor()

            self.vl = QgsProject.instance().mapLayersByName('Parcels')[0]
            self.iface.setActiveLayer(self.vl)

            reply = QMessageBox.question(self.iface.mainWindow(), 'DESTINATION:',
                                         'Click OK and select the destination for the moved supplemental.',
                                         QMessageBox.Ok, QMessageBox.Cancel)

            if reply == QMessageBox.Ok:

                self.iface.mapCanvas().selectionChanged.connect(self.select_dest)
                self.iface.actionSelectFreehand().trigger()

            else:
                QgsMessageLog.logMessage('DEBUG: Supp MOVE cancelled.', 'sGIS', level=Qgis.Info)

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))
            self.iface.mapCanvas().selectionChanged.disconnect(self.select_dest)
            return

    def select_dest(self): # NEED TO GRAB ALL VARIABLES FROM ORIGINAL JOB

        try:
            self.iface.mapCanvas().selectionChanged.disconnect(self.select_dest)
        except Exception as e:
            pass

        vLayer = self.iface.activeLayer()
        feats_count = vLayer.selectedFeatureCount()

        if feats_count != 0:
            msg = QMessageBox()
            msg.setWindowTitle('DESTINATION')
            msg.setText(str(feats_count) + ' feature(s) have been selected. Continue?')
            edit = msg.addButton('Continue', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(edit)
            QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
            msg.exec_()
            msg.deleteLater()
            QGuiApplication.restoreOverrideCursor()

            if msg.clickedButton() is edit:
                pass

            else:
                QgsMessageLog.logMessage('DEBUG: Supp MOVE cancelled.', 'sGIS', level=Qgis.Info)
                return
        else:
            return

        try:
            # QgsMessageLog.logMessage('SELECT DEST ENTERED.', 'sGIS', level=Qgis.Info)

            self.iface.setActiveLayer(self.vl)
            self.iface.actionCopyFeatures().trigger() # copy from parcels
            self.fields = self.iface.activeLayer().fields() # grab fields from parcels

            # create tmp layer for merge
            self.tmpLayer = QgsVectorLayer('MultiPolygon?crs=ESRI:102683', 'tmp_merge', 'memory')
            QgsProject.instance().addMapLayers([self.tmpLayer])
            self.iface.actionIdentify().trigger()

            # add fields to tmp layer
            self.iface.setActiveLayer(self.tmpLayer)
            self.iface.actionToggleEditing()
            self.layerData = self.tmpLayer.dataProvider()
            self.layerData.addAttributes(self.fields)
            self.iface.activeLayer().commitChanges()

            self.iface.actionToggleEditing().trigger()
            self.iface.actionPasteFeatures().trigger()
            self.iface.mainWindow().findChild(QAction, 'mActionMergeFeatures').trigger()
            self.iface.activeLayer().commitChanges()
            self.iface.messageBar().clearWidgets()

            ## PASTED TO TMP_LAYER

            self.iface.setActiveLayer(self.tmpLayer)
            self.layerData = self.tmpLayer.dataProvider()
            self.iface.actionToggleEditing().trigger()
            self.layerData.addAttributes([QgsField("supp_type", QVariant.String)])
            self.layerData.addAttributes([QgsField("id", QVariant.String)])
            self.iface.activeLayer().commitChanges()

            feats_count = self.tmpLayer.selectedFeatureCount()

            if feats_count == 0:
                self.iface.activeLayer().selectAll()
                self.new_tmp = self.tmpLayer.selectedFeatures()[0]
            elif feats_count == 1:
                self.new_tmp = self.tmpLayer.selectedFeatures()[0]

            try:
                idx = self.layerData.fieldNameIndex('supp_type')
                self.tmpLayer.startEditing()
                self.tmpLayer.changeAttributeValue(self.new_tmp.id(), idx, str(self.oldSuppType))
                idx = self.layerData.fieldNameIndex('id')
                self.tmpLayer.changeAttributeValue(self.new_tmp.id(), idx, str(self.oldID))
                self.tmpLayer.commitChanges()

                self.iface.actionCopyFeatures().trigger()

                if self.oldSuppType == 'X':
                    self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
                else:
                    self.vl = QgsProject.instance().mapLayersByName('supplementals')[0]

                self.iface.setActiveLayer(self.vl)
                self.iface.actionToggleEditing().trigger()
                self.iface.actionPasteFeatures().trigger()
                self.iface.activeLayer().commitChanges()

                self.newF = self.vl.selectedFeatures()[0]

                job_no2 = self.newF['job_no']
                lat_lon = self.newF['lat_lon']
                job_id2 = self.newF['sid']

                ll = len(lat_lon)

                if ll <= 30:
                    pass
                else:
                    lat_lon = formatLL(lat_lon)
                    self.iface.actionToggleEditing().trigger()
                    layerData = self.vl.dataProvider()
                    idx = layerData.fieldNameIndex('lat_lon')
                    self.vl.changeAttributeValue(self.newF.id(), idx, lat_lon)
                    self.iface.activeLayer().commitChanges()

                i = 0
                for a in self.oldF.fields():
                    if str(self.oldF[a.name()]) == 'NULL':
                        i = i + 1
                    else:
                        if i in (0,2,13,14,79,82,83,84,85,86):
                            i = i + 1
                            pass
                        else:
                            self.vl.startEditing()
                            self.vl.changeAttributeValue(self.newF.id(), i, str(self.oldF[a.name()]))
                            self.vl.commitChanges()
                            i = i + 1

                # may not reach here but exception ok?
                status = "The feature has been moved successfully."
                QMessageBox.information(None, "SUCCESS", str(status))

            except Exception as e:
                status = "The feature has been moved successfully."
                QMessageBox.information(None, "SUCCESS*", str(status))
                pass

            self.oldF_id = self.oldF.id()
            self.vl.startEditing()
            self.vl.deleteFeature(self.oldF_id)
            self.vl.commitChanges()

            self.vl = QgsProject.instance().mapLayersByName('supplementals')[0]
            self.iface.setActiveLayer(self.vl)
            try:
                QgsProject.instance().removeMapLayer(self.tmpLayer.id())

            except Exception as e:
                pass

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))
            return


class sgis_parcel(object):

    def __init__(self, iface):

        self.iface = iface

    def initGui(self):

        self.action = QAction("adding NEW parcel...", self.iface.mainWindow())
        self.action.triggered.connect(self.run)
        self.action.trigger()

    def run(self):

        self.vl = QgsProject.instance().mapLayersByName('Parcels')[0]
        self.iface.setActiveLayer(self.vl)
        cLayer = self.iface.mapCanvas().currentLayer()
        cLayer.featureAdded.connect(self.feature_added)
        self.iface.actionToggleEditing().trigger()
        self.iface.mainWindow().findChild(QAction, 'mActionAddFeature').trigger()

    def feature_added(self, fid):

        cLayer = self.iface.mapCanvas().currentLayer()
        cLayer.featureAdded.disconnect()
        cLayer.commitChanges()

        QMessageBox.critical(self.iface.mainWindow(), "WARNING!",
                             "This process will take several minutes to complete.  Please be patient." )

        t1 = sgis_addParcel(newParcel(self))
        t1.start()
        t1.join()


class sgis_exportDXF(object):

    def __init__(self, iface, cV, pV, aeV, veV, hV, sV, rdsV):

        self.iface = iface
        self.cV = cV
        self.pV = pV
        self.aeV = aeV
        self.veV = veV
        self.hV = hV
        self.sV = sV
        self.rdsV = rdsV

    def initGui(self, cV, pV, aeV, veV, hV, sV, rdsV):

        self.run(cV, pV, aeV, veV, hV, sV, rdsV)

    def run(self, cV, pV, aeV, veV, hV, sV, rdsV):

        self.cV = cV
        self.pV = pV
        self.aeV = aeV
        self.veV = veV
        self.hV = hV
        self.sV = sV
        self.rdsV = rdsV

        self.iface.mapCanvas().selectionChanged.connect(self.select_changed)
        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(self.vl)
        aLayer = self.iface.activeLayer()
        feats_count = aLayer.selectedFeatureCount()

        if feats_count == 0:
            msg = QMessageBox()
            msg.setWindowTitle('Selection')
            msg.setText('Choose the correct job feature and click OK to proceed.')
            cont = msg.addButton('Continue', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(cont)
            msg.exec_()
            msg.deleteLater()
            if msg.clickedButton() is cont:
                self.iface.actionSelect().trigger()
                
            else:
                QgsMessageLog.logMessage('DEBUG: exportDXF cancelled.', 'sGIS', level=Qgis.Info)
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception as e:
                    pass
                return
        else:
            feat = self.iface.activeLayer().selectedFeatures()[0]
            jobNo = str(feat["job_no"])

            msg = QMessageBox()
            msg.setWindowTitle('Selection')
            msg.setText(str(jobNo) + ' is selected. Do you want to use that selection or choose a '
                                           'new selection?')
            cont = msg.addButton('Continue', QMessageBox.AcceptRole)
            new = msg.addButton('New Selection', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(cont)
            msg.exec_()
            msg.deleteLater()

            if msg.clickedButton() is cont:
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception as e:
                    pass
                self.select_changed()
            elif msg.clickedButton() is new:
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception as e:
                    pass

                msg = QMessageBox()
                msg.setWindowTitle('Selection')
                msg.setText('Choose the correct job feature and click OK to proceed.')
                cont = msg.addButton('Continue', QMessageBox.AcceptRole)
                cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
                msg.setDefaultButton(cont)
                msg.exec_()
                msg.deleteLater()
                if msg.clickedButton() is cont:
                    for a in self.iface.mainWindow().children():

                        if a.objectName() == 'mActionDeselectAll':
                            a.trigger()
                            QgsMessageLog.logMessage('FIRST RUN: Previous selection has been cleared.',
                                                     'sGIS', level=Qgis.Info)
                            self.iface.actionSelect().trigger()
                    try:
                        self.iface.mapCanvas().selectionChanged.connect(self.select_changed)
                    except Exception as e:
                        pass

                elif msg.clickedButton() is cancel:
                    try:
                        self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                    except Exception as e:
                        pass

            elif msg.clickedButton() is cancel:
                QgsMessageLog.logMessage('DEBUG: exportDXF cancelled.', 'sGIS', level=Qgis.Info)
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception as e:
                    pass

        QgsMessageLog.logMessage('DXF Generation Complete.', 'sGIS', level=Qgis.Info)
        resetLegend(self)
        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(self.vl)

    def select_changed(self):

        try:
            self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
        except Exception as e:
            pass

        try:
            feat = self.iface.activeLayer().selectedFeatures()[0]
            jobNo = str(feat["job_no"])

            msg = QMessageBox()
            msg.setWindowTitle('SOURCE Job')
            msg.setText('JobNo: ' + jobNo + ' has been selected. Continue?')
            edit = msg.addButton('Select Job', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(edit)
            QGuiApplication.setOverrideCursor(Qt.ArrowCursor)
            msg.exec_()
            msg.deleteLater()
            QGuiApplication.restoreOverrideCursor()

            if msg.clickedButton() is edit:
                import datetime
                year = datetime.datetime.today().strftime('%Y')
                jobYear = '20' + jobNo[:2]

                if jobYear == year:
                    year = year
                else:
                    year = jobYear

                path = os.path.join(jobsPath, year, jobNo)  # need to programattically grab year
                jobFolders(jobNo)
                gispath = os.path.join(path, "GIS")

                town = str(feat["town"])

                self.iface.actionCopyFeatures().trigger()
                self.fields = self.iface.activeLayer().fields()

                self.tmpLayer = QgsVectorLayer('MultiPolygon?crs=EPSG:102684', 'tmp_buffer', 'memory')
                QgsProject.instance().addMapLayers([self.tmpLayer])
                self.iface.actionIdentify().trigger()
                self.vl = QgsProject.instance().mapLayersByName('tmp_buffer')[0]
                self.iface.setActiveLayer(self.vl)

                self.iface.actionToggleEditing()
                self.layerData = self.tmpLayer.dataProvider()
                self.layerData.addAttributes(self.fields)
                self.iface.activeLayer().commitChanges()

                self.iface.actionToggleEditing().trigger()
                self.iface.actionPasteFeatures().trigger()
                self.iface.activeLayer().commitChanges()

                self.bufferPOLY('tmp_buffer')

                contoursSHP = gispath + "\\contours.shp"
                contoursDXF = gispath + "\\contours.dxf"
                contoursGeoJSON = gispath + "\\contours.geojson"
                aZoneSHP = gispath + "\\aZone.shp"
                aZoneDXF = gispath + "\\aZone.dxf"
                veZoneSHP = gispath + "\\veZone.shp"
                veZoneDXF = gispath + "\\veZone.dxf"
                parcelsSHP = gispath + "\\parcels.shp"
                parcelsDXF = gispath + "\\parcels.dxf"
                hat2015SHP = gispath + "\\hat2015.shp"
                hat2015DXF = gispath + "\\hat2015.dxf"
                slopeTIFF = gispath + "\\midcoastSlope.tiff"
                slopeSHP = gispath + "\\midcoastSlope.shp"
                slopeSHPfinal = gispath + "\\midcoastSlopeFINAL.shp"
                slopeDXF = gispath + "\\midcoastSlope.dxf"
                rdsSHP = gispath + "\\rds.shp"
                rdsDXF = gispath + "\\rds.dxf"

                aZoneLayer = QgsProject.instance().mapLayersByName('S_Fld_Haz_Ar')[0]
                veZoneLayer = QgsProject.instance().mapLayersByName('S_Fld_Haz_Ar VE')[0]
                parcelsLayer = QgsProject.instance().mapLayersByName('Parcels')[0]
                hat2015Layer = QgsProject.instance().mapLayersByName('HAT_2015')[0]
                rdsLayer = QgsProject.instance().mapLayersByName('ng911rdss')[0]
                self.slopeLayer = \
                QgsProject.instance().mapLayersByName('MidcoastSlopeGreaterThan20PctGreaterThan2AcMask')[0]

                self.layerOutput = QgsProject.instance().mapLayersByName('output')[0]

                import subprocess

                filenames = []

                # NEED CLEAN EXIT IF UNSUPPORTED TOWN.
                if self.cV == 2:
                    self.intersectCONTOURS(town)
                    self.iLayer = QgsProject.instance().mapLayersByName('Intersection')[0]
                    self.multi2single(self.iLayer)
                    self.sLayer = QgsProject.instance().mapLayersByName('Single parts')[0]
                    self.exportGeoJSON(self.sLayer, contoursGeoJSON)
                    self.exportSHP(self.iLayer, contoursSHP)
                    command = ["ogr2ogr", "-skipfailures", "-f", "DXF", "-zField", """Elev_FT""", contoursDXF,
                               contoursSHP]
                    subprocess.check_call(command, stderr=subprocess.DEVNULL, stdout=subprocess.DEVNULL)
                    filenames.append(gispath + "\\contours.dxf")

                if self.aeV == 2:
                    self.intersectALL(aZoneLayer, town)
                    self.iLayer = QgsProject.instance().mapLayersByName('Intersection')[0]
                    self.exportSHP(self.iLayer, aZoneSHP)
                    command = ["ogr2ogr", "-skipfailures", "-f", "DXF", aZoneDXF, aZoneSHP]
                    subprocess.check_call(command, stderr=subprocess.DEVNULL, stdout=subprocess.DEVNULL)
                    filenames.append(gispath + "\\aZone.dxf")
                if self.veV == 2:
                    self.intersectALL(veZoneLayer, town)
                    self.iLayer = QgsProject.instance().mapLayersByName('Intersection')[0]
                    self.exportSHP(self.iLayer, veZoneSHP)
                    command = ["ogr2ogr", "-skipfailures", "-f", "DXF", veZoneDXF, veZoneSHP]
                    subprocess.check_call(command, stderr=subprocess.DEVNULL, stdout=subprocess.DEVNULL)
                    filenames.append(gispath + "\\veZone.dxf")
                if self.rdsV == 2:
                    self.intersectALL(rdsLayer, town)
                    self.iLayer = QgsProject.instance().mapLayersByName('Intersection')[0]
                    self.exportSHP(self.iLayer, rdsSHP)
                    command = ["ogr2ogr", "-skipfailures", "-f", "DXF", rdsDXF, rdsSHP]
                    subprocess.check_call(command, stderr=subprocess.DEVNULL, stdout=subprocess.DEVNULL)
                    filenames.append(gispath + "\\rds.dxf")
                if self.pV == 2:
                    self.intersectALL(parcelsLayer, town)
                    self.iLayer = QgsProject.instance().mapLayersByName('Intersection')[0]
                    self.exportSHP(self.iLayer, parcelsSHP)
                    command = ["ogr2ogr", "-skipfailures", "-f", "DXF", parcelsDXF, parcelsSHP]
                    subprocess.check_call(command, stderr=subprocess.DEVNULL, stdout=subprocess.DEVNULL)
                    filenames.append(gispath + "\\parcels.dxf")
                if self.hV == 2:
                    self.intersectALL(hat2015Layer, town)
                    self.iLayer = QgsProject.instance().mapLayersByName('Intersection')[0]
                    self.exportSHP(self.iLayer, hat2015SHP)
                    command = ["ogr2ogr", "-skipfailures", "-f", "DXF", hat2015DXF, hat2015SHP]
                    subprocess.check_call(command, stderr=subprocess.DEVNULL, stdout=subprocess.DEVNULL)
                    filenames.append(gispath + "\\hat2015.dxf")
                if self.sV == 2:
                    self.exportTIFF(self.slopeLayer, slopeTIFF)
                    layerSlope = QgsProject.instance().mapLayersByName('tmp_slope')[0]
                    self.polygonize(layerSlope, slopeSHP)
                    layerSlopeSHP = QgsProject.instance().mapLayersByName('tmp_slopeSHP')[0]
                    dataProvider = layerSlopeSHP.dataProvider()
                    dataProvider.addAttributes([QgsField("Layer", QVariant.String)])

                    for f in layerSlopeSHP.getFeatures():
                        self.iface.actionToggleEditing().trigger()
                        idx = dataProvider.fieldNameIndex('Layer')
                        layerSlopeSHP.changeAttributeValue(f.id(), idx, 'GOV-GIS-SLOPE')
                        layerSlopeSHP.updateFields()
                        self.iface.activeLayer().commitChanges()

                    self.tsLayer = QgsProject.instance().mapLayersByName('tmp_slopeSHP')[0]
                    self.exportSHP(self.tsLayer, slopeSHPfinal)
                    command = ["ogr2ogr", "-skipfailures", "-f", "DXF", slopeDXF, slopeSHPfinal]
                    subprocess.check_call(command, stderr=subprocess.DEVNULL, stdout=subprocess.DEVNULL)
                    filenames.append(gispath + "\\midcoastSlope.dxf")

                def merge(source, target):
                    importer = Importer(source, target)
                    importer.import_modelspace()
                    importer.finalize()

                base_dxf = ezdxf.readfile(filenames[0])

                for filename in filenames:
                    merge_dxf = ezdxf.readfile(filename)
                    QgsMessageLog.logMessage('filename: ' + '"' + str(filename) + '"', 'sGIS', level=Qgis.Info)
                    merge(merge_dxf, base_dxf)

                base_dxf.saveas(gispath + "\\CAD_merged.dxf")

                layer = QgsProject.instance().mapLayersByName('output')[0]
                QgsProject.instance().removeMapLayer(layer.id())
                self.iface.messageBar().clearWidgets()
                resetLegend(self)

                try:
                    layer = QgsProject.instance().mapLayersByName('Intersection')[0]
                    QgsProject.instance().removeMapLayer(layer.id())
                    resetLegend(self)
                except Exception as e:
                    resetLegend(self)
                    pass

                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception as e:
                    pass
            else:
                self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                QgsMessageLog.logMessage('DEBUG: exportDXF cancelled.', 'sGIS', level=Qgis.Info)
                resetLegend(self)
                return

        except Exception as e:
            msg = QMessageBox()
            msg.setWindowTitle('CONTOURS UNAVAILABLE!')
            msg.setText('There are no contours available for this town.  Please try again without contours.')
            cont = msg.addButton('Continue', QMessageBox.AcceptRole)
            msg.setDefaultButton(cont)
            msg.exec_()
            msg.deleteLater()
            if msg.clickedButton() is cont:
                self.iface.actionSelect().trigger()
                layer = QgsProject.instance().mapLayersByName('output')[0]
                QgsProject.instance().removeMapLayer(layer.id())
                self.iface.messageBar().clearWidgets()
                QgsMessageLog.logMessage('DEBUG: NO CONTOURS - Exiting...', 'sGIS', level=Qgis.Info)
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception as e:
                    pass
                return
                
            else:
                return

        except Exception as e:
            layer = QgsProject.instance().mapLayersByName('output')[0]
            QgsProject.instance().removeMapLayer(layer.id())
            self.iface.messageBar().clearWidgets()
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))

    def bufferPOLY(self, sInputLayer):

        Processing.initialize()
        inputL = QgsProject.instance().mapLayersByName(sInputLayer)[0]

        self.fb = QgsProcessingFeedback()
        self.context = QgsProcessingContext
        self.outputLayer = processing.run("native:buffer", {
            'INPUT': sInputLayer,
            'DISSOLVE': False,
            'DISTANCE': 400,
            'END_CAP_STYLE': 2,
            'JOIN_STYLE': 0,
            'MITER_LIMIT': 2,
            'SEGMENTS': 10,
            'OUTPUT': 'memory:',
            'PREDICATE': [0]}, feedback=self.fb)['OUTPUT']

        try:
            QgsProject.instance().addMapLayer(self.outputLayer)
            self.iface.setActiveLayer(self.outputLayer)
            # self.selectLastFeature(self.outputLayer)
            layer = QgsProject.instance().mapLayersByName('tmp_buffer')[0]
            QgsProject.instance().removeMapLayer(layer.id())
            return 1

        except Exception as e:

            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))

    def exportTIFF(self, layer, filename):

        layerOutput = QgsProject.instance().mapLayersByName('output')[0]

        Processing.initialize()
        self.fb = QgsProcessingFeedback()
        self.context = QgsProcessingContext

        self.outputLayer = processing.run("gdal:cliprasterbymasklayer", {
            'INPUT': layer,
            'KEEP_RESOLUTION': False,
            'MASK': layerOutput,
            'NODATA': 0, 'OPTIONS': '',
            'OUTPUT': filename,}, feedback=self.fb)['OUTPUT']

        # add raster layer to canvas
        self.iface.addRasterLayer(filename, "tmp_slope")

    def exportSHP(self, layer, filename):

        extent = self.iface.mapCanvas().extent()
        context = QgsProject.instance().transformContext()

        options = QgsVectorFileWriter.SaveVectorOptions()
        options.driverName = "ESRI Shapefile"
        options.fileEncoding = "UTF-8"
        options.overrideGeometryType=QgsWkbTypes.MultiLineStringZ
        options.filterExtent = extent
        options.ct = QgsCoordinateTransform(layer.crs(), QgsCoordinateReferenceSystem.fromEpsgId(102684),
                                                 QgsProject.instance())
        QgsVectorFileWriter.writeAsVectorFormatV2(layer, filename, context, options)
        QgsProject.instance().removeMapLayer(layer.id())

    def exportGeoJSON(self, layer, filename):

        extent = self.iface.mapCanvas().extent()
        context = QgsProject.instance().transformContext()

        options = QgsVectorFileWriter.SaveVectorOptions()
        options.driverName = "GeoJSON"
        options.fileEncoding = "UTF-8"
        options.overrideGeometryType=QgsWkbTypes.LineStringZ
        options.filterExtent = extent
        options.skipAttributeCreation = True
        options.forceMulti = False
        options.includeZ = True

        options.ct = QgsCoordinateTransform(layer.crs(), QgsCoordinateReferenceSystem.fromEpsgId(102684),
                                                 QgsProject.instance())
        QgsVectorFileWriter.writeAsVectorFormatV2(layer, filename, context, options)
        QgsProject.instance().removeMapLayer(layer.id())

    def multi2single(self, layer):

        Processing.initialize()
        self.fb = QgsProcessingFeedback()
        self.context = QgsProcessingContext

        algSingle = processing.runAndLoadResults("qgis:multiparttosingleparts",
                                                 {'INPUT': layer, 'OUTPUT': 'TEMPORARY_OUTPUT'})


    def polygonize(self, layer, filename):

        Processing.initialize()
        self.fb = QgsProcessingFeedback()
        self.context = QgsProcessingContext

        self.outputLayer = processing.run("gdal:polygonize", {
            'BAND': 1, 'EIGHT_CONNECTEDNESS': False, 'FIELD': 'DN',
            'INPUT': layer,
            'OUTPUT': filename,}, feedback=self.fb)['OUTPUT']

        l = QgsVectorLayer(filename, "tmp_slopeSHP", "ogr")
        QgsProject.instance().addMapLayer(l)
        QgsProject.instance().removeMapLayer(layer.id())

    def intersectCONTOURS(self, town):

        Processing.initialize()
        layerOutput = QgsProject.instance().mapLayersByName('output')[0]
        self.fb = QgsProcessingFeedback()
        self.context = QgsProcessingContext
        if str(town) == 'Southport':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Southport')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Bath':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Bath')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Boothbay':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Boothbay')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Boothbay Harbor':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Boothbay-Harbor')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Bremen':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Bremen')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Bristol':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Bristol')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Damariscotta':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Damariscotta')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Edgecomb':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Edgecomb')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Friendship':
            self.vl = QgsProject.instance().mapLayersByName('3DContours_Friendship')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Georgetown':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Georgetown')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Harpswell':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Harpswell')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Jefferson':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Jefferson')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Newcastle':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Newcastle')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Nobleboro':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Nobleboro')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Phippsburg':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Phippsburg')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'South Bristol':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_South_Bristol')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Waldoboro':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Waldoboro')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'West Bath':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_West_Bath')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Westport Island':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Westport_Island')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Wiscasset':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Wiscasset')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Woolwich':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Woolwich')[0]
            self.iface.setActiveLayer(self.vl)
        else:
            return 'X'

        processing.runAndLoadResults("qgis:intersection",
                                     {'INPUT': self.vl,
                                      'OUTPUT': 'memory:tmp',
                                      'OVERLAY': layerOutput,
                                      }, feedback=self.fb)

    def intersectALL(self, inputLayer, town):

        Processing.initialize()
        self.fb = QgsProcessingFeedback()
        self.context = QgsProcessingContext
        self.vl = inputLayer
        outputLayer = QgsProject.instance().mapLayersByName('output')[0]

        if self.vl == QgsProject.instance().mapLayersByName('Parcels')[0]:
            self.vl.setSubsetString("town = '{}' ".format(town))

        if self.vl == QgsProject.instance().mapLayersByName('ng911rdss')[0]:
            self.vl.setSubsetString("TOWN = '{}' ".format(town))

        self.iface.setActiveLayer(self.vl)

        processing.runAndLoadResults("qgis:intersection",
                                     {'INPUT': self.vl,
                                      'OUTPUT': 'memory:tmp',
                                      'OVERLAY': outputLayer,
                                      }, feedback=self.fb)

        if self.vl == QgsProject.instance().mapLayersByName('Parcels')[0]:
            self.vl.setSubsetString("")

        if self.vl == QgsProject.instance().mapLayersByName('ng911rdss')[0]:
            self.vl.setSubsetString("")


class sgis_exportPotree(object):

    def __init__(self, iface, ujnV, pnV, fLV, fPV, fOV, oX, oY, oZ, oS, oR, iC):

        self.iface = iface
        self.useJobNo = ujnV
        self.pageName = pnV
        self.fileLAS = fLV
        self.filePLY = fPV
        self.fileOBJ = fOV
        self.objX = oX
        self.objY = oY
        self.objZ = oZ
        self.objScale = oS
        self.objRotation = oR
        self.includeContours = iC


    def initGui(self, ujnV, pnV, fLV, fPV, fOV, oX, oY, oZ, oS, oR, iC):

        self.run(ujnV, pnV, fLV, fPV, fOV, oX, oY, oZ, oS, oR, iC)

    def run(self, ujnV, pnV, fLV, fPV, fOV, oX, oY, oZ, oS, oR, iC):

        self.useJobNo = ujnV
        self.pageName = pnV
        self.fileLAS = fLV
        self.filePLY = fPV
        self.fileOBJ = fOV
        self.objX = oX
        self.objY = oY
        self.objZ = oZ
        self.objScale = oS
        self.objRotation = oR
        self.objParams = []
        self.includeContours = iC
        self.filePanos = 'YES'

        import math
        try:
            self.objRotation = float(self.objRotation)
            self.objRotation = self.objRotation *(math.pi/180)
        except Exception as e:
            self.objRotation = self.objRotation

        QgsMessageLog.logMessage('Starting PotreeConverter.EXE...', 'sGIS', level=Qgis.Info)
        self.iface.mapCanvas().selectionChanged.connect(self.select_changed)
        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(self.vl)
        aLayer = self.iface.activeLayer()
        feats_count = aLayer.selectedFeatureCount()

        if feats_count == 0:
            msg = QMessageBox()
            msg.setWindowTitle('Selection')
            msg.setText('Choose the correct job feature and click OK to proceed.')
            cont = msg.addButton('Continue', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(cont)
            msg.exec_()
            msg.deleteLater()
            if msg.clickedButton() is cont:
                self.iface.actionSelect().trigger()
            else:
                QgsMessageLog.logMessage('DEBUG: exportPotree cancelled.', 'sGIS', level=Qgis.Info)
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception as e:
                    pass
                return
        else:
            job = self.iface.activeLayer().selectedFeatures()[0]
            jobNo = job["job_no"]
            msg = QMessageBox()
            msg.setWindowTitle('SELECT Job')
            msg.setText(str(jobNo) + ' is selected. Do you want to export that selection or choose a new selection?')
            cont = msg.addButton('Continue', QMessageBox.AcceptRole)
            new = msg.addButton('New Selection', QMessageBox.AcceptRole)
            cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
            msg.setDefaultButton(cont)
            msg.exec_()
            msg.deleteLater()

            if msg.clickedButton() is cont:
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception as e:
                    pass
                self.select_changed()
            elif msg.clickedButton() is new:
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception as e:
                    pass

                msg = QMessageBox()
                msg.setWindowTitle('Selection')
                msg.setText('Choose the correct job feature and click OK to proceed.')
                cont = msg.addButton('Continue', QMessageBox.AcceptRole)
                cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
                msg.setDefaultButton(cont)
                msg.exec_()
                msg.deleteLater()
                if msg.clickedButton() is cont:
                    for a in self.iface.mainWindow().children():

                        if a.objectName() == 'mActionDeselectAll':
                            a.trigger()
                            QgsMessageLog.logMessage('FIRST RUN: Previous selection has been cleared.',
                                                     'sGIS', level=Qgis.Info)
                            self.iface.actionSelect().trigger()

                elif msg.clickedButton() is cancel:
                    try:
                        self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                    except Exception as e:
                        pass

            elif msg.clickedButton() is cancel:
                QgsMessageLog.logMessage('DEBUG: exportPotree cancelled.', 'sGIS', level=Qgis.Info)
                try:
                    self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                except Exception as e:
                    pass

        QgsMessageLog.logMessage('Potree Generation Complete.', 'sGIS', level=Qgis.Info)
        resetLegend(self)
        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(self.vl)

    def select_changed(self):

        try:
            self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
        except Exception as e:
            pass

        try:
            feat = self.iface.activeLayer().selectedFeatures()[0]
            jobNo = str(feat["job_no"])
            import datetime
            import subprocess
            year = datetime.datetime.today().strftime('%Y')
            jobYear = '20' + jobNo[:2]

            if jobYear == year:
                year = year
            else:
                year = jobYear

            if str(self.useJobNo) == 'True':
                self.pageName = jobNo
            else:
                self.pageName = self.pageName

            self.objParams.append(self.pageName)
            self.objParams.append(self.objX)
            self.objParams.append(self.objY)
            self.objParams.append(self.objZ)
            self.objParams.append(self.objScale)
            self.objParams.append(self.objRotation)
            QgsMessageLog.logMessage('.OBJ Params: ' + str(self.objParams), 'sGIS', level=Qgis.Info)

            path = os.path.join(jobsPath, year, jobNo)  # need to programattically grab year
            jipath = os.path.join(path, "Job_Info")
            supath = os.path.join(path, "survey")
            panopath = supath + "\\360Pano\\"
            jobFolders(jobNo)
            town = str(feat["town"])

            self.strObjParams = str(self.objParams)
            self.strObjParams = self.strObjParams.replace('\'', '')
            self.strObjParams = self.strObjParams.replace(' ', '')
            self.strObjParams = self.strObjParams.replace('[', '')
            self.strObjParams = self.strObjParams.replace(']', '')

            QgsMessageLog.logMessage('fileLAS: ' + str(self.fileLAS), 'sGIS', level=Qgis.Info)
            command = ["c:\\program files\\qgis 3.40.1\\bin\\converter\\PotreeConverter.exe", self.fileLAS, "-o", supath, "-p",
                       self.strObjParams,
                       "--overwrite"]
            subprocess.check_call(command, stderr=subprocess.STDOUT, shell=False)

            if str(self.filePanos) == '':
                QgsMessageLog.logMessage('EXCLUDE PANOS', 'sGIS', level=Qgis.Info)
            else:
                if os.path.exists(supath + "\\pointclouds\\" + self.pageName + "\\360Pano"):
                    shutil.rmtree(supath + "\\pointclouds\\" + self.pageName + "\\360Pano")
                    shutil.copytree(panopath, supath + "\\pointclouds\\" + self.pageName + "\\360Pano")
                else:
                    if not os.path.exists(panopath):
                        os.makedirs(panopath)
                        shutil.copytree(panopath, supath + "\\pointclouds\\" + self.pageName + "\\360Pano")
                    else:
                        shutil.copytree(panopath, supath + "\\pointclouds\\" + self.pageName + "\\360Pano")
    
            if str(self.filePLY) == '':
                QgsMessageLog.logMessage('EXCLUDE .PLY', 'sGIS', level=Qgis.Info)
            else:
                QgsMessageLog.logMessage('filePLY: ' + str(self.filePLY), 'sGIS', level=Qgis.Info)
                shutil.copyfile(self.filePLY, supath + "\\pointclouds\\" + self.pageName + "\\" + self.pageName + ".ply")

            if str(self.fileOBJ) == '':
                QgsMessageLog.logMessage('EXCLUDE .OBJ', 'sGIS', level=Qgis.Info)
            else:
                QgsMessageLog.logMessage('fileOBJ: ' + str(self.fileOBJ), 'sGIS', level=Qgis.Info)
                shutil.copyfile(self.fileOBJ, supath + "\\pointclouds\\" + self.pageName + "\\" + self.pageName + ".obj")

            if str(self.includeContours) == 'NO':
                QgsMessageLog.logMessage('EXCLUDE CONTOURS', 'sGIS', level=Qgis.Info)
            else:
                self.fileGeoJSON = gispath + "\\contours.geojson"
                QgsMessageLog.logMessage('fileGeoJSON: ' + str(self.fileGeoJSON), 'sGIS', level=Qgis.Info)
                shutil.copyfile(self.fileGeoJSON, supath + "\\pointclouds\\" + self.pageName + "\\" + self.pageName + ".geojson")

            self.pcFolder = supath + "\\pointclouds\\" + self.pageName
            self.htmlFile = supath + "\\" + self.pageName + ".html"
            self.wwwRootFolder = wwwRootFolder
            self.wwwPCFolder = self.wwwRootFolder + '\\pointclouds\\' + self.pageName
            self.wwwURL = wwwURL + self.pageName + '.html'
            QgsMessageLog.logMessage('wwwRootFolder: ' + str(self.wwwRootFolder), 'sGIS', level=Qgis.Info)
            QgsMessageLog.logMessage('url: ' + str(self.wwwURL), 'sGIS', level=Qgis.Info)

            from distutils import file_util, dir_util

            distutils.file_util.copy_file(self.htmlFile, self.wwwRootFolder)
            # distutils.dir_util.copy_tree(self.pcFolder, self.wwwPCFolder)
            
            if not os.path.exists(self.wwwPCFolder):
                os.makedirs(self.wwwPCFolder)
                os.makedirs(self.wwwPCFolder + "\\360Pano")
                distutils.dir_util.copy_tree(self.pcFolder, self.wwwPCFolder)

            else:
              if not os.path.exists(self.wwwPCFolder + "\\360Pano"):
                    os.makedirs(self.wwwPCFolder + "\\360Pano")
                    distutils.dir_util.copy_tree(self.pcFolder, self.wwwPCFolder)
              else:
                  distutils.dir_util.copy_tree(self.pcFolder, self.wwwPCFolder)            

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))


class sgis_CADexportOrig_dialog(QDialog, Ui_sgis_cadOutputs_form):

    def __init__(self, iface):
        QDialog.__init__(self)
        self.iface = iface
        self.setupUi(self)
        buttonBox = self.buttonBox
        buttonBox.accepted.connect(partial(self.launch_form))

    def launch_form(self):

        contours = self.findChild(QCheckBox, "cb_contours")
        parcels = self.findChild(QCheckBox, "cb_parcels")
        ae = self.findChild(QCheckBox, "cb_ae")
        ve = self.findChild(QCheckBox, "cb_ve")
        hat2015 = self.findChild(QCheckBox, "cb_hat2015")
        slope = self.findChild(QCheckBox, "cb_slope")
        rds = self.findChild(QCheckBox, "cb_rdss")

        cValue = contours.checkState()
        pValue = parcels.checkState()
        aeValue = ae.checkState()
        veValue = ve.checkState()
        hatValue = hat2015.checkState()
        slopeValue = slope.checkState()
        rdsValue = rds.checkState()

        self.eDXF_dialog = sgis_exportDXF(self.iface, cValue, pValue, aeValue, veValue, hatValue, slopeValue, rdsValue)
        self.eDXF_dialog.initGui(cValue, pValue, aeValue, veValue, hatValue, slopeValue, rdsValue)


class sgis_CADexportNEW_dialog(QDialog, Ui_sgis_cadOutputs_form):

    def __init__(self, iface):
        QDialog.__init__(self)
        self.iface = iface
        self.setupUi(self)
        buttonBox = self.buttonBox
        buttonBox.accepted.connect(partial(self.launch_form))

    def launch_form(self):

        contours = self.findChild(QCheckBox, "cb_contours")
        parcels = self.findChild(QCheckBox, "cb_parcels")
        ae = self.findChild(QCheckBox, "cb_ae")
        ve = self.findChild(QCheckBox, "cb_ve")
        hat2015 = self.findChild(QCheckBox, "cb_hat2015")
        slope = self.findChild(QCheckBox, "cb_slope")
        rds = self.findChild(QCheckBox, "cb_rdss")

        cValue = contours.checkState()
        pValue = parcels.checkState()
        aeValue = ae.checkState()
        veValue = ve.checkState()
        hatValue = hat2015.checkState()
        slopeValue = slope.checkState()
        rdsValue = rds.checkState()

        self.eCAD_dialog = sgis_exportCAD(self.iface, cValue, pValue, aeValue, veValue, hatValue, slopeValue, rdsValue)
        self.eCAD_dialog.initGui(cValue, pValue, aeValue, veValue, hatValue, slopeValue, rdsValue)
        

class sgis_potreeExport_dialog(QDialog, Ui_PotreeDialog):

    def __init__(self, iface):
        QDialog.__init__(self)
        self.iface = iface
        self.setupUi(self)
        buttonBox = self.buttonBox
        buttonBox.accepted.connect(partial(self.launch_form))

    def launch_form(self):

        useJobNo = self.findChild(QRadioButton, "btnUseJobNo")
        useJobNoValue = useJobNo.isChecked()
        pageName = self.findChild(QLineEdit, "pageName")
        pageNameValue = pageName.text()
        fileLAS = self.findChild(QgsFileWidget, "fileLAS")
        fileLASValue = fileLAS.filePath()
        filePLY = self.findChild(QgsFileWidget, "filePLY")
        filePLYValue = filePLY.filePath()
        fileOBJ = self.findChild(QgsFileWidget, "fileOBJ")
        fileOBJValue = fileOBJ.filePath()
        objX = self.findChild(QLineEdit, "objX")
        objXValue = objX.text()
        objY = self.findChild(QLineEdit, "objY")
        objYValue = objY.text()
        objZ = self.findChild(QLineEdit, "objZ")
        objZValue = objZ.text()
        objScale = self.findChild(QLineEdit, "objScale")
        objScaleValue = objScale.text()
        objRotation = self.findChild(QLineEdit, "objRotation")
        objRotationValue = objRotation.text()
        includeContours = self.findChild(QComboBox, "includeContours")
        icValue = includeContours.currentText()

        self.ePotree_dialog = sgis_exportPotree(self.iface, useJobNoValue, pageNameValue, fileLASValue, filePLYValue,
                                                  fileOBJValue, objXValue, objYValue, objZValue, objScaleValue,
                                                  objRotationValue, icValue)
        self.ePotree_dialog.initGui(useJobNoValue, pageNameValue, fileLASValue, filePLYValue, fileOBJValue, objXValue,
                                    objYValue, objZValue, objScaleValue, objRotationValue, icValue)


class sgis_addParcel(threading.Thread):

    def __init__(self, function_that_downloads):
        threading.Thread.__init__(self)
        self.runnable = function_that_downloads

    def run(self):
        self.runnable()


class sgis_supp_dialog(QDialog, Ui_sgis_supp_pre_form):

    def __init__(self, iface, pType):
        QDialog.__init__(self)
        self.iface = iface
        self.setupUi(self)
        buttonBox = self.buttonBox
        buttonBox.accepted.connect(partial(self.launch_form, pType))
        buttonBox.rejected.connect(self.finished)

    def launch_form(self, pType):

        supp_type = self.findChild(QComboBox, "supp_type")

        sType = str(supp_type.currentText())

        self.newLPJob_dialog = sgis_newLPJob(self.iface, sType, pType)
        self.newLPJob_dialog.initGui(sType, pType)


    def finished(self, **kwargs):
        self.done(1)
        

class sgis_exportCAD(object):

    def __init__(self, iface, cV, pV, aeV, veV, hV, sV, rdsV):

        self.iface = iface
        self.cV = cV
        self.pV = pV
        self.aeV = aeV
        self.veV = veV
        self.hV = hV
        self.sV = sV
        self.rdsV = rdsV

    def initGui(self, cV, pV, aeV, veV, hV, sV, rdsV):

        self.run(cV, pV, aeV, veV, hV, sV, rdsV)

    def run(self, cV, pV, aeV, veV, hV, sV, rdsV):

        self.cV = cV
        self.pV = pV
        self.aeV = aeV
        self.veV = veV
        self.hV = hV
        self.sV = sV
        self.rdsV = rdsV

        self.iface.mapCanvas().selectionChanged.connect(self.select_changed)
        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(self.vl)
        aLayer = self.iface.activeLayer()
        feats_count = aLayer.selectedFeatureCount()

        msg = QMessageBox()
        msg.setWindowTitle('User Export')
        msg.setText('Extents will be used for intersections.  Please ensure you have a job selected before continuing.  Click OK to proceed.')
        cont = msg.addButton('Continue', QMessageBox.AcceptRole)
        cancel = msg.addButton('Cancel', QMessageBox.RejectRole)
        msg.setDefaultButton(cont)
        msg.exec_()
        msg.deleteLater()
        
        if msg.clickedButton() is cont:
        
          try:
            feat = self.iface.activeLayer().selectedFeatures()[0]
          except Exception as e:
            self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "No Selection",
                                 "Please ensure that you have a job feature selected\nand attempt to "
                                 "generate the outputs again.\n\n"
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))
            return
                   
          jobNo = str(feat["job_no"])
          town = str(feat["town"])
          self.iface.actionSelect().trigger()
          p = QgsProject.instance()
          ext = self.iface.mapCanvas().extent()
          ext_g = QgsGeometry.fromRect(ext)
          QgsMessageLog.logMessage('Using extents to create mask...', 'sGIS', level=Qgis.Info)

          mask = QgsVectorLayer(f'?query=SELECT SetSRID(ST_GeomFromText(\'{ext_g.asWkt()}\'), {int(p.crs().authid().replace("ESRI:",""))})', 'ExtentPolyTemp', 'virtual')
          QgsProject.instance().addMapLayer(mask)

          sym = QgsSymbol.defaultSymbol(mask.geometryType())
          sym.setOpacity(1)
          s = sym.symbolLayers()[0]
          s.setColor(QColor("transparent"))
          s.setStrokeColor(QColor('fuchsia'))
          # s.setStrokeWidth(1)

          mask.setRenderer( QgsSingleSymbolRenderer( sym ) )
          mask.updateExtents()
          mask.triggerRepaint()
          
          self.vl = QgsProject.instance().mapLayersByName('ExtentPolyTemp')[0]
          self.iface.setActiveLayer(self.vl)
          self.vl.selectAll()
          feat = self.iface.activeLayer().selectedFeatures()[0]
          self.select_changed(jobNo, town)
          
        else:
            QgsMessageLog.logMessage('DEBUG: CAD Export (New) cancelled.', 'sGIS', level=Qgis.Info)
            try:
                self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
            except Exception as e:
                pass
            return
        
        resetLegend(self)
        
        cleanupCAD.run(self)
        QgsMessageLog.logMessage('DXF Generation Complete, TempFiles Deleted...', 'sGIS', level=Qgis.Info)

    def select_changed(self, jobNo, town):

        try:
            self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
        except Exception as e:
            pass

        import datetime
        year = datetime.datetime.today().strftime('%Y')
        jobYear = '20' + jobNo[:2]

        if jobYear == year:
            year = year
        else:
            year = jobYear

        path = os.path.join(jobsPath, year, jobNo)  # need to programattically grab year
        jobFolders(jobNo)
        gispath = os.path.join(path, "GIS")

        self.iface.actionCopyFeatures().trigger()
        self.fields = self.iface.activeLayer().fields()

        self.tmpLayer = self.vl
        self.iface.actionIdentify().trigger()

        self.bufferPOLY('ExtentPolyTemp')

        contoursSHP = gispath + "\\contours.shp"
        contoursDXF = gispath + "\\contours.dxf"
        contoursGeoJSON = gispath + "\\contours.geojson"
        aZoneSHP = gispath + "\\aZone.shp"
        aZoneDXF = gispath + "\\aZone.dxf"
        veZoneSHP = gispath + "\\veZone.shp"
        veZoneDXF = gispath + "\\veZone.dxf"
        parcelsSHP = gispath + "\\parcels.shp"
        parcelsDXF = gispath + "\\parcels.dxf"
        hat2015SHP = gispath + "\\hat2015.shp"
        hat2015DXF = gispath + "\\hat2015.dxf"
        slopeTIFF = gispath + "\\midcoastSlope.tiff"
        slopeSHP = gispath + "\\midcoastSlope.shp"
        slopeSHPfinal = gispath + "\\midcoastSlopeFINAL.shp"
        slopeDXF = gispath + "\\midcoastSlope.dxf"
        rdsSHP = gispath + "\\rds.shp"
        rdsDXF = gispath + "\\rds.dxf"

        aZoneLayer = QgsProject.instance().mapLayersByName('S_Fld_Haz_Ar')[0]
        veZoneLayer = QgsProject.instance().mapLayersByName('S_Fld_Haz_Ar VE')[0]
        parcelsLayer = QgsProject.instance().mapLayersByName('Parcels')[0]
        hat2015Layer = QgsProject.instance().mapLayersByName('HAT_2015')[0]
        rdsLayer = QgsProject.instance().mapLayersByName('ng911rdss')[0]
        self.slopeLayer = \
        QgsProject.instance().mapLayersByName('MidcoastSlopeGreaterThan20PctGreaterThan2AcMask')[0]

        self.layerOutput = QgsProject.instance().mapLayersByName('output')[0]

        import subprocess

        filenames = []

        # NEED CLEAN EXIT IF UNSUPPORTED TOWN.
        if self.cV == 2:
            self.intersectCONTOURS(town)
            try:
              self.iLayer = QgsProject.instance().mapLayersByName('Intersection')[0]
              self.multi2single(self.iLayer)
              self.sLayer = QgsProject.instance().mapLayersByName('Single parts')[0]
              self.exportGeoJSON(self.sLayer, contoursGeoJSON)
              self.exportSHP(self.iLayer, contoursSHP)
              command = ["ogr2ogr", "-skipfailures", "-f", "DXF", "-zField", """Elev_FT""", contoursDXF,
                         contoursSHP]
              subprocess.check_call(command, stderr=subprocess.DEVNULL, stdout=subprocess.DEVNULL)
              filenames.append(gispath + "\\contours.dxf")
            except Exception as e:
                # QgsMessageLog.logMessage('EXCEPTION: ' + str(e), 'sGIS', level=Qgis.Info)
                msg = QMessageBox()
                msg.setWindowTitle('CONTOURS UNAVAILABLE!')
                msg.setText('There are no contours available for this town.  Please try again without contours.')
                cont = msg.addButton('Continue', QMessageBox.AcceptRole)
                msg.setDefaultButton(cont)
                msg.exec_()
                msg.deleteLater()
                if msg.clickedButton() is cont:
                    self.iface.actionSelect().trigger()
                    layer = QgsProject.instance().mapLayersByName('output')[0]
                    QgsProject.instance().removeMapLayer(layer.id())
                    self.iface.messageBar().clearWidgets()
                    QgsMessageLog.logMessage('DEBUG: NO CONTOURS - Exiting...', 'sGIS', level=Qgis.Info)
                    try:
                        self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
                    except Exception as e:
                        pass
                    return
                    
                else:
                    return

        if self.aeV == 2:
            self.intersectALL(aZoneLayer, town)
            self.iLayer = QgsProject.instance().mapLayersByName('Intersection')[0]
            self.exportSHP(self.iLayer, aZoneSHP)
            command = ["ogr2ogr", "-skipfailures", "-f", "DXF", aZoneDXF, aZoneSHP]
            subprocess.check_call(command, stderr=subprocess.DEVNULL, stdout=subprocess.DEVNULL)
            filenames.append(gispath + "\\aZone.dxf")
        if self.veV == 2:
            self.intersectALL(veZoneLayer, town)
            self.iLayer = QgsProject.instance().mapLayersByName('Intersection')[0]
            self.exportSHP(self.iLayer, veZoneSHP)
            command = ["ogr2ogr", "-skipfailures", "-f", "DXF", veZoneDXF, veZoneSHP]
            subprocess.check_call(command, stderr=subprocess.DEVNULL, stdout=subprocess.DEVNULL)
            filenames.append(gispath + "\\veZone.dxf")
        if self.rdsV == 2:
            self.intersectALL(rdsLayer, town)
            self.iLayer = QgsProject.instance().mapLayersByName('Intersection')[0]
            self.exportSHP(self.iLayer, rdsSHP)
            command = ["ogr2ogr", "-skipfailures", "-f", "DXF", rdsDXF, rdsSHP]
            subprocess.check_call(command, stderr=subprocess.DEVNULL, stdout=subprocess.DEVNULL)
            filenames.append(gispath + "\\rds.dxf")
        if self.pV == 2:
            self.intersectALL(parcelsLayer, town)
            self.iLayer = QgsProject.instance().mapLayersByName('Intersection')[0]
            self.exportSHP(self.iLayer, parcelsSHP)
            command = ["ogr2ogr", "-skipfailures", "-f", "DXF", parcelsDXF, parcelsSHP]
            subprocess.check_call(command, stderr=subprocess.DEVNULL, stdout=subprocess.DEVNULL)
            filenames.append(gispath + "\\parcels.dxf")
        if self.hV == 2:
            self.intersectALL(hat2015Layer, town)
            self.iLayer = QgsProject.instance().mapLayersByName('Intersection')[0]
            self.exportSHP(self.iLayer, hat2015SHP)
            command = ["ogr2ogr", "-skipfailures", "-f", "DXF", hat2015DXF, hat2015SHP]
            subprocess.check_call(command, stderr=subprocess.DEVNULL, stdout=subprocess.DEVNULL)
            filenames.append(gispath + "\\hat2015.dxf")
        if self.sV == 2:
            self.exportTIFF(self.slopeLayer, slopeTIFF)
            layerSlope = QgsProject.instance().mapLayersByName('tmp_slope')[0]
            self.polygonize(layerSlope, slopeSHP)
            layerSlopeSHP = QgsProject.instance().mapLayersByName('tmp_slopeSHP')[0]
            dataProvider = layerSlopeSHP.dataProvider()
            dataProvider.addAttributes([QgsField("Layer", QVariant.String)])

            for f in layerSlopeSHP.getFeatures():
                self.iface.actionToggleEditing().trigger()
                idx = dataProvider.fieldNameIndex('Layer')
                layerSlopeSHP.changeAttributeValue(f.id(), idx, 'GOV-GIS-SLOPE')
                layerSlopeSHP.updateFields()
                self.iface.activeLayer().commitChanges()

            self.tsLayer = QgsProject.instance().mapLayersByName('tmp_slopeSHP')[0]
            self.exportSHP(self.tsLayer, slopeSHPfinal)
            command = ["ogr2ogr", "-skipfailures", "-f", "DXF", slopeDXF, slopeSHPfinal]
            subprocess.check_call(command, stderr=subprocess.DEVNULL, stdout=subprocess.DEVNULL)
            filenames.append(gispath + "\\midcoastSlope.dxf")

        def merge(source, target):
            importer = Importer(source, target)
            importer.import_modelspace()
            importer.finalize()

        base_dxf = ezdxf.readfile(filenames[0])

        for filename in filenames:
            merge_dxf = ezdxf.readfile(filename)
            QgsMessageLog.logMessage('filename: ' + '"' + str(filename) + '"', 'sGIS', level=Qgis.Info)
            merge(merge_dxf, base_dxf)

        base_dxf.saveas(gispath + "\\GIS_UserExport.dxf")
        
        path = os.path.join(jobsPath, year, jobNo)  # need to programattically grab year
        gispath = os.path.join(path, "GIS")
        
        from pathlib import Path
        import time
        keep = ['GIS_UserExport.dxf','CAD_merged.dxf','contours.geojson']
        root = Path(gispath)
        
        for path in root.iterdir():
        
            try:
                if path.name not in keep:
                    path.unlink()
            except Exception as e:
        
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                QgsMessageLog.logMessage("File Deletion Error - Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(exc_tb.tb_lineno) + ' ' + str(e), 'sGIS', level=Qgis.Info)
            
        layer = QgsProject.instance().mapLayersByName('output')[0]
        QgsProject.instance().removeMapLayer(layer.id())
        self.iface.messageBar().clearWidgets()
        resetLegend(self)

        try:
            layer = QgsProject.instance().mapLayersByName('Intersection')[0]
            QgsProject.instance().removeMapLayer(layer.id())
            resetLegend(self)
        except Exception as e:
            resetLegend(self)
            pass

        try:
            self.iface.mapCanvas().selectionChanged.disconnect(self.select_changed)
        except Exception as e:
            pass
        resetLegend(self)
       
    def bufferPOLY(self, sInputLayer):

        Processing.initialize()
        inputL = QgsProject.instance().mapLayersByName(sInputLayer)[0]

        self.fb = QgsProcessingFeedback()
        self.context = QgsProcessingContext
        # QgsMessageLog.logMessage(sInputLayer, 'sGIS', level=Qgis.Info)
        self.outputLayer = processing.run("native:buffer", {
            'INPUT': sInputLayer,
            'DISSOLVE': False,
            'DISTANCE': 400,
            'END_CAP_STYLE': 2,
            'JOIN_STYLE': 0,
            'MITER_LIMIT': 2,
            'SEGMENTS': 10,
            'OUTPUT': 'memory:',
            'PREDICATE': [0]}, feedback=self.fb)['OUTPUT']

        try:
            QgsProject.instance().addMapLayer(self.outputLayer)
            self.iface.setActiveLayer(self.outputLayer)
            layer = QgsProject.instance().mapLayersByName('ExtentPolyTemp')[0]
            QgsProject.instance().removeMapLayer(layer.id())
            return 1

        except Exception as e:

            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))

    def exportTIFF(self, layer, filename):

        layerOutput = QgsProject.instance().mapLayersByName('output')[0]

        Processing.initialize()
        self.fb = QgsProcessingFeedback()
        self.context = QgsProcessingContext

        self.outputLayer = processing.run("gdal:cliprasterbymasklayer", {
            'INPUT': layer,
            'KEEP_RESOLUTION': False,
            'MASK': layerOutput,
            'NODATA': 0, 'OPTIONS': '',
            'OUTPUT': filename,}, feedback=self.fb)['OUTPUT']

        # add raster layer to canvas
        self.iface.addRasterLayer(filename, "tmp_slope")

    def exportSHP(self, layer, filename):

        extent = self.iface.mapCanvas().extent()
        context = QgsProject.instance().transformContext()

        options = QgsVectorFileWriter.SaveVectorOptions()
        options.driverName = "ESRI Shapefile"
        options.fileEncoding = "UTF-8"
        options.overrideGeometryType=QgsWkbTypes.MultiLineStringZ
        options.filterExtent = extent
        options.ct = QgsCoordinateTransform(layer.crs(), QgsCoordinateReferenceSystem.fromEpsgId(102684),
                                                 QgsProject.instance())
        QgsVectorFileWriter.writeAsVectorFormatV2(layer, filename, context, options)
        QgsProject.instance().removeMapLayer(layer.id())

    def exportGeoJSON(self, layer, filename):

        extent = self.iface.mapCanvas().extent()
        context = QgsProject.instance().transformContext()

        options = QgsVectorFileWriter.SaveVectorOptions()
        options.driverName = "GeoJSON"
        options.fileEncoding = "UTF-8"
        options.overrideGeometryType=QgsWkbTypes.LineStringZ
        options.filterExtent = extent
        options.skipAttributeCreation = True
        options.forceMulti = False
        options.includeZ = True

        options.ct = QgsCoordinateTransform(layer.crs(), QgsCoordinateReferenceSystem.fromEpsgId(102684),
                                                 QgsProject.instance())
        QgsVectorFileWriter.writeAsVectorFormatV2(layer, filename, context, options)
        QgsProject.instance().removeMapLayer(layer.id())

    def multi2single(self, layer):

        Processing.initialize()
        self.fb = QgsProcessingFeedback()
        self.context = QgsProcessingContext

        algSingle = processing.runAndLoadResults("qgis:multiparttosingleparts",
                                                 {'INPUT': layer, 'OUTPUT': 'TEMPORARY_OUTPUT'})


    def polygonize(self, layer, filename):

        Processing.initialize()
        self.fb = QgsProcessingFeedback()
        self.context = QgsProcessingContext

        self.outputLayer = processing.run("gdal:polygonize", {
            'BAND': 1, 'EIGHT_CONNECTEDNESS': False, 'FIELD': 'DN',
            'INPUT': layer,
            'OUTPUT': filename,}, feedback=self.fb)['OUTPUT']

        l = QgsVectorLayer(filename, "tmp_slopeSHP", "ogr")
        QgsProject.instance().addMapLayer(l)
        QgsProject.instance().removeMapLayer(layer.id())

    def intersectCONTOURS(self, town):

        Processing.initialize()
        layerOutput = QgsProject.instance().mapLayersByName('output')[0]
        self.fb = QgsProcessingFeedback()
        self.context = QgsProcessingContext
        if str(town) == 'Southport':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Southport')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Bath':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Bath')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Boothbay':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Boothbay')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Boothbay Harbor':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Boothbay-Harbor')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Bremen':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Bremen')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Bristol':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Bristol')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Damariscotta':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Damariscotta')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Edgecomb':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Edgecomb')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Friendship':
            self.vl = QgsProject.instance().mapLayersByName('3DContours_Friendship')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Georgetown':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Georgetown')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Harpswell':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Harpswell')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Jefferson':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Jefferson')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Newcastle':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Newcastle')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Nobleboro':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Nobleboro')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Phippsburg':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Phippsburg')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'South Bristol':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_South_Bristol')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Waldoboro':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Waldoboro')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'West Bath':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_West_Bath')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Westport Island':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Westport_Island')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Wiscasset':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Wiscasset')[0]
            self.iface.setActiveLayer(self.vl)
        elif str(town) == 'Woolwich':
            self.vl = QgsProject.instance().mapLayersByName('3DContours2ft_Woolwich')[0]
            self.iface.setActiveLayer(self.vl)
        else:
            return 'X'

        processing.runAndLoadResults("qgis:intersection",
                                     {'INPUT': self.vl,
                                      'OUTPUT': 'memory:tmp',
                                      'OVERLAY': layerOutput,
                                      }, feedback=self.fb)

    def intersectALL(self, inputLayer, town):

        Processing.initialize()
        self.fb = QgsProcessingFeedback()
        self.context = QgsProcessingContext
        self.vl = inputLayer
        outputLayer = QgsProject.instance().mapLayersByName('output')[0]

        if self.vl == QgsProject.instance().mapLayersByName('Parcels')[0]:
            self.vl.setSubsetString("town = '{}' ".format(town))

        if self.vl == QgsProject.instance().mapLayersByName('ng911rdss')[0]:
            self.vl.setSubsetString("TOWN = '{}' ".format(town))

        self.iface.setActiveLayer(self.vl)

        processing.runAndLoadResults("qgis:intersection",
                                     {'INPUT': self.vl,
                                      'OUTPUT': 'memory:tmp',
                                      'OVERLAY': outputLayer,
                                      }, feedback=self.fb)

        if self.vl == QgsProject.instance().mapLayersByName('Parcels')[0]:
            self.vl.setSubsetString("")

        if self.vl == QgsProject.instance().mapLayersByName('ng911rdss')[0]:
            self.vl.setSubsetString("")


class cleanupCAD(object):

    def __init__(self, iface):
        self.iface = iface

    def run(self):
        QgsMessageLog.logMessage('Attempting cleanup...', 'sGIS', level=Qgis.Info)
        self.vl = QgsProject.instance().mapLayersByName('jobs')[0]
        self.iface.setActiveLayer(self.vl)
        try:
          feat = self.iface.activeLayer().selectedFeatures()[0]
        except Exception as e:
          exc_type, exc_obj, exc_tb = sys.exc_info()
          fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
          QMessageBox.critical(self.iface.mainWindow(), "No Selection",
                               "Please ensure that you have a job feature selected\nand attempt to "
                               "generate the outputs again.\n\n"
                               "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                   exc_tb.tb_lineno) + ' ' + str(e))
          return
                   
        jobNo = str(feat["job_no"])

        import datetime
        year = datetime.datetime.today().strftime('%Y')     
        jobYear = '20' + jobNo[:2]
        if jobYear == year:
            year = year
        else:
            year = jobYear   
        path = os.path.join(jobsPath, jobYear, jobNo)  # need to programattically grab year
        gispath = os.path.join(path, "GIS")
        
        from pathlib import Path
        import time
        keep = ['GIS_UserExport.dxf','CAD_merged.dxf','contours.geojson']
        root = Path(gispath)
        
        for path in root.iterdir():
        
            try:
                if path.name not in keep:
                    path.unlink()
            except Exception as e:
        
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                QgsMessageLog.logMessage("File Deletion Error - Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(exc_tb.tb_lineno) + ' ' + str(e), 'sGIS', level=Qgis.Info)


class sgis_allOutputs(object):

    def __init__(self, iface):
        self.iface = iface

    def run(self):
        QgsMessageLog.logMessage('Generating Folder Face Label...', 'sGIS', level=Qgis.Info)
        self.folderFace_dialog = sgis_printFolderLabel(self.iface)
        self.folderFace_dialog.run()
        QgsMessageLog.logMessage('DONE!', 'sGIS', level=Qgis.Info)
        
        QgsMessageLog.logMessage('Generating Yellow Sheet...', 'sGIS', level=Qgis.Info)
        self.yellowSheet_dialog = sgis_printYellowSheet(self.iface)
        self.yellowSheet_dialog.run()
        QgsMessageLog.logMessage('DONE!', 'sGIS', level=Qgis.Info)
        
        QgsMessageLog.logMessage('Generating MapTable...', 'sGIS', level=Qgis.Info)
        self.mapTable_dialog = sgis_printMapTable(self.iface)
        self.mapTable_dialog.run()
        QgsMessageLog.logMessage('DONE!', 'sGIS', level=Qgis.Info)

        QgsMessageLog.logMessage('Generating Contacts...', 'sGIS', level=Qgis.Info)
        self.contacts_dialog = sgis_printContacts(self.iface)
        self.contacts_dialog.run()
        QgsMessageLog.logMessage('DONE!', 'sGIS', level=Qgis.Info)

        QgsMessageLog.logMessage('Generating MapView...', 'sGIS', level=Qgis.Info)
        self.mapView_dialog = sgis_printMapView(self.iface)
        self.mapView_dialog.run()
        QgsMessageLog.logMessage('DONE!', 'sGIS', level=Qgis.Info)
                
        QgsMessageLog.logMessage('Generating SiteMap...', 'sGIS', level=Qgis.Info)
        self.siteMap_dialog = sgis_printSiteMap(self.iface)
        self.siteMap_dialog.run()
        QgsMessageLog.logMessage('DONE!', 'sGIS', level=Qgis.Info)

